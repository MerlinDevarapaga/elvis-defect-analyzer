"""
Generate AI-inferred RCA and Countermeasures for Elvis defects — Batch 2 (21 tickets).
Reads all_defects_batch2.json, writes output to Excel matching 'Till 17th March' template.
"""
import json
import sys
import io
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8', errors='replace')

EXCEL_OUT = r"C:\Users\mdevarapaga\Downloads\MSIL_DA2.8_Defect_List_6-Apr-2026_RCA_Output.xlsx"

# ── AI-generated RCA and Countermeasures per ticket ──────────────────────────
# Batch 1 (15 tickets from previous run)
RCA_BATCH1 = {
    "3727322": {
        "rca_identified": "No",
        "fix_in_progress": "Yes",
        "root_cause": (
            "Probable RCA: BT discovery asymmetry — HU's Bluetooth stack (likely BlueZ/Fluoride) "
            "is failing to respond to inquiry scan from certain phone models during device discovery, "
            "while the HU itself is broadcasting and visible to phones. "
            "Likely cause: Incompatible BT inquiry mode (standard vs interlaced) or BT controller "
            "firmware not properly handling page scan parameters for specific BT chipsets."
        ),
        "counter_measure": (
            "1. Collect BT HCI snoop logs from HU during failed discovery.\n"
            "2. Verify inquiry scan type (standard vs interlaced) and page scan interval config.\n"
            "3. Check BT firmware version against known compatibility matrix.\n"
            "4. Test with BT controller parameter tuning (inquiry scan window/interval).\n"
            "5. If firmware gap — update BT FW; if stack issue — patch Fluoride inquiry handler."
        ),
        "closure": "Post BT log analysis — target R9 (2026-04-17)",
    },
    "3725174": {
        "rca_identified": "No",
        "fix_in_progress": "No",
        "root_cause": (
            "Probable RCA: BT pairing state machine race condition — When 2nd phone initiates pairing "
            "while HFP/A2DP profiles are active on 1st phone, DCSM fails to register 2nd device in "
            "paired device DB. Pairing key exchange succeeds at BT stack level but device entry not "
            "propagated to HMI paired-device list. Missing callback in DCSM multi-device handler."
        ),
        "counter_measure": (
            "1. Review DCSM bonding callback flow for multi-device pairing.\n"
            "2. Add logging at DCSM layer to trace device-add events after bond completion.\n"
            "3. Verify BT stack bond_state_changed callback reaches DCSM for 2nd device.\n"
            "4. Fix: Ensure DCSM registers paired device in DB regardless of active profile count.\n"
            "5. Validate with 3+ phone pairing scenario."
        ),
        "closure": "Pending DCSM analysis — TBD",
    },
    "3725645": {
        "rca_identified": "Yes",
        "fix_in_progress": "Yes",
        "root_cause": (
            "Confirmed RCA: EVSHAL memory leak in Qualcomm camera HAL. Design flaw — mDirectRendering "
            "flag never becomes true since importExternalBuffers not called in MSIL use case, causing "
            "buffer handles to be allocated but never released. Continuous leak leads to SOC memory "
            "exhaustion, causing HU boot failure."
        ),
        "counter_measure": (
            "1. Fix applied: Release buffer handles when mDirectRendering is false (Gerrit 611568, 610972).\n"
            "2. Patch in packages/services/Car and vendor/qcom-proprietary/ais.\n"
            "3. Validated on OND — no memory leak observed.\n"
            "4. Awaiting code review and pre-integration promotion to R9.\n"
            "5. Long-term: Add memory monitoring watchdog for EVSHAL buffer pool."
        ),
        "closure": "Fix validated, awaiting R9 integration (target 2026-04-02)",
    },
    "3723553": {
        "rca_identified": "No",
        "fix_in_progress": "Yes",
        "root_cause": (
            "Probable RCA: SWU installation progress persistence failure. When HU is rebooted "
            "mid-installation, OTA agent does not checkpoint progress to persistent storage. "
            "On reboot, reads 0% and restarts installation from scratch. Missing fsync/commit "
            "of installation progress state."
        ),
        "counter_measure": (
            "1. Review OTA agent's installation progress persistence mechanism.\n"
            "2. Add periodic checkpoint writes (every 5% or per-partition) to persistent storage.\n"
            "3. Implement resume-from-checkpoint logic in OTA agent's reboot recovery path.\n"
            "4. Validate with forced reboot at 30%, 50%, 80% progress.\n"
            "5. Ensure AB partition scheme supports partial installation resume."
        ),
        "closure": "Pending SWU team analysis — target 2026-03-27",
    },
    "3719292": {
        "rca_identified": "Yes",
        "fix_in_progress": "Yes",
        "root_cause": (
            "Confirmed RCA: Incorrect/Missing Design — Old inventory in tree.xml. SWU activation "
            "flow uses tree.xml for component inventory validation. Stale default entries caused "
            "Redbend initialization blocker, HU stuck at 'Activation in Progress' indefinitely."
        ),
        "counter_measure": (
            "1. Fix: Updated tree.xml with new defaults including MSIL-specific configurations.\n"
            "2. Resolved Redbend initialization blocker (Gerrit 596428).\n"
            "3. Preventive: Add tree.xml inventory validation in SWU pre-check step.\n"
            "4. Add integration test for USB offline update activation flow."
        ),
        "closure": "Fix submitted — pending verification on R8.1+ build",
    },
    "3719329": {
        "rca_identified": "Yes",
        "fix_in_progress": "Yes",
        "root_cause": (
            "Confirmed RCA: Same as 3719292 — Old inventory in tree.xml causes Redbend init "
            "misbehavior, triggering premature activation popup at 0% installation. SWU state "
            "machine incorrectly transitions to ACTIVATION_READY before installation begins."
        ),
        "counter_measure": (
            "1. Fix: Same as 3719292 — updated tree.xml with MSIL configs + Redbend init fix.\n"
            "2. Gerrit 596428.\n"
            "3. Preventive: Add state machine guard — block ACTIVATION_READY if progress < 100%.\n"
            "4. Add assertion/log for unexpected state transitions in SWU flow."
        ),
        "closure": "Fix submitted — pending verification on R8.1+ build",
    },
    "3726560": {
        "rca_identified": "No",
        "fix_in_progress": "No (In Reproduction)",
        "root_cause": (
            "Probable RCA: BT stack crash/restart loop. Failure to pair AND unpair + BT toggling "
            "On/Off with black screen suggests BT system service crashing repeatedly. Likely "
            "null pointer or resource exhaustion in BT bonding manager on concurrent pair/unpair, "
            "causing service restart loop. Black screen = HMI loses BT service connection."
        ),
        "counter_measure": (
            "1. Collect logcat/tombstone during reproduction — check for BT service crashes.\n"
            "2. Review BT bonding manager for concurrent pair/unpair race conditions.\n"
            "3. Add crash recovery — prevent BT toggle loop on service restart.\n"
            "4. HMI: Add fallback UI instead of black screen when BT service unavailable.\n"
            "5. Test with Vivo T3X and Moto G85 (Android 15)."
        ),
        "closure": "In reproduction phase — TBD",
    },
    "3726402": {
        "rca_identified": "No",
        "fix_in_progress": "Yes",
        "root_cause": (
            "Probable RCA: Camera subsystem init causing boot loop. After battery cycle, camera "
            "HAL or EVS init fails, triggering watchdog restart. Multiple restarts with welcome "
            "logo/music = system reaches early boot but crashes before full init. Race condition "
            "between camera service startup and SOC power sequencing."
        ),
        "counter_measure": (
            "1. Capture boot logs across restart cycles — identify crashing service.\n"
            "2. Check camera HAL init timeout and watchdog config.\n"
            "3. Add graceful degradation — skip camera init on repeated failures.\n"
            "4. Review SOC power sequencing timing after battery cycle.\n"
            "5. Planned fix in R9 (target 2026-04-10)."
        ),
        "closure": "Planned fix in R9 — target 2026-04-10",
    },
    "3729547": {
        "rca_identified": "No",
        "fix_in_progress": "Yes (Analysis pending)",
        "root_cause": (
            "Probable RCA: SVS calibration data wiped during factory reset but not re-initialized "
            "on first boot. Camera pipeline starts but has no valid calibration data to render. "
            "View icons render (HMI layer works) but camera rendering layer has no valid input."
        ),
        "counter_measure": (
            "1. Verify which SVS data/configs are cleared during factory reset.\n"
            "2. Ensure SVS default calibration is restored as part of factory reset recovery.\n"
            "3. Add SVS self-check: if calibration data missing, re-init with defaults.\n"
            "4. Clone to Platform for investigation.\n"
            "5. Test: Factory reset -> reverse gear -> verify SVS renders correctly."
        ),
        "closure": "Analysis pending (cloned to Platform) — target 2026-03-31",
    },
    "3714420": {
        "rca_identified": "No",
        "fix_in_progress": "No",
        "root_cause": (
            "Probable RCA: HMI keyboard language selection UI does not persist/highlight active "
            "language state. Keyboard works in Hindi but UI selection indicator doesn't read "
            "current language setting on picker screen. Missing binding between language "
            "preference value and list item highlight state."
        ),
        "counter_measure": (
            "1. Review keyboard language selection HMI — check if selected language ID is bound to highlight.\n"
            "2. Fix: Read active keyboard language from settings and set highlight on matching item.\n"
            "3. Low priority (C3) — awaiting MSIL OPL confirmation.\n"
            "4. Target: R9.1."
        ),
        "closure": "C3 priority — deferred to R9.1, awaiting MSIL OPL",
    },
    "3727656": {
        "rca_identified": "Yes",
        "fix_in_progress": "Yes (In Integration)",
        "root_cause": (
            "Confirmed RCA: Incorrect Construction. AA projection shows black screen when AA media "
            "source selected while another projection (CP) is connected but not active. AA app fails "
            "to handle media source switch with competing projection session. Fix in Gerrit 604786."
        ),
        "counter_measure": (
            "1. Fix in Gerrit 604786 (AA projection app patch).\n"
            "2. Handles media source switch correctly with multiple projections connected.\n"
            "3. Planned for R9 integration (target 2026-04-07).\n"
            "4. Verify with: AA + CP connected -> switch to AA media source."
        ),
        "closure": "In integration — target R9 (2026-04-07)",
    },
    "3724991": {
        "rca_identified": "No",
        "fix_in_progress": "No (Validate on latest SW)",
        "root_cause": (
            "Probable RCA: Camera switch disconnect event not propagated to DID 0x1000 diagnostic "
            "data. IOC does not update Position 28 Bit 0 to '0' on disconnect. Likely GPIO interrupt "
            "or ADC threshold for disconnect not configured, or DID handler only writes on connect "
            "(edge-triggered rising only, not falling)."
        ),
        "counter_measure": (
            "1. Review IOC camera switch GPIO/ADC — verify both connect and disconnect trigger DID update.\n"
            "2. Check DID 0x1000 write handler for Position 28 Bit 0 — handle both states.\n"
            "3. Validate on latest SW.\n"
            "4. Fix IOC diagnostic handler to update DID on disconnect event.\n"
            "5. Target: 2026-04-24."
        ),
        "closure": "Pending validation on latest SW — target 2026-04-24",
    },
    "3710131": {
        "rca_identified": "Yes",
        "fix_in_progress": "Yes",
        "root_cause": (
            "Confirmed RCA: Additional speed-related handling missing in SVS 360 view. When 360 view "
            "opens at speed >10 km/h, previous frame buffer momentarily displayed before speed-based "
            "restriction kicks in. Rendering pipeline renders stale frame before speed check logic."
        ),
        "counter_measure": (
            "1. Reorder SVS rendering pipeline: perform speed check BEFORE frame render.\n"
            "2. Clear/blank frame buffer before opening 360 view at any speed.\n"
            "3. Architecture review in progress.\n"
            "4. Planned fix: R8.1 (target 2026-04-24).\n"
            "5. Validate: Open 360 at >10 km/h — no stale frame flash."
        ),
        "closure": "Architecture review ongoing — target R8.1 (2026-04-24)",
    },
    "3703960": {
        "rca_identified": "No",
        "fix_in_progress": "Yes (Analysis pending)",
        "root_cause": (
            "Probable RCA: 3D rendering artifact in SVS 360 view — grey patch at bottom of 3D "
            "vehicle model. Likely UV mapping gap or Z-buffer clipping where ground plane intersects "
            "vehicle model geometry. Could be missing texture tile in bird's-eye stitching."
        ),
        "counter_measure": (
            "1. Check 3D vehicle model mesh — inspect bottom vertices for UV mapping gaps.\n"
            "2. Review ground plane Z-offset relative to vehicle model in 360 render config.\n"
            "3. If texture issue — fix model asset; if rendering — adjust Z-clipping plane.\n"
            "4. Planned fix: R8.1 (target 2026-04-24)."
        ),
        "closure": "Analysis pending — target R8.1 (2026-04-24)",
    },
    "3726579": {
        "rca_identified": "No",
        "fix_in_progress": "No (In Reproduction)",
        "root_cause": (
            "Probable RCA: SOC boot failure after battery cycle — system hangs at black screen. "
            "Similar to 3726402. Core system service crash during early boot prevents HMI launch. "
            "Timing-dependent race condition in power management or system services init after "
            "abrupt power loss. Possible filesystem corruption or incomplete shutdown state."
        ),
        "counter_measure": (
            "1. Capture serial/UART logs during black screen to identify stuck service.\n"
            "2. Check for filesystem corruption markers (fsck) after battery cycle.\n"
            "3. Review power management shutdown sequence for clean state on abrupt power loss.\n"
            "4. Add watchdog-based recovery to force full reboot if black screen exceeds timeout.\n"
            "5. Test: 20 battery cycles — verify no persistent black screen."
        ),
        "closure": "In reproduction phase — TBD",
    },
}

# Batch 2 (21 new tickets)
RCA_BATCH2 = {
    "3733355": {
        "rca_identified": "No",
        "fix_in_progress": "Yes (Planned R9.1)",
        "root_cause": (
            "Probable RCA: Incomplete Hindi localization in HMI. After switching system language "
            "to Hindi, many UI elements, popups, condition messages, and warning texts remain in "
            "English or show incorrectly. Root cause: Hindi string resources are either missing "
            "or not mapped in HMI resource bundles for system-level screens (warnings, popups, "
            "condition messages). Likely incomplete translation coverage in language resource files."
        ),
        "counter_measure": (
            "1. Audit all HMI string resource files — identify missing Hindi translations.\n"
            "2. Complete Hindi localization for all system UI, popups, and warning messages.\n"
            "3. HMI team has planned detailed review and fixing in R9.1 phase.\n"
            "4. Validate: Switch to Hindi -> verify all screens, popups, warnings display Hindi.\n"
            "5. Add localization coverage test as part of HMI regression suite."
        ),
        "closure": "Hindi localization planned for R9.1 — target 2026-04-24",
    },
    "3733654": {
        "rca_identified": "No",
        "fix_in_progress": "No (In Reproduction)",
        "root_cause": (
            "Probable RCA: Audio routing mismatch during source switch from FM to AM. When user "
            "switches from FM to AM via 'All Sources', the tuner audio pipeline does not correctly "
            "re-route from FM decoder to AM decoder. The audio mixer continues playing the FM "
            "stream while the UI shows AM source. Likely cause: Source switch command reaches HMI "
            "but the tuner HAL does not receive or process the band-switch command, keeping FM "
            "audio path active."
        ),
        "counter_measure": (
            "1. Review tuner HAL band-switch command flow when switching FM->AM via All Sources.\n"
            "2. Add logging at tuner HAL layer to confirm band-switch command delivery.\n"
            "3. Verify audio routing table updates correctly on source change.\n"
            "4. Fix: Ensure tuner HAL processes band change before audio mixer starts AM playback.\n"
            "5. Test: FM->AM switch with multiple connected devices (BT, CP, USB)."
        ),
        "closure": "In reproduction — TBD",
    },
    "3733737": {
        "rca_identified": "No",
        "fix_in_progress": "Yes (Analysis in progress)",
        "root_cause": (
            "Probable RCA: AM source state not persisted through maintenance reset. During "
            "maintenance reboot (ACC/IGN removal), the last active source (AM) is not saved "
            "to persistent storage or the restoration logic skips AM source specifically. "
            "Likely the source persistence mechanism only covers FM/BT/USB but not AM, "
            "or the AM source ID is not included in the last-source restoration list."
        ),
        "counter_measure": (
            "1. Review source persistence logic — verify AM source is included in save/restore list.\n"
            "2. Check if AM source ID is correctly decoded during post-maintenance-reset boot.\n"
            "3. Fix: Add AM to the persisted source list alongside FM.\n"
            "4. Advanced analysis in progress (per RespNote 02-Apr).\n"
            "5. Validate: Set AM active -> maintenance reset -> verify AM restores on boot."
        ),
        "closure": "Advanced analysis in progress — target 2026-04-10",
    },
    "3733996": {
        "rca_identified": "No",
        "fix_in_progress": "No",
        "root_cause": (
            "Probable RCA: AA low battery popup rendering issue. When Android Auto reconnects, "
            "the phone's low battery notification triggers a system popup overlay. The popup "
            "renders with incorrect Z-order (black background) and the close button's click "
            "handler is not bound (disabled state). Likely cause: The popup is rendered by the "
            "projection overlay layer instead of the native HMI dialog layer, causing incorrect "
            "background and missing button state initialization."
        ),
        "counter_measure": (
            "1. Review AA popup overlay rendering path — check if low-battery popup uses correct dialog layer.\n"
            "2. Fix: Route phone notification popups through native HMI dialog framework (not projection overlay).\n"
            "3. Ensure close button click handler is properly bound during popup initialization.\n"
            "4. Workaround: Battery cycle recovers — add auto-dismiss timer as fallback.\n"
            "5. Validate: AA reconnect with low battery phone -> verify popup renders correctly."
        ),
        "closure": "Pending analysis — TBD",
    },
    "3734657": {
        "rca_identified": "No",
        "fix_in_progress": "No",
        "root_cause": (
            "Probable RCA: Media (AV) Off state machine stuck. When BT music is active and user "
            "enables Media Off, the state transition to disable Media Off fails because the BT "
            "A2DP audio stream is still active. The Media Off toggle logic doesn't handle the "
            "transition back to active source when an A2DP stream is playing. Selecting BT source "
            "from All Sources also fails because the Media Off state blocks source switch."
        ),
        "counter_measure": (
            "1. Review Media Off state machine — ensure disable transition works with active A2DP.\n"
            "2. Fix: When disabling Media Off, check for active BT stream and resume A2DP playback.\n"
            "3. Source switch from All Sources should force-exit Media Off state.\n"
            "4. Validate: Enable Media Off with BT active -> disable -> verify BT resumes.\n"
            "5. Test: Media Off toggle with each source type (BT, FM, AM, USB)."
        ),
        "closure": "Pending analysis — TBD",
    },
    "3734662": {
        "rca_identified": "No",
        "fix_in_progress": "No",
        "root_cause": (
            "Probable RCA: HFP call routing issue during multi-party/add-call scenario. When user "
            "adds a second call from active call screen, the telephony manager incorrectly routes "
            "the 2nd call as private (audio to phone earpiece instead of HU speaker). Likely cause: "
            "The add-call intent does not carry the audio routing preference flag, or the HFP AG "
            "(Audio Gateway) on the phone defaults to private mode for conference/add-call."
        ),
        "counter_measure": (
            "1. Review HFP add-call flow — check if audio routing flag is set for HU speaker output.\n"
            "2. Verify AT+CHLD command handling for call addition in BT phone module.\n"
            "3. Fix: Force SCO audio routing to HU when 2nd call is initiated from HU call screen.\n"
            "4. Validate: Active call -> add call -> verify 2nd call audio plays on HU speakers.\n"
            "5. Test across multiple phone models (Samsung, iPhone, Motorola)."
        ),
        "closure": "Pending analysis — target 2026-04-24",
    },
    "3734687": {
        "rca_identified": "No",
        "fix_in_progress": "No",
        "root_cause": (
            "Probable RCA: Call status UI not updated when call is initiated from incoming message "
            "popup. When user taps call-back from a BT message notification popup, the HFP call "
            "is initiated but the HMI call status screen is not launched. Likely cause: The message "
            "popup's call-back action trigger uses a different intent path than the standard dialer, "
            "bypassing the call status screen launch callback in the BT Phone HMI module."
        ),
        "counter_measure": (
            "1. Review message popup call-back intent flow — trace from popup action to call screen.\n"
            "2. Fix: Ensure call-back from message popup uses same call initiation path as dialer.\n"
            "3. Or: Register callback listener in BT Phone HMI for all HFP call initiation paths.\n"
            "4. Validate: Receive message -> tap call from popup -> verify call status screen shows.\n"
            "5. Test with Samsung S24, iPhone 16, Android phone (as reported in Fleet test)."
        ),
        "closure": "Pending analysis — TBD",
    },
    "3734429": {
        "rca_identified": "No",
        "fix_in_progress": "No",
        "root_cause": (
            "Probable RCA: A2DP audio routing lost after ACC cycle during CarPlay connection attempt. "
            "During ACC Off/On, if CarPlay connection was being established, the BT reconnection "
            "sequence prioritizes CP/IAP2 profile over A2DP. After ACC On, BT reconnects but A2DP "
            "AVRCP control stays with phone (plays from phone speaker) instead of routing to HU. "
            "Likely cause: DCSM reconnection priority logic doesn't re-establish A2DP sink after "
            "CarPlay connection fails/interrupts."
        ),
        "counter_measure": (
            "1. Review DCSM BT reconnection priority after ACC cycle with pending CP connection.\n"
            "2. Fix: After ACC On, ensure A2DP sink role is re-established regardless of CP state.\n"
            "3. Add explicit AVRCP target check after BT reconnection to confirm HU is audio sink.\n"
            "4. Validate: BT music playing -> start CP connect -> ACC Off/On -> verify BT plays on HU.\n"
            "5. Workaround: Disconnect and reconnect BT manually."
        ),
        "closure": "Pending analysis — target 2026-04-07",
    },
    "3732170": {
        "rca_identified": "No",
        "fix_in_progress": "Yes",
        "root_cause": (
            "Probable RCA: AA disconnect button not functional in Paired Devices screen. The AA "
            "toggle/button in BT paired devices list is unresponsive — tapping it does not trigger "
            "AA session disconnect. Likely cause: The AA disconnect action handler is not registered "
            "or bound to the AA button widget in the paired devices list adapter. The button renders "
            "but has no click listener attached."
        ),
        "counter_measure": (
            "1. Review paired devices list adapter — check AA button click listener registration.\n"
            "2. Fix: Bind AA disconnect action handler to the AA button in device list item.\n"
            "3. Ensure AA session termination API is called when button is tapped.\n"
            "4. Workaround: Unpair device to disconnect AA (as noted in recovery).\n"
            "5. Validate: Connect AA -> tap AA button in paired devices -> verify disconnects."
        ),
        "closure": "In processing — target 2026-04-17",
    },
    "3733333": {
        "rca_identified": "No",
        "fix_in_progress": "Yes",
        "root_cause": (
            "Probable RCA: HU crash (black screen) triggered by rapid home key taps during app "
            "launcher rearrange mode. When AA is connected and user enters app rearrange mode, "
            "rapid home key presses cause a race condition between the rearrange animation/state "
            "and the home navigation handler. Likely a null reference or unhandled state in the "
            "launcher activity when home is pressed during drag mode, causing system_server "
            "or launcher process crash."
        ),
        "counter_measure": (
            "1. Review launcher rearrange mode — add guard against home key during drag state.\n"
            "2. Fix: Disable or queue home key events while rearrange animation is in progress.\n"
            "3. Add try-catch in launcher's onHomePressed handler for rearrange state.\n"
            "4. Previously wrongly rejected (per InternalStatement) — re-opened for proper fix.\n"
            "5. Validate: AA connected -> rearrange mode -> rapid home taps -> verify no crash."
        ),
        "closure": "In processing — TBD",
    },
    "3733699": {
        "rca_identified": "No",
        "fix_in_progress": "Yes",
        "root_cause": (
            "Probable RCA: Inconsistent BT/Projection auto-reconnection after ACC cycle. When "
            "multiple devices are connected (iPhone via CP, Motorola via AA+HFP, iPhone 12 mini "
            "via A2DP), the auto-reconnection sequence after ACC Off/On does not reliably restore "
            "all profiles. DCSM reconnection priority logic fails to reconnect all profiles in "
            "the correct order. Likely cause: Reconnection timeout per device is too short, or "
            "the priority queue doesn't handle 3+ device reconnection."
        ),
        "counter_measure": (
            "1. Review DCSM multi-device auto-reconnection priority and timeout settings.\n"
            "2. Increase reconnection timeout window for 3+ paired devices.\n"
            "3. Fix: Implement retry logic for failed profile reconnections after ACC cycle.\n"
            "4. Validate: 3 devices connected (CP + AA + A2DP) -> ACC cycle -> verify all reconnect.\n"
            "5. Add reconnection status logging for fleet/drive test diagnostics."
        ),
        "closure": "In processing — target 2026-04-17",
    },
    "3733702": {
        "rca_identified": "No",
        "fix_in_progress": "Yes",
        "root_cause": (
            "Probable RCA: BT and Projection profile mismatch after ACC cycle. After ACC Off/On, "
            "the BT profiles (HFP/A2DP) reconnect to a different device than the projection profile "
            "(CP/AA), causing a mismatch where phone calls route to one device but projection shows "
            "another. Likely cause: DCSM reconnects BT profiles in order of last-connected but "
            "projection re-establishes based on USB/WiFi detection, creating a desync between "
            "BT device binding and projection session binding."
        ),
        "counter_measure": (
            "1. Review DCSM profile-device binding synchronization after ACC reconnection.\n"
            "2. Fix: Ensure projection session and BT profiles are bound to same device after ACC cycle.\n"
            "3. Add cross-check between BT reconnected device and projection session device.\n"
            "4. Validate: CP + AA + BT devices -> ACC cycle -> verify correct profile-device mapping.\n"
            "5. Add device-profile mapping verification in DCSM reconnection completion callback."
        ),
        "closure": "In processing — target 2026-04-17",
    },
    "3733909": {
        "rca_identified": "No",
        "fix_in_progress": "No",
        "root_cause": (
            "Probable RCA: Wireless CarPlay disconnect blocked because BT is turned off. When "
            "WLCP is active, navigating to Settings > Bluetooth > Paired Devices shows BT is off. "
            "Since paired devices list requires BT on, user cannot access WLCP device to disconnect. "
            "Likely cause: WLCP session uses WiFi Direct for data but disables BT to prevent "
            "interference, or a BT power state conflict between WLCP stack and BT settings UI."
        ),
        "counter_measure": (
            "1. Review WLCP BT power management — check if WLCP session forcibly turns off BT.\n"
            "2. Fix Option A: Keep BT enabled when WLCP is active (WLCP can coexist with BT).\n"
            "3. Fix Option B: Add WLCP disconnect option in Projection settings (not dependent on BT).\n"
            "4. Workaround: Manually turn on BT, then disconnect WLCP.\n"
            "5. Validate: WLCP active -> navigate to paired devices -> verify BT is on and can disconnect."
        ),
        "closure": "Pending analysis — TBD",
    },
    "3734271": {
        "rca_identified": "No",
        "fix_in_progress": "Yes",
        "root_cause": (
            "Probable RCA: HMI stuck on loading screen after cold boot. CarPlay component "
            "(FGroup: Projection/CarPlay) initialization is blocking the HMI rendering pipeline. "
            "After cold boot, the CarPlay service or a dependent system service hangs during "
            "initialization, preventing the home screen from rendering. Issue is intermittent "
            "(Sometimes) suggesting a timing-dependent deadlock or resource contention during "
            "boot sequence."
        ),
        "counter_measure": (
            "1. Capture boot timing logs — identify which service blocks HMI rendering.\n"
            "2. Review CarPlay service init — add timeout and fallback if init exceeds threshold.\n"
            "3. Fix: Decouple HMI home screen rendering from projection service initialization.\n"
            "4. Add boot watchdog — if loading screen exceeds 30s, force-kill blocking service.\n"
            "5. Planned fix R9 (target 2026-04-08). Recovery: Cold boot."
        ),
        "closure": "Planned fix R9 — target 2026-04-08",
    },
    "3734865": {
        "rca_identified": "No",
        "fix_in_progress": "No",
        "root_cause": (
            "Probable RCA: AA ringtone audio not routed to HU speakers during incoming call. "
            "When Android Auto is active and an incoming call is received, the phone's ringtone "
            "audio stream is not routed through the HU audio output. Likely cause: The HFP "
            "incoming call ringtone (in-band ringing) is not properly configured for the AA "
            "session, or the BT audio focus manager does not request ringtone audio focus from "
            "the AA audio pipeline."
        ),
        "counter_measure": (
            "1. Review HFP in-band ringing configuration during active AA session.\n"
            "2. Check if AA audio focus manager grants ringtone stream priority.\n"
            "3. Fix: Configure in-band ringing to route through HU speaker during AA session.\n"
            "4. If phone uses out-of-band ringing — add HFP +BSIR handling to switch to in-band.\n"
            "5. Validate: AA active -> incoming call -> verify ringtone audible on HU speakers."
        ),
        "closure": "Pending BT analysis — TBD",
    },
    "3734870": {
        "rca_identified": "Yes",
        "fix_in_progress": "Yes (In Integration)",
        "root_cause": (
            "Confirmed RCA: Incorrect Construction — Missing audio permission in AndroidManifest.xml. "
            "Wireless Android Auto voice recognition not working because AA was receiving silent "
            "audio data (no microphone input). The RECORD_AUDIO permission was missing in the AA "
            "app manifest, causing the microphone stream to return empty/silent buffers."
        ),
        "counter_measure": (
            "1. Fix: Added missing RECORD_AUDIO permission in AA manifest.\n"
            "2. AA now receives proper microphone input for voice recognition.\n"
            "3. Planned for R9 integration (target 2026-04-07).\n"
            "4. Preventive: Add permission audit check in AA build pipeline.\n"
            "5. Validate: Wireless AA -> trigger VR -> verify voice input is captured and processed."
        ),
        "closure": "Fix in integration — target R9 (2026-04-07)",
    },
    "3734607": {
        "rca_identified": "No",
        "fix_in_progress": "Yes",
        "root_cause": (
            "Probable RCA: HFP profile not reconnecting when AA session is switched between devices. "
            "During drive test, switching AA from Device-1 to Device-2, the DCSM disconnects "
            "Device-1 AA but does not re-establish HFP profile for Device-2. The HFP connection "
            "remains bound to Device-1 (now disconnected) while AA moves to Device-2. Likely cause: "
            "DCSM AA switch handler disconnects all profiles from Device-1 but only reconnects AA "
            "profile on Device-2, leaving HFP orphaned."
        ),
        "counter_measure": (
            "1. Review DCSM AA device-switch handler — ensure HFP is reconnected on new AA device.\n"
            "2. Fix: When switching AA between devices, transfer all BT profiles (HFP + A2DP + AA).\n"
            "3. Add HFP connection state verification after AA device switch completion.\n"
            "4. Validate: AA on Device-1 -> switch to Device-2 -> verify HFP connected on Device-2.\n"
            "5. Drive test scenario with Motorola 50 fusion + S22."
        ),
        "closure": "In processing — TBD",
    },
    "3734601": {
        "rca_identified": "No",
        "fix_in_progress": "Yes",
        "root_cause": (
            "Probable RCA: DCSM device connection list not updated after AA session change. When only "
            "one device is connected via AA, the HU shows two devices as AA-connected. Likely cause: "
            "When Device-1 AA session ends and Device-2 connects, DCSM does not clear Device-1's AA "
            "connection flag in the connected devices list. The UI reads stale connection state — "
            "a missing state cleanup in DCSM's onSessionDisconnected callback."
        ),
        "counter_measure": (
            "1. Review DCSM onSessionDisconnected — ensure AA connection flag is cleared for old device.\n"
            "2. Fix: Clear all projection connection flags for disconnected device in DCSM state.\n"
            "3. Add connected device list refresh after each AA session change.\n"
            "4. Validate: Connect AA Device-1 -> switch to Device-2 -> verify only Device-2 shows connected.\n"
            "5. Drive test with Motorola 50 fusion and S22."
        ),
        "closure": "In processing — TBD",
    },
    "3734599": {
        "rca_identified": "No",
        "fix_in_progress": "Yes",
        "root_cause": (
            "Probable RCA: A2DP connection blocked for 2nd device when AA is active on 1st device. "
            "DCSM connection policy may restrict A2DP connections when an AA session is active, "
            "treating AA as exclusive audio source. Or the BT stack's A2DP acceptor is paused/blocked "
            "while AA audio pipeline holds the audio focus. Likely cause: Connection policy rule "
            "that limits concurrent A2DP + AA sessions across different devices."
        ),
        "counter_measure": (
            "1. Review DCSM connection policy for A2DP when AA is active on another device.\n"
            "2. Fix: Allow A2DP connection for 2nd device while AA is active (BT audio coexistence).\n"
            "3. Ensure A2DP audio is routed correctly when both AA and A2DP are connected.\n"
            "4. Validate: AA on Device-1 -> connect Device-2 A2DP -> verify A2DP connects.\n"
            "5. Drive test scenario with Motorola 50 fusion and S22."
        ),
        "closure": "In processing — TBD",
    },
    "3734957": {
        "rca_identified": "No",
        "fix_in_progress": "No (In Reproduction)",
        "root_cause": (
            "Probable RCA: AA projection freezes when VR session is interrupted by incoming call. "
            "During active wired AA session in vehicle, initiating VR multiple times followed by "
            "an incoming call creates a conflict between VR audio session and call audio session. "
            "After accepting the call and pressing home button (left edge swipe), the AA projection "
            "surface cannot re-launch. Likely cause: VR session holds audio focus, call steals it, "
            "and AA projection surface is not restored after call ends."
        ),
        "counter_measure": (
            "1. Review AA VR + incoming call audio focus transition handling.\n"
            "2. Fix: Ensure AA projection surface is restored after VR->Call->Home sequence.\n"
            "3. Add audio focus release handler when VR session is interrupted by incoming call.\n"
            "4. Validate: AA active -> VR -> incoming call -> accept -> home -> verify AA re-launches.\n"
            "5. In reproduction phase — vehicle test scenario."
        ),
        "closure": "In reproduction — TBD",
    },
    "3734858": {
        "rca_identified": "No",
        "fix_in_progress": "No",
        "root_cause": (
            "Probable RCA: Downlink call audio (far-end voice) not routed to HU speakers during "
            "Wireless AA call. When a call is active via Wireless AA, the SCO (Synchronous Connection "
            "Oriented) audio link for downlink is not established or the audio routing table does not "
            "map SCO downlink to HU speaker output. Likely cause: Wireless AA uses WiFi for data but "
            "call audio still needs BT HFP SCO link — the SCO link may not be established for WLAA "
            "calls, or audio mixer input for SCO downlink is muted/disconnected."
        ),
        "counter_measure": (
            "1. Review SCO link establishment for calls during Wireless AA session.\n"
            "2. Check audio routing table — verify SCO downlink mapped to HU speaker output.\n"
            "3. Fix: Ensure HFP SCO link is active for call audio when Wireless AA uses WiFi for data.\n"
            "4. Check audio mixer — verify SCO downlink input is not muted.\n"
            "5. Validate: Wireless AA -> call -> verify far-end voice audible on HU speakers."
        ),
        "closure": "Pending BT/audio analysis — TBD",
    },
}


def main():
    # Load both batches of defect data
    all_defects = {}
    for fname in ['all_defects_dump.json', 'all_defects_batch2.json']:
        try:
            with open(fname, 'r', encoding='utf-8') as f:
                all_defects.update(json.load(f))
        except FileNotFoundError:
            print(f"Warning: {fname} not found, skipping.")

    # Merge both RCA batches
    all_rca = {}
    all_rca.update(RCA_BATCH1)
    all_rca.update(RCA_BATCH2)

    rows = []
    for tid_str, rca_info in all_rca.items():
        d = all_defects.get(tid_str, {})
        rows.append({
            "Requested Priority": d.get("PriorityID", ""),
            "Occurrence": d.get("Occurance", ""),
            "Platform/Project": d.get("SubProject", "") or "Intelligent Cockpit Platform",
            "Functional group": d.get("FGroup", ""),
            "Ticket ID": int(tid_str),
            "Title": d.get("Title", ""),
            "Reported on": d.get("EnterDateTime", ""),
            "Root cause identified\n[Yes/No]": rca_info["rca_identified"],
            " Fix in Progress [Yes/No]": rca_info["fix_in_progress"],
            "What is the root cause?": rca_info["root_cause"],
            "What is the counter measure?": rca_info["counter_measure"],
            "When this ticket can be closed?": rca_info["closure"],
        })

    df = pd.DataFrame(rows)

    # Write to Excel
    wb = Workbook()
    ws = wb.active
    ws.title = "RCA Output"

    headers = list(df.columns)
    header_font = Font(bold=True, size=11, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell_align = Alignment(vertical="top", wrap_text=True)
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )
    yes_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    no_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")

    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border

    for row_idx, (_, row) in enumerate(df.iterrows(), 2):
        for col_idx, header in enumerate(headers, 1):
            val = row[header]
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.alignment = cell_align
            cell.border = thin_border
            if header == "Root cause identified\n[Yes/No]":
                cell.fill = yes_fill if val == "Yes" else no_fill
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    col_widths = {
        "Requested Priority": 14, "Occurrence": 12, "Platform/Project": 22,
        "Functional group": 18, "Ticket ID": 12, "Title": 55,
        "Reported on": 14, "Root cause identified\n[Yes/No]": 14,
        " Fix in Progress [Yes/No]": 18, "What is the root cause?": 75,
        "What is the counter measure?": 75, "When this ticket can be closed?": 30,
    }
    for col_idx, header in enumerate(headers, 1):
        col_letter = ws.cell(row=1, column=col_idx).column_letter
        ws.column_dimensions[col_letter].width = col_widths.get(header, 15)

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    wb.save(EXCEL_OUT)
    print(f"Done! Saved {len(rows)} tickets to: {EXCEL_OUT}")
    print(f"Sheet: 'RCA Output'\n")

    # Summary
    yes_count = sum(1 for r in rows if r["Root cause identified\n[Yes/No]"] == "Yes")
    no_count = len(rows) - yes_count
    print(f"RCA Identified: {yes_count} Yes, {no_count} No (total: {len(rows)})")
    print()
    for r in rows:
        tid = r["Ticket ID"]
        rca = r["Root cause identified\n[Yes/No]"]
        fg = r["Functional group"]
        title = r["Title"][:55]
        print(f"  {tid} | {rca:3s} | {fg:30s} | {title}")


if __name__ == "__main__":
    main()
