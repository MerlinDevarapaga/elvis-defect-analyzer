"""
Inflow & Outflow Overview Report — Generates a domain-wise summary of
ticket inflow, outflow (Integrating/Concluding), and rejected tickets
for the last 4 weeks (overall totals, no week-wise breakdown).

Columns:
  - Inflow:    all tickets created (EnterDateTime) in the period
  - Outflow:   tickets currently in Integrating/Concluding whose
               FirstIntegrDateTime falls in the period (non-rejected)
  - Rejected:  tickets created in the period that are rejected

Filters:
  - ProjectID = 'MSIL_DA2.8'
  - IsDeleted = 'N'

Usage:
    python scripts/weekly_inflow_integrating_report.py
    python scripts/weekly_inflow_integrating_report.py --excel
    python scripts/weekly_inflow_integrating_report.py --excel output.xlsx
"""
import os
import sys
import io
import argparse
from datetime import datetime, date, timedelta
from collections import defaultdict

sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8", errors="replace")

from dotenv import load_dotenv
import mysql.connector

# Load .env from workspace root
_script_dir = os.path.dirname(os.path.abspath(__file__))
for _env in [os.path.join(_script_dir, "..", ".env"), os.path.join(_script_dir, ".env")]:
    if os.path.exists(_env):
        load_dotenv(_env)
        break

# ---------------------------------------------------------------------------
# Configuration
# ---------------------------------------------------------------------------
PROJECT_ID = "MSIL_DA2.8"
NUM_WEEKS = 4

# FGroup display names (consistent with workshop)
FGROUP_DISPLAY = {
    "Media": "Media",
    "Bluetooth": "BT",
    "Projection": "Projection",
    "Audio": "Audio",
    "IOC": "IOC",
    "Camera": "Camera",
    "WiFi": "Wifi",
    "Systems - Core": "Systems Core",
    "Systems - Infra": "Sys Infra",
    "Systems - SWU Software Update": "SWUP",
    "USB": "USB",
    "SVS": "SVS",
}


def get_connection():
    return mysql.connector.connect(
        host=os.getenv("ELVIS_DB_HOST"),
        user=os.getenv("ELVIS_DB_USER"),
        password=os.getenv("ELVIS_DB_PASSWORD"),
        database=os.getenv("ELVIS_DB_NAME"),
        port=int(os.getenv("ELVIS_DB_PORT", 3306)),
    )


def get_period(num_weeks):
    """Return (start_date, end_date) covering the last N full weeks (Mon-Sun)."""
    today = date.today()
    current_monday = today - timedelta(days=today.weekday())
    start = current_monday - timedelta(weeks=num_weeks)
    end = current_monday - timedelta(days=1)  # last Sunday
    return start, end


def display_fg(fg):
    return FGROUP_DISPLAY.get((fg or "Unknown").strip(), (fg or "Unknown").strip())


def fetch_inflow(start_date, end_date):
    """Fetch ALL tickets created in the period (including rejected)."""
    conn = get_connection()
    cursor = conn.cursor(dictionary=True)
    end_exclusive = end_date + timedelta(days=1)
    query = """
        SELECT TicketID, FGroup, Rejected
        FROM tbl_ElvisSR
        WHERE ProjectID = %s
          AND EnterDateTime >= %s
          AND EnterDateTime < %s
          AND IsDeleted = 'N'
        ORDER BY FGroup
    """
    cursor.execute(query, [PROJECT_ID, start_date.strftime("%Y-%m-%d"),
                           end_exclusive.strftime("%Y-%m-%d")])
    rows = cursor.fetchall()
    cursor.close()
    conn.close()
    return rows


def fetch_outflow(start_date, end_date):
    """Fetch non-rejected tickets currently in Integrating/Concluding whose
    FirstIntegrDateTime falls in the period."""
    conn = get_connection()
    cursor = conn.cursor(dictionary=True)
    end_exclusive = end_date + timedelta(days=1)
    current_steps = ("Integrating", "Concluding")
    step_ph = ", ".join(["%s"] * len(current_steps))
    query = f"""
        SELECT TicketID, FGroup
        FROM tbl_ElvisSR
        WHERE ProjectID = %s
          AND TicketStepID IN ({step_ph})
          AND FirstIntegrDateTime >= %s
          AND FirstIntegrDateTime < %s
          AND IsDeleted = 'N'
          AND Rejected = 'N'
        ORDER BY FGroup
    """
    params = [PROJECT_ID] + list(current_steps) + [
        start_date.strftime("%Y-%m-%d"),
        end_exclusive.strftime("%Y-%m-%d"),
    ]
    cursor.execute(query, params)
    rows = cursor.fetchall()
    cursor.close()
    conn.close()
    return rows


def build_report():
    """Build domain-wise totals for inflow, outflow, rejected."""
    start, end = get_period(NUM_WEEKS)
    period_label = f"{start.strftime('%d-%b-%Y')} to {end.strftime('%d-%b-%Y')}"

    print(f"  Period: {period_label}")
    print(f"  Fetching inflow ...")
    inflow_rows = fetch_inflow(start, end)
    print(f"  Fetching outflow (Integrating/Concluding) ...")
    outflow_rows = fetch_outflow(start, end)

    # Count inflow and rejected (from inflow) per domain
    inflow_counts = defaultdict(int)
    rejected_counts = defaultdict(int)
    for row in inflow_rows:
        fg = display_fg(row["FGroup"])
        inflow_counts[fg] += 1
        if row.get("Rejected") == "Y":
            rejected_counts[fg] += 1

    # Count outflow per domain
    outflow_counts = defaultdict(int)
    for row in outflow_rows:
        fg = display_fg(row["FGroup"])
        outflow_counts[fg] += 1

    all_domains = sorted(set(inflow_counts.keys()) | set(outflow_counts.keys()))

    return all_domains, inflow_counts, outflow_counts, rejected_counts, period_label


def print_report(domains, inflow, outflow, rejected, period_label):
    """Pretty-print the overview table."""
    name_w = max(len("Domain"), max((len(d) for d in domains), default=10))
    col_w = 10

    print(f"\n{'=' * 70}")
    print(f"  MSIL DA2.8 — Inflow & Outflow Overview (Last {NUM_WEEKS} Weeks)")
    print(f"  Period: {period_label}")
    print(f"  Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')}")
    print(f"{'=' * 70}")

    header = f"{'Domain':<{name_w}}  {'Inflow':>{col_w}}  {'Outflow':>{col_w}}  {'Rejected':>{col_w}}"
    print(header)
    print("-" * len(header))

    total_in, total_out, total_rej = 0, 0, 0
    for domain in domains:
        i = inflow.get(domain, 0)
        o = outflow.get(domain, 0)
        r = rejected.get(domain, 0)
        total_in += i
        total_out += o
        total_rej += r
        i_s = str(i) if i else "·"
        o_s = str(o) if o else "·"
        r_s = str(r) if r else "·"
        print(f"{domain:<{name_w}}  {i_s:>{col_w}}  {o_s:>{col_w}}  {r_s:>{col_w}}")

    print("-" * len(header))
    print(f"{'Total':<{name_w}}  {total_in:>{col_w}}  {total_out:>{col_w}}  {total_rej:>{col_w}}")
    print()


def write_excel(domains, inflow, outflow, rejected, period_label, output_path):
    """Write the overview table to Excel."""
    try:
        import openpyxl
        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    except ImportError:
        print("Error: openpyxl not installed. Run: pip install openpyxl")
        return

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Overview"

    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    total_font = Font(bold=True, size=11)
    total_fill = PatternFill(start_color="D9E2F3", end_color="D9E2F3", fill_type="solid")
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )

    # Title
    ws.merge_cells("A1:D1")
    tc = ws.cell(row=1, column=1,
                 value=f"MSIL DA2.8 — Inflow & Outflow Overview (Last {NUM_WEEKS} Weeks)")
    tc.font = Font(bold=True, size=14)
    tc.alignment = Alignment(horizontal="center")

    ws.cell(row=2, column=1, value=f"Period: {period_label}").font = Font(italic=True, size=9)
    ws.cell(row=3, column=1,
            value=f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')}").font = Font(italic=True, size=9)

    start_row = 5
    headers = ["Domain", "Inflow", "Outflow", "Rejected"]
    for ci, h in enumerate(headers, 1):
        cell = ws.cell(row=start_row, column=ci, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", wrap_text=True)
        cell.border = thin_border

    total_in, total_out, total_rej = 0, 0, 0
    for ri, domain in enumerate(domains, start_row + 1):
        i = inflow.get(domain, 0)
        o = outflow.get(domain, 0)
        r = rejected.get(domain, 0)
        total_in += i
        total_out += o
        total_rej += r

        ws.cell(row=ri, column=1, value=domain).border = thin_border
        for ci, val in enumerate([i, o, r], 2):
            cell = ws.cell(row=ri, column=ci, value=val if val else "")
            cell.border = thin_border
            cell.alignment = Alignment(horizontal="center")

    # Total row
    tr = start_row + len(domains) + 1
    for ci, val in enumerate(["Total", total_in, total_out, total_rej], 1):
        cell = ws.cell(row=tr, column=ci, value=val)
        cell.font = total_font
        cell.fill = total_fill
        cell.border = thin_border
        if ci > 1:
            cell.alignment = Alignment(horizontal="center")

    # Auto-width
    for ci in range(1, 5):
        max_len = max(len(str(ws.cell(row=r, column=ci).value or ""))
                      for r in range(start_row, tr + 1))
        ws.column_dimensions[openpyxl.utils.get_column_letter(ci)].width = max(max_len + 4, 12)

    wb.save(output_path)
    print(f"\nExcel saved to: {output_path}")


def main():
    global NUM_WEEKS
    parser = argparse.ArgumentParser(description="Inflow & Outflow Overview Report")
    parser.add_argument("--excel", nargs="?", const="auto", default=None,
                        help="Export to Excel. Optionally specify output path.")
    parser.add_argument("--weeks", type=int, default=NUM_WEEKS,
                        help=f"Number of weeks to look back (default: {NUM_WEEKS})")
    args = parser.parse_args()

    NUM_WEEKS = args.weeks

    print("Fetching data from Elvis DB ...")
    domains, inflow, outflow, rejected, period_label = build_report()

    if not domains:
        print("No data found for the specified period.")
        return

    print_report(domains, inflow, outflow, rejected, period_label)

    if args.excel is not None:
        if args.excel == "auto":
            output_dir = os.path.join(_script_dir, "..", "docs", "output")
            os.makedirs(output_dir, exist_ok=True)
            output_path = os.path.join(output_dir,
                                       f"weekly_inflow_integrating_{date.today().strftime('%Y%m%d')}.xlsx")
        else:
            output_path = args.excel
        write_excel(domains, inflow, outflow, rejected, period_label, output_path)


if __name__ == "__main__":
    main()
