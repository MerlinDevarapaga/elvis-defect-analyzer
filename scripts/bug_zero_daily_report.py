"""
DA2.8 May End Bug Zero — Daily Tracking Report

Generates a daily report in the Template.xlsx format with:
- Domain-wise Total Open, Inflow, Outflow for the last 9 days
- Email-ready summary for Bug Zero progress

Bug Zero Filter:
  (ReferenceNumber <= 2) AND (FG_SWRev != 'P8_YTB_NA')
  AND (Priority IN (A(1), top) OR (B(2)/C(3) AND Occurrence != Once))

Usage:
    python bug_zero_daily_report.py
"""
import os, sys, io
from datetime import datetime, timedelta, date
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8", errors="replace")
from dotenv import load_dotenv
import mysql.connector

_script_dir = os.path.dirname(os.path.abspath(__file__))
for _env in [os.path.join(_script_dir, ".env"), os.path.join(_script_dir, "..", ".env")]:
    if os.path.exists(_env):
        load_dotenv(_env)
        break

# ---------------------------------------------------------------------------
# FGroup values are used directly as domain names — no mapping needed.
# These are the actual Elvis FGroup values from the Bug Zero filter.
# ---------------------------------------------------------------------------
FGROUP_TO_DOMAIN = {}  # identity mapping — FGroup is the domain

# Template domains = actual FGroup values in alphabetical order
TEMPLATE_DOMAINS = [
    "Audio",
    "Bluetooth",
    "Camera",
    "Dead Reckoning- GNSS",
    "External Suppliers",
    "HMI IVI",
    "IOC",
    "Kernel",
    "Media",
    "Others",
    "Projection",
    "RCA",
    "Requirements",
    "Security",
    "SVS",
    "Systems - Core",
    "Systems - Infra",
    "Systems - SWU Software Update",
    "Systems Engineering",
    "TCU communication",
    "Tuner",
    "USB",
    "Voice Recognition",
    "Webportal",
    "WiFi",
]

# Bug Zero filter WHERE clause
BUG_ZERO_WHERE = """
    `ProjectID` = 'MSIL_DA2.8'
    AND `IsDeleted` = 'N'
    AND (`ReferenceNumber` IS NULL OR `ReferenceNumber` <= 2)
    AND (`FG_SWRev` IS NULL OR `FG_SWRev` != 'P8_YTB_NA')
    AND (
        `PriorityID` IN ('A(1)', 'top')
        OR (`PriorityID` = 'B(2)' AND `Occurance` != 'Once')
        OR (`PriorityID` = 'C(3)' AND `Occurance` != 'Once')
    )
"""

# Steps considered "open" (pre-integrating only)
OPEN_STEPS = ("Categorizing", "Reproduction", "Processing")
OUTFLOW_STEPS = ("Closed",)


def get_connection():
    return mysql.connector.connect(
        host=os.getenv("ELVIS_DB_HOST"),
        user=os.getenv("ELVIS_DB_USER"),
        password=os.getenv("ELVIS_DB_PASSWORD"),
        database=os.getenv("ELVIS_DB_NAME"),
        port=int(os.getenv("ELVIS_DB_PORT", 3306)),
        connection_timeout=15,
    )


def map_domain(fgroup):
    if not fgroup or fgroup.strip() == "":
        return "Others"
    return FGROUP_TO_DOMAIN.get(fgroup, fgroup)


def fetch_bug_zero_data(cursor, num_days=9):
    """Fetch all Bug Zero tickets and compute daily inflow/outflow by domain."""
    today = date.today()
    dates = [today - timedelta(days=i) for i in range(num_days)]  # today, yesterday, ...

    # ── 1. Current total open per domain (pre-integrating only) ──
    open_steps_sql = "','".join(OPEN_STEPS)
    cursor.execute(f"""
        SELECT `FGroup`, COUNT(*) as cnt
        FROM tbl_ElvisSR
        WHERE {BUG_ZERO_WHERE}
          AND `TicketStepID` IN ('{open_steps_sql}')
        GROUP BY `FGroup`
    """)
    total_open = {}
    for r in cursor.fetchall():
        domain = map_domain(r["FGroup"])
        total_open[domain] = total_open.get(domain, 0) + r["cnt"]

    # ── 2. Daily inflow (tickets entered on each date) ──
    earliest = dates[-1]
    cursor.execute(f"""
        SELECT `FGroup`, DATE(`EnterDateTime`) as enter_date, COUNT(*) as cnt
        FROM tbl_ElvisSR
        WHERE {BUG_ZERO_WHERE}
          AND DATE(`EnterDateTime`) >= %s
        GROUP BY `FGroup`, DATE(`EnterDateTime`)
    """, (earliest,))
    daily_inflow = {}  # {date: {domain: count}}
    for d in dates:
        daily_inflow[d] = {dom: 0 for dom in TEMPLATE_DOMAINS}
    for r in cursor.fetchall():
        d = r["enter_date"]
        if isinstance(d, datetime):
            d = d.date()
        if d in daily_inflow:
            domain = map_domain(r["FGroup"])
            daily_inflow[d][domain] = daily_inflow[d].get(domain, 0) + r["cnt"]

    # ── 3. Daily outflow: tickets moved to Integrating OR rejected ──
    # 3a. Moved to Integrating (by FirstIntegrDateTime date)
    cursor.execute(f"""
        SELECT `FGroup`, DATE(`FirstIntegrDateTime`) as outflow_date, COUNT(*) as cnt
        FROM tbl_ElvisSR
        WHERE {BUG_ZERO_WHERE}
          AND `Rejected` = 'N'
          AND DATE(`FirstIntegrDateTime`) >= %s
        GROUP BY `FGroup`, DATE(`FirstIntegrDateTime`)
    """, (earliest,))
    daily_outflow = {}
    for d in dates:
        daily_outflow[d] = {dom: 0 for dom in TEMPLATE_DOMAINS}
    for r in cursor.fetchall():
        d = r["outflow_date"]
        if isinstance(d, datetime):
            d = d.date()
        if d in daily_outflow:
            domain = map_domain(r["FGroup"])
            daily_outflow[d][domain] = daily_outflow[d].get(domain, 0) + r["cnt"]

    # 3b. Rejected tickets (moved to Concluding, use LastChangeDateTime as reject date)
    cursor.execute(f"""
        SELECT `FGroup`, DATE(`LastChangeDateTime`) as outflow_date, COUNT(*) as cnt
        FROM tbl_ElvisSR
        WHERE {BUG_ZERO_WHERE}
          AND `Rejected` = 'Y'
          AND DATE(`LastChangeDateTime`) >= %s
        GROUP BY `FGroup`, DATE(`LastChangeDateTime`)
    """, (earliest,))
    for r in cursor.fetchall():
        d = r["outflow_date"]
        if isinstance(d, datetime):
            d = d.date()
        if d in daily_outflow:
            domain = map_domain(r["FGroup"])
            daily_outflow[d][domain] = daily_outflow[d].get(domain, 0) + r["cnt"]

    return dates, total_open, daily_inflow, daily_outflow


def generate_excel(dates, total_open, daily_inflow, daily_outflow):
    """Generate Excel in template format."""
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    except ImportError:
        import subprocess
        subprocess.check_call([sys.executable, "-m", "pip", "install", "openpyxl", "-q"])
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    wb = openpyxl.load_workbook(
        os.path.join(os.path.expanduser("~"), "Downloads", "Template.xlsx")
    )
    ws = wb["Sheet1"]

    # Row 5: date headers (columns G, I, K, M, O, Q, S, U, W)
    date_cols = [7, 9, 11, 13, 15, 17, 19, 21, 23]  # G, I, K, M, O, Q, S, U, W
    for i, (col_idx, d) in enumerate(zip(date_cols, dates)):
        ws.cell(row=5, column=col_idx, value=d.strftime("%d-%b-%Y"))
        ws.cell(row=5, column=col_idx).font = Font(bold=True, size=9)
        ws.cell(row=5, column=col_idx).alignment = Alignment(horizontal="center")

    # Clear old template rows (8-35) and rewrite with actual FGroups
    for row_idx in range(8, 36):
        for col_idx in range(5, 25):
            ws.cell(row=row_idx, column=col_idx).value = None

    # Row 8 onwards: domain data (E=FGroup, F=total open, G/H=inflow/outflow day1, ...)
    for row_idx, domain in enumerate(TEMPLATE_DOMAINS, 8):
        ws.cell(row=row_idx, column=5, value=domain)
        # F: Total Open
        ws.cell(row=row_idx, column=6, value=total_open.get(domain, 0))

        # Daily inflow/outflow
        for day_idx, d in enumerate(dates):
            inflow_col = date_cols[day_idx]      # G, I, K, ...
            outflow_col = date_cols[day_idx] + 1  # H, J, L, ...
            inflow_val = daily_inflow.get(d, {}).get(domain, 0)
            outflow_val = daily_outflow.get(d, {}).get(domain, 0)
            ws.cell(row=row_idx, column=inflow_col, value=inflow_val)
            ws.cell(row=row_idx, column=outflow_col, value=outflow_val)

    # Grand Total row = after last domain
    grand_row = 8 + len(TEMPLATE_DOMAINS)
    ws.cell(row=grand_row, column=5, value="Grand Total")
    ws.cell(row=grand_row, column=5).font = Font(bold=True)
    ws.cell(row=grand_row, column=6, value=sum(total_open.values()))
    ws.cell(row=grand_row, column=6).font = Font(bold=True)
    for day_idx, d in enumerate(dates):
        inflow_col = date_cols[day_idx]
        outflow_col = date_cols[day_idx] + 1
        total_in = sum(daily_inflow.get(d, {}).values())
        total_out = sum(daily_outflow.get(d, {}).values())
        ws.cell(row=grand_row, column=inflow_col, value=total_in)
        ws.cell(row=grand_row, column=inflow_col).font = Font(bold=True)
        ws.cell(row=grand_row, column=outflow_col, value=total_out)
        ws.cell(row=grand_row, column=outflow_col).font = Font(bold=True)

    # ── Summary Sheet ──
    if "Summary" in wb.sheetnames:
        del wb["Summary"]
    ws2 = wb.create_sheet("Summary")

    title_font = Font(bold=True, size=14, color="1F4E79")
    header_font = Font(bold=True, size=11, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    yellow_fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")

    ws2.cell(row=1, column=1, value="DA2.8 Bug Zero — Daily Progress Report").font = title_font
    ws2.cell(row=2, column=1, value=f"Report Date: {dates[0].strftime('%d-%b-%Y (%A)')}")
    
    today_total = sum(total_open.values())
    may_end = date(2026, 5, 31)
    days_left = (may_end - dates[0]).days
    
    ws2.cell(row=3, column=1, value=f"Total Open Tickets: {today_total}")
    ws2.cell(row=3, column=1).font = Font(bold=True, size=12, color="FF0000" if today_total > 0 else "008000")
    ws2.cell(row=4, column=1, value=f"Days Remaining to May End: {days_left}")
    ws2.cell(row=4, column=1).font = Font(bold=True, size=12)
    
    if days_left > 0:
        daily_target = today_total / days_left
        ws2.cell(row=5, column=1, value=f"Required Daily Closure Rate: {daily_target:.0f} tickets/day")
        ws2.cell(row=5, column=1).font = Font(bold=True, size=11)

    # Today's inflow vs outflow
    today_in = sum(daily_inflow.get(dates[0], {}).values())
    today_out = sum(daily_outflow.get(dates[0], {}).values())
    net = today_out - today_in
    ws2.cell(row=7, column=1, value="Today's Summary").font = Font(bold=True, size=12)
    ws2.cell(row=8, column=1, value="Inflow")
    ws2.cell(row=8, column=2, value=today_in)
    ws2.cell(row=9, column=1, value="Outflow (Integrated + Rejected)")
    ws2.cell(row=9, column=2, value=today_out)
    ws2.cell(row=10, column=1, value="Net Reduction")
    ws2.cell(row=10, column=2, value=net)
    ws2.cell(row=10, column=2).fill = green_fill if net > 0 else red_fill

    # Top 5 domains by open count
    ws2.cell(row=12, column=1, value="Top Domains by Open Count").font = Font(bold=True, size=12)
    ws2.cell(row=13, column=1, value="Domain").font = header_font
    ws2.cell(row=13, column=1).fill = header_fill
    ws2.cell(row=13, column=2, value="Open").font = header_font
    ws2.cell(row=13, column=2).fill = header_fill
    ws2.cell(row=13, column=3, value="% of Total").font = header_font
    ws2.cell(row=13, column=3).fill = header_fill

    sorted_domains = sorted(total_open.items(), key=lambda x: -x[1])
    for i, (dom, cnt) in enumerate(sorted_domains[:10], 14):
        ws2.cell(row=i, column=1, value=dom)
        ws2.cell(row=i, column=2, value=cnt)
        pct = (cnt / today_total * 100) if today_total else 0
        ws2.cell(row=i, column=3, value=f"{pct:.1f}%")

    # Trend (last 7 days net)
    trend_row = 14 + min(10, len(sorted_domains)) + 2
    ws2.cell(row=trend_row, column=1, value="Daily Trend (Last 9 Days)").font = Font(bold=True, size=12)
    ws2.cell(row=trend_row+1, column=1, value="Date").font = header_font
    ws2.cell(row=trend_row+1, column=1).fill = header_fill
    ws2.cell(row=trend_row+1, column=2, value="Inflow").font = header_font
    ws2.cell(row=trend_row+1, column=2).fill = header_fill
    ws2.cell(row=trend_row+1, column=3, value="Outflow").font = header_font
    ws2.cell(row=trend_row+1, column=3).fill = header_fill
    ws2.cell(row=trend_row+1, column=4, value="Net").font = header_font
    ws2.cell(row=trend_row+1, column=4).fill = header_fill

    for i, d in enumerate(dates, trend_row + 2):
        day_in = sum(daily_inflow.get(d, {}).values())
        day_out = sum(daily_outflow.get(d, {}).values())
        day_net = day_out - day_in
        ws2.cell(row=i, column=1, value=d.strftime("%d-%b (%a)"))
        ws2.cell(row=i, column=2, value=day_in)
        ws2.cell(row=i, column=3, value=day_out)
        ws2.cell(row=i, column=4, value=day_net)
        ws2.cell(row=i, column=4).fill = green_fill if day_net > 0 else (red_fill if day_net < 0 else yellow_fill)

    # Auto-width
    for col in ws2.columns:
        max_len = 0
        col_letter = col[0].column_letter
        for cell in col:
            if cell.value:
                max_len = max(max_len, len(str(cell.value)))
        ws2.column_dimensions[col_letter].width = min(max_len + 3, 40)

    # ── Email Body Sheet ──
    if "Email Body" in wb.sheetnames:
        del wb["Email Body"]
    ws3 = wb.create_sheet("Email Body")
    
    email_lines = [
        f"Subject: DA2.8 Bug Zero Progress — {dates[0].strftime('%d-%b-%Y')} | {today_total} Open | {days_left} Days Left",
        "",
        f"Hi Team,",
        "",
        f"Please find the Bug Zero daily status update below:",
        "",
        f"📊 Overall Status:",
        f"  • Total Open Tickets: {today_total}",
        f"  • Days Remaining (till May 31): {days_left}",
        f"  • Required Daily Closure Rate: {today_total / max(days_left, 1):.0f} tickets/day",
        "",
        f"📈 Today's Movement:",
        f"  • Inflow: {today_in}",
        f"  • Outflow (Integrated + Rejected): {today_out}",
        f"  • Net: {'↓' if net > 0 else '↑'} {abs(net)} {'(GOOD — reducing)' if net > 0 else '(ALERT — increasing)'}",
        "",
        f"🔝 Top 5 Domains by Open Count:",
    ]
    for dom, cnt in sorted_domains[:5]:
        pct = (cnt / today_total * 100) if today_total else 0
        email_lines.append(f"  • {dom}: {cnt} ({pct:.0f}%)")
    
    email_lines.extend([
        "",
        "Action Required:",
        "  → Focus closure on top domains",
        "  → Escalate blockers immediately",
        "  → Review rejected tickets for re-processing",
        "",
        "Detailed report attached.",
        "",
        "Regards,",
        "Defect Management Team",
    ])

    for i, line in enumerate(email_lines, 1):
        ws3.cell(row=i, column=1, value=line)
    ws3.column_dimensions["A"].width = 80

    # Save
    downloads = os.path.join(os.path.expanduser("~"), "Downloads")
    timestamp = dates[0].strftime("%Y%m%d")
    filepath = os.path.join(downloads, f"DA28_BugZero_Daily_{timestamp}.xlsx")
    wb.save(filepath)
    return filepath, email_lines


def main():
    print("=" * 60)
    print("DA2.8 Bug Zero — Daily Report Generator")
    print("=" * 60)

    conn = get_connection()
    cursor = conn.cursor(dictionary=True)

    print("\nFetching data from Elvis DB...")
    dates, total_open, daily_inflow, daily_outflow = fetch_bug_zero_data(cursor, num_days=9)

    cursor.close()
    conn.close()

    today_total = sum(total_open.values())
    today = dates[0]
    may_end = date(2026, 5, 31)
    days_left = (may_end - today).days

    print(f"\nReport Date: {today.strftime('%d-%b-%Y (%A)')}")
    print(f"Total Open: {today_total}")
    print(f"Days to May 31: {days_left}")
    print(f"Required closure rate: {today_total / max(days_left, 1):.0f} tickets/day")

    today_in = sum(daily_inflow.get(today, {}).values())
    today_out = sum(daily_outflow.get(today, {}).values())
    print(f"\nToday: Inflow={today_in}, Outflow={today_out}, Net={'↓' if today_out >= today_in else '↑'}{abs(today_out - today_in)}")

    print("\nTop domains:")
    for dom, cnt in sorted(total_open.items(), key=lambda x: -x[1])[:5]:
        print(f"  {dom:<25} {cnt}")

    # Generate Excel
    filepath, email_lines = generate_excel(dates, total_open, daily_inflow, daily_outflow)
    print(f"\nExcel exported: {filepath}")

    # Print email body
    print("\n" + "=" * 60)
    print("EMAIL BODY (copy-paste):")
    print("=" * 60)
    for line in email_lines:
        print(line)


if __name__ == "__main__":
    main()
