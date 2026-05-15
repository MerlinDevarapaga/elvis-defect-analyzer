"""
Daily Triage Tool — Fetches recent inflow tickets from Elvis DB,
categorizes/prioritizes them, checks for priority misassignment using
the Harman Bug Taxonomy matrix, and suggests actions.

Bug Taxonomy:
  Final Priority = f(Severity Rank, Frequency, Recovery Condition)
  - Severity Ranks: Top, A, B, C
  - Frequency: Always (7-10/10), Sometimes (3-6/10), Rare (1-2/10)
  - Recovery: Difficult (cold boot/reflash), Easy (warm boot/S2R), Automatic

Output: Excel file with triage summary + priority mismatch flags.

Usage:
    python scripts/daily_triage.py
    python scripts/daily_triage.py --days 3
    python scripts/daily_triage.py --excel
    python scripts/daily_triage.py --excel output.xlsx --days 5
"""
import os
import sys
import io
import re
import argparse
from datetime import datetime, date, timedelta
from collections import defaultdict

sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8", errors="replace")

from dotenv import load_dotenv
import mysql.connector

# Load .env from workspace root
_script_dir = os.path.dirname(os.path.abspath(__file__))
for _env in [os.path.join(_script_dir, "..", ".env"), os.path.join(_script_dir, ".env")]:
    if os.path.exists(_env):
        load_dotenv(_env)
        break

# ---------------------------------------------------------------------------
# Configuration
# ---------------------------------------------------------------------------
PROJECT_ID = "MSIL_DA2.8"
DAYS_BACK = 2  # default: last 2 days of inflow
PRE_INTEGRATING_STEPS = ("Categorizing", "Processing", "Reproduction")

# Priority mapping: DB value → taxonomy label
PRIORITY_DB_TO_LABEL = {
    "top": "Top",
    "A(1)": "A",
    "B(2)": "B",
    "C(3)": "C",
}
PRIORITY_LABEL_TO_DB = {v: k for k, v in PRIORITY_DB_TO_LABEL.items()}

# Occurrence mapping: DB value → taxonomy frequency
OCCURRENCE_TO_FREQUENCY = {
    "Always": "Always",
    "Sometimes": "Sometimes",
    "Once": "Rare",
    "Rare": "Rare",
    "": "Unknown",
    None: "Unknown",
}

# FGroup display names
FGROUP_DISPLAY = {
    "Media": "Media",
    "Bluetooth": "BT",
    "Projection": "Projection",
    "Audio": "Audio",
    "IOC": "IOC",
    "Camera": "Camera",
    "WiFi": "Wifi",
    "Systems - Core": "Systems Core",
    "Systems - Infra": "Sys Infra",
    "Systems - SWU Software Update": "SWUP",
    "USB": "USB",
    "SVS": "SVS",
}

# ---------------------------------------------------------------------------
# Keyword-based FGroup inference from title/description
# ---------------------------------------------------------------------------
FGROUP_KEYWORDS = {
    "Projection": [r"projection", r"carplay", r"android\s*auto", r"\bAA\b", r"\bCP\b", r"mirrorlink", r"WPC"],
    "BT": [r"bluetooth", r"\bBT\b", r"pairing", r"\bA2DP\b", r"\bHFP\b", r"phone\s*book", r"contacts?"],
    "Media": [r"\bmedia\b", r"\bDAB\b", r"\bFM\b", r"\bAM\b", r"tuner", r"radio", r"playback", r"track", r"album"],
    "USB": [r"\bUSB\b", r"pendrive", r"mass\s*storage"],
    "Audio": [r"\baudio\b", r"volume", r"sound", r"speaker", r"amplifier", r"\bAMP\b", r"mute"],
    "Camera": [r"camera", r"\bRVC\b", r"\bRVS\b", r"rear\s*view", r"surround\s*view"],
    "Wifi": [r"wi-?fi", r"hotspot", r"wireless\s*lan", r"WLAN"],
    "SVS": [r"\bSVS\b", r"surround\s*view\s*system"],
    "SWUP": [r"\bOTA\b", r"\bSWUP\b", r"software\s*update", r"\bSW-OTA\b", r"FOTA"],
    "IOC": [r"\bIOC\b", r"diagnostic", r"\bDID\b", r"\bNRC\b"],
    "Systems Core": [r"system\s*core", r"\bHU\b", r"head\s*unit", r"boot", r"crash", r"restart", r"display", r"status\s*bar", r"factory\s*mode", r"dealer\s*mode"],
    "Sys Infra": [r"sys\s*infra", r"infra", r"diagnostic\s*checklist"],
    "Security": [r"security", r"intrusion", r"authentication", r"\bIDS\b", r"firewall"],
    "HMI IVI": [r"\bHMI\b", r"\bIVI\b", r"settings", r"gallery", r"apps?", r"screen", r"navigation", r"keyboard", r"language"],
    "Tuner": [r"tuner", r"multiplex", r"\bDAB\b", r"\bFM\b"],
}

# ---------------------------------------------------------------------------
# Bug Taxonomy Classification Matrix
# Keys: (severity_rank, frequency, recovery) → final_priority
# Severity: Top, A, B, C
# Frequency: Always, Sometimes, Rare
# Recovery: Difficult, Easy, Automatic
# ---------------------------------------------------------------------------
TAXONOMY_MATRIX = {
    # Top severity: always Top regardless of frequency/recovery
    ("Top", "Always", "Difficult"): "Top",
    ("Top", "Always", "Easy"): "Top",
    ("Top", "Always", "Automatic"): "Top",
    ("Top", "Sometimes", "Difficult"): "Top",
    ("Top", "Sometimes", "Easy"): "Top",
    ("Top", "Sometimes", "Automatic"): "Top",
    ("Top", "Rare", "Difficult"): "Top",
    ("Top", "Rare", "Easy"): "Top",
    ("Top", "Rare", "Automatic"): "Top",
    # A severity
    ("A", "Always", "Difficult"): "Top",
    ("A", "Always", "Easy"): "A",
    ("A", "Always", "Automatic"): "A",
    ("A", "Sometimes", "Difficult"): "A",
    ("A", "Sometimes", "Easy"): "B",
    ("A", "Sometimes", "Automatic"): "B",
    ("A", "Rare", "Difficult"): "A",
    ("A", "Rare", "Easy"): "B",
    ("A", "Rare", "Automatic"): "C",
    # B severity
    ("B", "Always", "Difficult"): "B",
    ("B", "Always", "Easy"): "B",
    ("B", "Always", "Automatic"): "B",
    ("B", "Sometimes", "Difficult"): "B",
    ("B", "Sometimes", "Easy"): "B",
    ("B", "Sometimes", "Automatic"): "C",
    ("B", "Rare", "Difficult"): "B",
    ("B", "Rare", "Easy"): "C",
    ("B", "Rare", "Automatic"): "C",
    # C severity: always C
    ("C", "Always", "Difficult"): "C",
    ("C", "Always", "Easy"): "C",
    ("C", "Always", "Automatic"): "C",
    ("C", "Sometimes", "Difficult"): "C",
    ("C", "Sometimes", "Easy"): "C",
    ("C", "Sometimes", "Automatic"): "C",
    ("C", "Rare", "Difficult"): "C",
    ("C", "Rare", "Easy"): "C",
    ("C", "Rare", "Automatic"): "C",
}

# ---------------------------------------------------------------------------
# Keyword-based heuristics for severity inference from Title/Description
# ---------------------------------------------------------------------------
TOP_KEYWORDS = [
    r"regulat", r"safety", r"walk\s*home", r"privacy", r"legal",
    r"unusable", r"hu\s*restart", r"head\s*unit\s*restart",
    r"complete\s*loss", r"complete\s*functionality\s*loss",
    r"brick", r"no\s*boot", r"does\s*not\s*start",
]

A_KEYWORDS = [
    r"system\s*reset", r"temporary\s*freeze", r"temp\s*freeze",
    r"temporary\s*crash", r"audio\s*loss", r"no\s*audio",
    r"partial\s*functionality\s*loss", r"black\s*screen",
    r"white\s*screen", r"vehicle\s*config", r"freeze",
    r"crash", r"reboot", r"reset", r"not\s*working",
    r"functionality\s*not\s*available", r"feature\s*not\s*working",
]

B_KEYWORDS = [
    r"minor\s*display", r"sound\s*glitch", r"noise",
    r"sub\s*feature\s*fail", r"partial\s*degradation",
    r"truncated\s*string", r"display\s*issue", r"glitch",
    r"intermittent", r"delay", r"slow\s*response",
    r"ui\s*issue", r"alignment", r"overlap",
]

C_KEYWORDS = [
    r"minor\s*graphical", r"grammatical", r"spelling",
    r"uniformity", r"cosmetic", r"nice\s*to\s*have",
    r"personal\s*preference", r"negligible", r"barely\s*noticeable",
    r"typo", r"font\s*size", r"color\s*mismatch",
]

# Recovery inference keywords
DIFFICULT_RECOVERY_KEYWORDS = [
    r"ignition\s*cycle", r"cold\s*boot", r"reflash",
    r"initialization", r"factory\s*reset", r"does\s*not\s*recover",
    r"power\s*cycle", r"ecu\s*reset",
]

EASY_RECOVERY_KEYWORDS = [
    r"warm\s*boot", r"s2r", r"standby", r"user\s*re-?operation",
    r"source\s*change", r"reconnect", r"re-?pair",
    r"re-?connect", r"toggle",
]

AUTO_RECOVERY_KEYWORDS = [
    r"auto\w*\s*recov", r"resolves\s*itself", r"self\s*correct",
    r"goes\s*away", r"disappears\s*after", r"recovers\s*auto",
    r"within\s*seconds", r"momentary",
]


def get_connection():
    return mysql.connector.connect(
        host=os.getenv("ELVIS_DB_HOST"),
        user=os.getenv("ELVIS_DB_USER"),
        password=os.getenv("ELVIS_DB_PASSWORD"),
        database=os.getenv("ELVIS_DB_NAME"),
        port=int(os.getenv("ELVIS_DB_PORT", 3306)),
    )


def fetch_recent_inflow(days_back, priority_filter=None, pre_integrating_only=False):
    """Fetch MSIL DA2.8 tickets created in the last N days.

    Args:
        days_back: Number of days to look back.
        priority_filter: List of PriorityID values to filter (e.g. ['top', 'A(1)'])
        pre_integrating_only: If True, only tickets in pre-integrating steps.
    """
    conn = get_connection()
    cursor = conn.cursor(dictionary=True)
    start_date = date.today() - timedelta(days=days_back)

    query = """
        SELECT TicketID, Title, ProblemDescription, PriorityID, Occurance,
               FGroup, TicketStepID, EnterDateTime, SequenceOfActivity,
               Category, ProblemType, Owner, StateID
        FROM tbl_ElvisSR
        WHERE ProjectID = %s
          AND EnterDateTime >= %s
          AND IsDeleted = 'N'
    """
    params = [PROJECT_ID, start_date.strftime("%Y-%m-%d")]

    if priority_filter:
        placeholders = ", ".join(["%s"] * len(priority_filter))
        query += f"  AND PriorityID IN ({placeholders})\n"
        params.extend(priority_filter)

    if pre_integrating_only:
        placeholders = ", ".join(["%s"] * len(PRE_INTEGRATING_STEPS))
        query += f"  AND TicketStepID IN ({placeholders})\n"
        params.extend(PRE_INTEGRATING_STEPS)

    query += "ORDER BY EnterDateTime DESC"
    cursor.execute(query, params)
    rows = cursor.fetchall()
    cursor.close()
    conn.close()
    return rows


def infer_fgroup(title, description):
    """Infer expected FGroup from title and description keywords."""
    text = f"{title or ''} {description or ''}"
    text_lower = text.lower()

    scores = defaultdict(int)
    for fgroup, patterns in FGROUP_KEYWORDS.items():
        for pattern in patterns:
            if re.search(pattern, text, re.IGNORECASE):
                scores[fgroup] += 1

    if not scores:
        return "Unknown"

    # Return the FGroup with highest keyword matches
    return max(scores, key=scores.get)


def infer_severity(title, description):
    """Infer severity rank (Top/A/B/C) from title and description keywords."""
    text = f"{title or ''} {description or ''}".lower()

    for pattern in TOP_KEYWORDS:
        if re.search(pattern, text):
            return "Top"
    for pattern in A_KEYWORDS:
        if re.search(pattern, text):
            return "A"
    for pattern in B_KEYWORDS:
        if re.search(pattern, text):
            return "B"
    for pattern in C_KEYWORDS:
        if re.search(pattern, text):
            return "C"

    return "Unknown"


def infer_recovery(title, description):
    """Infer recovery condition from title and description keywords."""
    text = f"{title or ''} {description or ''}".lower()

    for pattern in DIFFICULT_RECOVERY_KEYWORDS:
        if re.search(pattern, text):
            return "Difficult"
    for pattern in EASY_RECOVERY_KEYWORDS:
        if re.search(pattern, text):
            return "Easy"
    for pattern in AUTO_RECOVERY_KEYWORDS:
        if re.search(pattern, text):
            return "Automatic"

    return "Unknown"


def get_expected_priority(severity, frequency, recovery):
    """Look up the expected priority from the taxonomy matrix."""
    if severity == "Unknown" or frequency == "Unknown" or recovery == "Unknown":
        return None
    key = (severity, frequency, recovery)
    return TAXONOMY_MATRIX.get(key)


def suggest_action(ticket, expected_priority, assigned_label):
    """Suggest triage action based on analysis."""
    actions = []
    step = str(ticket.get("TicketStepID") or "").strip()

    # Priority mismatch
    if expected_priority and assigned_label and expected_priority != assigned_label:
        priority_order = {"Top": 0, "A": 1, "B": 2, "C": 3}
        exp_rank = priority_order.get(expected_priority, 99)
        asg_rank = priority_order.get(assigned_label, 99)

        if exp_rank < asg_rank:
            actions.append(f"ESCALATE: Expected '{expected_priority}' but assigned '{assigned_label}' — priority may be too low")
        else:
            actions.append(f"REVIEW: Expected '{expected_priority}' but assigned '{assigned_label}' — priority may be too high")

    # Step-based suggestions
    if step == "Categorizing":
        actions.append("ASSIGN: Ticket needs categorization and domain assignment")
    elif step == "Processing":
        if not ticket.get("Owner"):
            actions.append("ASSIGN: No owner assigned yet")

    # SOA check for high priority
    soa = str(ticket.get("SequenceOfActivity") or "").strip()
    if assigned_label in ("Top", "A") and soa not in ("1-Urgent", "2-Very High"):
        actions.append(f"SOA-FLAG: Priority {assigned_label} but SOA is '{soa or 'Not Set'}' — consider elevating SOA")

    if not actions:
        actions.append("OK: No action needed")

    return "; ".join(actions)


def analyze_tickets(tickets):
    """Analyze all tickets and return triage results."""
    results = []

    for t in tickets:
        tid = t.get("TicketID")
        title = str(t.get("Title") or "").strip()
        desc = str(t.get("ProblemDescription") or "").strip()
        priority_db = str(t.get("PriorityID") or "").strip()
        occurrence_db = str(t.get("Occurance") or "").strip()
        fgroup = str(t.get("FGroup") or "Unknown").strip()
        step = str(t.get("TicketStepID") or "").strip()
        soa = str(t.get("SequenceOfActivity") or "").strip()
        entered = t.get("EnterDateTime")

        # Map DB values to taxonomy labels
        assigned_label = PRIORITY_DB_TO_LABEL.get(priority_db, priority_db)
        frequency = OCCURRENCE_TO_FREQUENCY.get(occurrence_db, "Unknown")

        # Infer severity and recovery from text
        inferred_severity = infer_severity(title, desc)
        inferred_recovery = infer_recovery(title, desc)

        # Get expected priority from taxonomy matrix
        expected_priority = get_expected_priority(inferred_severity, frequency, inferred_recovery)

        # Determine mismatch
        mismatch = ""
        if expected_priority and assigned_label:
            if expected_priority != assigned_label:
                mismatch = f"{assigned_label} → {expected_priority}"

        # Suggest action
        action = suggest_action(t, expected_priority, assigned_label)

        # Infer expected FGroup
        expected_fgroup = infer_fgroup(title, desc)
        assigned_fgroup = FGROUP_DISPLAY.get(fgroup, fgroup)
        fgroup_mismatch = ""
        if expected_fgroup != "Unknown" and expected_fgroup != assigned_fgroup:
            fgroup_mismatch = f"{assigned_fgroup} → {expected_fgroup}"

        results.append({
            "TicketID": tid,
            "Title": title[:100],
            "FGroup": assigned_fgroup,
            "Expected FGroup": expected_fgroup,
            "FGroup Mismatch": fgroup_mismatch,
            "Step": step,
            "Assigned Priority": assigned_label,
            "Occurrence": occurrence_db,
            "Inferred Severity": inferred_severity,
            "Inferred Recovery": inferred_recovery,
            "Expected Priority": expected_priority or "N/A",
            "Priority Mismatch": mismatch,
            "SOA": soa,
            "Suggested Action": action,
            "Entered": entered.strftime("%Y-%m-%d") if entered else "",
        })

    return results


def print_summary(results, days_back):
    """Print triage summary to console."""
    total = len(results)
    mismatches = [r for r in results if r["Priority Mismatch"]]
    escalations = [r for r in results if "ESCALATE" in r["Suggested Action"]]
    needs_assign = [r for r in results if "ASSIGN" in r["Suggested Action"]]

    # Domain breakdown
    domain_counts = defaultdict(int)
    for r in results:
        domain_counts[r["FGroup"]] += 1

    print()
    print("=" * 80)
    print(f"  DAILY TRIAGE REPORT — MSIL DA2.8 (Last {days_back} days inflow)")
    print(f"  Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')}")
    print("=" * 80)

    print(f"\n  Total Inflow Tickets: {total}")
    print(f"  Priority Mismatches:  {len(mismatches)}")
    print(f"  Need Escalation:      {len(escalations)}")
    print(f"  Need Assignment:      {len(needs_assign)}")

    # Domain breakdown
    print(f"\n  --- Domain Breakdown ---")
    for fg, cnt in sorted(domain_counts.items(), key=lambda x: x[1], reverse=True):
        print(f"    {fg:<20} {cnt}")

    # Priority mismatches detail
    if mismatches:
        print(f"\n  --- Priority Mismatches ({len(mismatches)}) ---")
        print(f"  {'Ticket':<10} {'Domain':<14} {'Assigned':<10} {'Expected':<10} {'Occurrence':<12} Title")
        print(f"  {'-'*10} {'-'*14} {'-'*10} {'-'*10} {'-'*12} {'-'*40}")
        for r in mismatches:
            print(f"  {r['TicketID']:<10} {r['FGroup']:<14} {r['Assigned Priority']:<10} "
                  f"{r['Expected Priority']:<10} {r['Occurrence']:<12} {r['Title'][:45]}")

    # Escalation needed
    if escalations:
        print(f"\n  --- Need Escalation ({len(escalations)}) ---")
        for r in escalations:
            print(f"  [{r['TicketID']}] {r['FGroup']} — {r['Priority Mismatch']} — {r['Title'][:50]}")

    # Action items
    if needs_assign:
        print(f"\n  --- Need Assignment ({len(needs_assign)}) ---")
        for r in needs_assign:
            print(f"  [{r['TicketID']}] {r['FGroup']} — Step: {r['Step']} — {r['Title'][:50]}")

    print("\n" + "=" * 80)


def export_excel(results, output_path, days_back):
    """Export triage results to Excel with color coding."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    wb = Workbook()
    ws = wb.active
    ws.title = "Daily Triage"

    # Header style
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
    cell_align = Alignment(vertical="top", wrap_text=True)
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    # Conditional fills
    red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")      # Escalation
    orange_fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")    # Review/mismatch
    green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")     # OK

    # Headers
    headers = [
        "Ticket ID", "Title", "Domain", "Expected FGroup", "FGroup Mismatch",
        "Step", "Assigned Priority", "Occurrence",
        "Inferred Severity", "Inferred Recovery",
        "Expected Priority", "Priority Mismatch", "SOA",
        "Suggested Action", "Entered",
    ]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = cell_align
        cell.border = thin_border

    # Data rows
    for row_idx, r in enumerate(results, 2):
        values = [
            r["TicketID"], r["Title"], r["FGroup"], r["Expected FGroup"],
            r["FGroup Mismatch"], r["Step"],
            r["Assigned Priority"], r["Occurrence"],
            r["Inferred Severity"], r["Inferred Recovery"],
            r["Expected Priority"], r["Priority Mismatch"],
            r["SOA"], r["Suggested Action"], r["Entered"],
        ]
        for col, val in enumerate(values, 1):
            cell = ws.cell(row=row_idx, column=col, value=val)
            cell.alignment = cell_align
            cell.border = thin_border

        # Color coding based on action
        action = r["Suggested Action"]
        if "ESCALATE" in action:
            for col in range(1, len(headers) + 1):
                ws.cell(row=row_idx, column=col).fill = red_fill
        elif r["Priority Mismatch"]:
            for col in range(1, len(headers) + 1):
                ws.cell(row=row_idx, column=col).fill = orange_fill
        elif "OK" in action:
            for col in range(1, len(headers) + 1):
                ws.cell(row=row_idx, column=col).fill = green_fill

    # Column widths
    col_widths = [10, 50, 14, 16, 18, 14, 12, 12, 14, 14, 12, 16, 14, 50, 12]
    for col, w in enumerate(col_widths, 1):
        ws.column_dimensions[ws.cell(row=1, column=col).column_letter].width = w

    # Freeze top row
    ws.freeze_panes = "A2"

    # --- Summary sheet ---
    ws_sum = wb.create_sheet("Summary")
    ws_sum.cell(row=1, column=1, value=f"Daily Triage Report — MSIL DA2.8")
    ws_sum.cell(row=1, column=1).font = Font(bold=True, size=14)
    ws_sum.cell(row=2, column=1, value=f"Period: Last {days_back} days | Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')}")

    total = len(results)
    mismatches = len([r for r in results if r["Priority Mismatch"]])
    escalations = len([r for r in results if "ESCALATE" in r["Suggested Action"]])
    needs_assign = len([r for r in results if "ASSIGN" in r["Suggested Action"]])

    summary_data = [
        ("Total Inflow", total),
        ("Priority Mismatches", mismatches),
        ("Need Escalation", escalations),
        ("Need Assignment", needs_assign),
    ]
    for i, (label, val) in enumerate(summary_data, 4):
        ws_sum.cell(row=i, column=1, value=label)
        ws_sum.cell(row=i, column=2, value=val)

    # Domain breakdown
    ws_sum.cell(row=9, column=1, value="Domain Breakdown").font = Font(bold=True)
    domain_counts = defaultdict(int)
    for r in results:
        domain_counts[r["FGroup"]] += 1
    for i, (fg, cnt) in enumerate(sorted(domain_counts.items(), key=lambda x: x[1], reverse=True), 10):
        ws_sum.cell(row=i, column=1, value=fg)
        ws_sum.cell(row=i, column=2, value=cnt)

    wb.save(output_path)
    print(f"\n  Excel saved: {output_path}")


def main():
    parser = argparse.ArgumentParser(description="Daily Triage Tool for MSIL DA2.8")
    parser.add_argument("--days", type=int, default=DAYS_BACK,
                        help=f"Days of inflow to analyze (default: {DAYS_BACK})")
    parser.add_argument("--excel", nargs="?", const="auto",
                        help="Export to Excel (optionally specify path)")
    parser.add_argument("--priority", nargs="+",
                        help="Filter by priority: top A B C (maps to DB values)")
    parser.add_argument("--pre-integrating", action="store_true",
                        help="Only show tickets in pre-integrating steps (Categorizing/Processing/Reproduction)")
    args = parser.parse_args()

    # Map user-friendly priority labels to DB values
    priority_filter = None
    if args.priority:
        priority_filter = []
        for p in args.priority:
            p_upper = p.strip().capitalize()
            if p_upper in PRIORITY_LABEL_TO_DB:
                priority_filter.append(PRIORITY_LABEL_TO_DB[p_upper])
            else:
                priority_filter.append(p.strip())  # pass as-is

    prio_label = f" [Priority: {', '.join(args.priority)}]" if args.priority else ""
    step_label = " [Pre-Integrating only]" if args.pre_integrating else ""
    print(f"  Fetching MSIL DA2.8 inflow for last {args.days} days...{prio_label}{step_label}")
    tickets = fetch_recent_inflow(args.days, priority_filter, args.pre_integrating)

    if not tickets:
        print("  No tickets found in the specified period.")
        return

    print(f"  Analyzing {len(tickets)} tickets...")
    results = analyze_tickets(tickets)

    # Console summary
    print_summary(results, args.days)

    # Excel export
    if args.excel is not None:
        if args.excel == "auto":
            output_dir = os.path.join(_script_dir, "..", "docs", "output")
            os.makedirs(output_dir, exist_ok=True)
            output_path = os.path.join(
                output_dir,
                f"daily_triage_{date.today().strftime('%Y%m%d')}.xlsx"
            )
        else:
            output_path = args.excel
        export_excel(results, output_path, args.days)


if __name__ == "__main__":
    main()
