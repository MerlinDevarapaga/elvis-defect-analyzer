"""
Find duplicate defects within existing pre-integrating tickets.
Uses both Title and ProblemDescription for similarity matching.
Each ticket appears at most once as a duplicate (no double-counting).
"""
import os
import sys
import io
import re
from datetime import datetime, date
from difflib import SequenceMatcher
from collections import defaultdict

sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8", errors="replace")

from dotenv import load_dotenv
import mysql.connector

_script_dir = os.path.dirname(os.path.abspath(__file__))
for _env in [os.path.join(_script_dir, "..", ".env"), os.path.join(_script_dir, ".env")]:
    if os.path.exists(_env):
        load_dotenv(_env)
        break

PROJECT_ID = "MSIL_DA2.8"
PRE_INTEGRATING_STEPS = ("Categorizing", "Processing", "Reproduction")
TITLE_THRESHOLD = 0.65
DESC_THRESHOLD = 0.50
COMBINED_THRESHOLD = 0.60


def get_connection():
    return mysql.connector.connect(
        host=os.getenv("ELVIS_DB_HOST"),
        user=os.getenv("ELVIS_DB_USER"),
        password=os.getenv("ELVIS_DB_PASSWORD"),
        database=os.getenv("ELVIS_DB_NAME"),
        port=int(os.getenv("ELVIS_DB_PORT", 3306)),
    )


def normalize(text):
    if not text:
        return ""
    t = text.lower().strip()
    t = re.sub(r"[^a-z0-9 ]", " ", t)
    t = re.sub(r"\s+", " ", t)
    return t


def similarity(a, b):
    if not a or not b:
        return 0.0
    return SequenceMatcher(None, a, b).ratio()


def main():
    conn = get_connection()
    cursor = conn.cursor(dictionary=True)

    placeholders = ", ".join(["%s"] * len(PRE_INTEGRATING_STEPS))
    query = f"""
        SELECT TicketID, Title, ProblemDescription, EnterDateTime, TicketStepID,
               FGroup, PriorityID, Occurance, MasterType
        FROM tbl_ElvisSR
        WHERE ProjectID = %s
          AND TicketStepID IN ({placeholders})
          AND FGroup NOT LIKE '%%HMI%%'
          AND IsDeleted = 'N'
        ORDER BY FGroup, EnterDateTime
    """
    cursor.execute(query, [PROJECT_ID] + list(PRE_INTEGRATING_STEPS))
    tickets = cursor.fetchall()
    cursor.close()
    conn.close()

    today = date.today()
    print("Total pre-integrating Non-HMI tickets: %d" % len(tickets))

    # Normalize titles and descriptions
    for t in tickets:
        t["_norm_title"] = normalize(t["Title"])
        desc = t.get("ProblemDescription") or ""
        # Trim description to first 500 chars for speed
        t["_norm_desc"] = normalize(desc[:500])
        ed = t["EnterDateTime"]
        if isinstance(ed, datetime):
            ed = ed.date()
        t["_age"] = (today - ed).days

    # Compare all pairs within same FGroup first (most likely duplicates)
    # Then cross-group for high title similarity
    by_fgroup = defaultdict(list)
    for t in tickets:
        fg = (t.get("FGroup") or "Unknown").strip()
        by_fgroup[fg].append(t)

    matches = []
    claimed = set()  # ticket IDs already used as a duplicate

    print("Comparing within FGroups...")
    for fg, group in by_fgroup.items():
        n = len(group)
        for i in range(n):
            for j in range(i + 1, n):
                a = group[i]
                b = group[j]
                title_sim = similarity(a["_norm_title"], b["_norm_title"])
                desc_sim = similarity(a["_norm_desc"], b["_norm_desc"]) if a["_norm_desc"] and b["_norm_desc"] else 0.0
                combined = title_sim * 0.6 + desc_sim * 0.4 if desc_sim > 0 else title_sim

                if title_sim >= TITLE_THRESHOLD or (combined >= COMBINED_THRESHOLD and desc_sim >= DESC_THRESHOLD):
                    # Older ticket is primary, newer is duplicate
                    if a["_age"] >= b["_age"]:
                        primary, dup = a, b
                    else:
                        primary, dup = b, a
                    matches.append({
                        "primary_id": primary["TicketID"],
                        "primary_title": primary["Title"],
                        "primary_date": str(primary["EnterDateTime"])[:10],
                        "primary_step": primary["TicketStepID"],
                        "primary_fg": primary.get("FGroup", ""),
                        "primary_pri": primary.get("PriorityID", ""),
                        "primary_age": primary["_age"],
                        "dup_id": dup["TicketID"],
                        "dup_title": dup["Title"],
                        "dup_date": str(dup["EnterDateTime"])[:10],
                        "dup_step": dup["TicketStepID"],
                        "dup_fg": dup.get("FGroup", ""),
                        "dup_pri": dup.get("PriorityID", ""),
                        "dup_age": dup["_age"],
                        "title_sim": title_sim,
                        "desc_sim": desc_sim,
                        "combined_sim": combined,
                    })

    # Sort by combined similarity descending
    matches.sort(key=lambda x: -x["combined_sim"])

    # Deduplicate: each ticket can be a duplicate only ONCE
    final = []
    for m in matches:
        if m["dup_id"] not in claimed:
            claimed.add(m["dup_id"])
            final.append(m)

    print("Potential duplicate pairs found: %d" % len(final))
    print()

    # Print summary
    print("=" * 140)
    print("DUPLICATE DEFECTS WITHIN PRE-INTEGRATING TICKETS")
    print("=" * 140)

    for i, m in enumerate(final[:50], 1):
        tsim = "%.0f%%" % (m["title_sim"] * 100)
        dsim = "%.0f%%" % (m["desc_sim"] * 100)
        csim = "%.0f%%" % (m["combined_sim"] * 100)
        print()
        print("%2d. [Combined=%s Title=%s Desc=%s]" % (i, csim, tsim, dsim))
        print("    PRIMARY: %s (%s) %dd [%s] [%s] %s" % (
            m["primary_id"], m["primary_date"], m["primary_age"],
            m["primary_step"], m["primary_fg"], m["primary_title"][:80]))
        print("    DUPLICATE: %s (%s) %dd [%s] [%s] %s" % (
            m["dup_id"], m["dup_date"], m["dup_age"],
            m["dup_step"], m["dup_fg"], m["dup_title"][:80]))

    # Excel export
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

    wb = Workbook()
    ws = wb.active
    ws.title = "Duplicate Defects"

    hf = Font(bold=True, color="FFFFFF", size=10)
    hfill = PatternFill(start_color="2E75B6", end_color="2E75B6", fill_type="solid")
    red_fill = PatternFill(start_color="FF6B6B", end_color="FF6B6B", fill_type="solid")
    orange_fill = PatternFill(start_color="FFB347", end_color="FFB347", fill_type="solid")
    green_fill = PatternFill(start_color="77DD77", end_color="77DD77", fill_type="solid")
    tb = Border(left=Side("thin"), right=Side("thin"), top=Side("thin"), bottom=Side("thin"))
    wrap = Alignment(wrap_text=True, vertical="top")

    headers = ["#", "Combined Sim%", "Title Sim%", "Desc Sim%",
               "Primary ID", "Primary Date", "Primary Age", "Primary Step", "Primary FGroup", "Primary Priority", "Primary Title",
               "Duplicate ID", "Dup Date", "Dup Age", "Dup Step", "Dup FGroup", "Dup Priority", "Duplicate Title",
               "Action Needed"]
    for c, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=c, value=h)
        cell.font = hf
        cell.fill = hfill
        cell.border = tb
        cell.alignment = Alignment(horizontal="center", wrap_text=True, vertical="center")

    for i, m in enumerate(final, 1):
        csim = round(m["combined_sim"] * 100)
        tsim = round(m["title_sim"] * 100)
        dsim = round(m["desc_sim"] * 100)
        vals = [
            i, "%d%%" % csim, "%d%%" % tsim, "%d%%" % dsim,
            m["primary_id"], m["primary_date"], m["primary_age"], m["primary_step"],
            m["primary_fg"], m["primary_pri"], m["primary_title"],
            m["dup_id"], m["dup_date"], m["dup_age"], m["dup_step"],
            m["dup_fg"], m["dup_pri"], m["dup_title"], ""
        ]
        for c, v in enumerate(vals, 1):
            cell = ws.cell(row=i + 1, column=c, value=v)
            cell.border = tb
            cell.alignment = wrap
        # Color-code combined similarity
        sc = ws.cell(row=i + 1, column=2)
        if csim >= 90:
            sc.fill = red_fill
        elif csim >= 75:
            sc.fill = orange_fill
        else:
            sc.fill = green_fill

    widths = {"A": 4, "B": 10, "C": 9, "D": 9,
              "E": 12, "F": 11, "G": 8, "H": 13, "I": 22, "J": 9, "K": 65,
              "L": 12, "M": 11, "N": 8, "O": 13, "P": 22, "Q": 9, "R": 65, "S": 20}
    for col, w in widths.items():
        ws.column_dimensions[col].width = w

    # Summary by FGroup
    ws2 = wb.create_sheet("Summary by Domain")
    fg_cnt = defaultdict(int)
    for m in final:
        fg_cnt[m["primary_fg"]] += 1
    ws2.cell(row=1, column=1, value="FGroup").font = hf
    ws2.cell(row=1, column=1).fill = hfill
    ws2.cell(row=1, column=1).border = tb
    ws2.cell(row=1, column=2, value="Duplicate Pairs").font = hf
    ws2.cell(row=1, column=2).fill = hfill
    ws2.cell(row=1, column=2).border = tb
    for r, (fg, cnt) in enumerate(sorted(fg_cnt.items(), key=lambda x: -x[1]), 2):
        ws2.cell(row=r, column=1, value=fg).border = tb
        ws2.cell(row=r, column=2, value=cnt).border = tb
    tr = len(fg_cnt) + 2
    ws2.cell(row=tr, column=1, value="TOTAL").border = tb
    ws2.cell(row=tr, column=1).font = Font(bold=True)
    ws2.cell(row=tr, column=2, value=len(final)).border = tb
    ws2.cell(row=tr, column=2).font = Font(bold=True)
    ws2.column_dimensions["A"].width = 30
    ws2.column_dimensions["B"].width = 16

    out = os.path.join(_script_dir, "..", "docs", "output", "existing_duplicate_defects.xlsx")
    os.makedirs(os.path.dirname(out), exist_ok=True)
    wb.save(out)
    print()
    print("Excel saved to: %s" % os.path.abspath(out))
    print("Total unique duplicate pairs: %d" % len(final))


if __name__ == "__main__":
    main()
