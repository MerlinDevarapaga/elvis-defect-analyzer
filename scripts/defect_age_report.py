"""
Defect Age Reporter — Generates a pivot table of Non-HMI defect counts
grouped by Domain Dev Team (FGroup) and age buckets (in days).

Filters:
  - ProjectID = 'MSIL_DA2.8'
  - TicketStepID IN ('Responsible', 'Processing')  (i.e. < Integrating)
  - FGroup NOT LIKE '%HMI%'                         (Non-HMI only)

Age = (today - EnterDateTime) in days

Usage:
    python defect_age_report.py
    python defect_age_report.py --excel output.xlsx
"""
import os
import sys
import json
import io
from datetime import datetime, date
from collections import defaultdict

sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8", errors="replace")

from dotenv import load_dotenv
import mysql.connector

# Load .env from workspace root
_script_dir = os.path.dirname(os.path.abspath(__file__))
_env_candidates = [
    os.path.join(_script_dir, ".env"),
    os.path.join(_script_dir, "..", ".env"),
    os.path.join(_script_dir, "..", "..", "..", "..", ".env"),
]
for _env in _env_candidates:
    if os.path.exists(_env):
        load_dotenv(_env)
        break

# ---------------------------------------------------------------------------
# Configuration
# ---------------------------------------------------------------------------
PROJECT_ID = "MSIL_DA2.8"
PRE_INTEGRATING_STEPS = ("Categorizing", "Processing", "Reproduction")
HMI_EXCLUDE_PATTERN = "%HMI%"

AGE_BUCKETS = [
    (0, 59),
    (60, 119),
    (120, 179),
    (180, 239),
]
BUCKET_LABELS = [f"{lo}-{hi}" for lo, hi in AGE_BUCKETS]

QUERY_COLUMNS = ["TicketID", "FGroup", "EnterDateTime", "TicketStepID", "PriorityID", "Title"]


def get_connection():
    return mysql.connector.connect(
        host=os.getenv("ELVIS_DB_HOST"),
        user=os.getenv("ELVIS_DB_USER"),
        password=os.getenv("ELVIS_DB_PASSWORD"),
        database=os.getenv("ELVIS_DB_NAME"),
        port=int(os.getenv("ELVIS_DB_PORT", 3306)),
    )


def fetch_pre_integration_defects():
    """Fetch all MSIL DA2.8 Non-HMI defects that are before Integrating step."""
    conn = get_connection()
    cursor = conn.cursor(dictionary=True)

    placeholders = ", ".join(["%s"] * len(PRE_INTEGRATING_STEPS))
    cols_sql = ", ".join(f"`{c}`" for c in QUERY_COLUMNS)

    query = (
        f"SELECT {cols_sql} FROM tbl_ElvisSR "
        f"WHERE ProjectID = %s "
        f"  AND TicketStepID IN ({placeholders}) "
        f"  AND FGroup NOT LIKE %s "
        f"  AND IsDeleted = 'N' "
        f"ORDER BY FGroup, EnterDateTime"
    )
    params = [PROJECT_ID] + list(PRE_INTEGRATING_STEPS) + [HMI_EXCLUDE_PATTERN]
    cursor.execute(query, params)
    rows = cursor.fetchall()

    cursor.close()
    conn.close()
    return rows


def compute_age(enter_datetime, reference_date=None):
    """Compute age in days from EnterDateTime to reference_date (default: today)."""
    if reference_date is None:
        reference_date = date.today()
    if isinstance(enter_datetime, datetime):
        enter_date = enter_datetime.date()
    elif isinstance(enter_datetime, date):
        enter_date = enter_datetime
    else:
        enter_date = datetime.strptime(str(enter_datetime)[:10], "%Y-%m-%d").date()
    return (reference_date - enter_date).days


def bucket_index(age_days):
    """Return the bucket index for the given age, or -1 if out of range."""
    for i, (lo, hi) in enumerate(AGE_BUCKETS):
        if lo <= age_days <= hi:
            return i
    return -1  # beyond all defined buckets


def build_pivot(rows):
    """Build a pivot: FGroup -> [bucket_counts...] + grand_total."""
    pivot = defaultdict(lambda: [0] * (len(AGE_BUCKETS) + 1))  # +1 for grand total

    skipped_over = 0
    for row in rows:
        fg = (row.get("FGroup") or "Unknown").strip()
        enter_dt = row.get("EnterDateTime")
        if not enter_dt:
            continue

        age = compute_age(enter_dt)
        idx = bucket_index(age)

        if idx >= 0:
            pivot[fg][idx] += 1
            pivot[fg][-1] += 1  # grand total column
        else:
            # Age exceeds highest bucket — still count in grand total
            pivot[fg][-1] += 1
            skipped_over += 1

    return dict(pivot), skipped_over


def print_table(pivot):
    """Pretty-print the pivot table sorted by Grand Total descending."""
    header_labels = BUCKET_LABELS + ["Grand Total"]

    # Sort by grand total descending
    sorted_teams = sorted(pivot.items(), key=lambda x: x[1][-1], reverse=True)

    # Column widths
    name_w = max(len("Domain Dev Teams"), max((len(k) for k in pivot), default=10))
    col_w = max(len(l) for l in header_labels)
    col_w = max(col_w, 6)

    # Header
    header = f"{'Domain Dev Teams':<{name_w}}"
    for label in header_labels:
        header += f"  {label:>{col_w}}"
    sep = "-" * len(header)

    print()
    print(f"MSIL DA2.8 — Non-HMI Defect Age Report (< Integrating)")
    print(f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')}")
    print(sep)
    print(header)
    print(sep)

    # Grand total row accumulator
    grand = [0] * len(header_labels)

    for team, counts in sorted_teams:
        row_str = f"{team:<{name_w}}"
        for i, c in enumerate(counts):
            if c == 0:
                row_str += f"  {'':>{col_w}}"
            else:
                row_str += f"  {c:>{col_w}}"
            grand[i] += c
        print(row_str)

    print(sep)
    total_str = f"{'Grand Total':<{name_w}}"
    for g in grand:
        total_str += f"  {g:>{col_w}}"
    print(total_str)
    print(sep)
    print(f"\nTotal defects: {grand[-1]}")


def write_excel(pivot, output_path):
    """Write the pivot table to an Excel file."""
    try:
        import openpyxl
        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    except ImportError:
        print("Error: openpyxl not installed. Run: pip install openpyxl")
        return

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Defect Age Report"

    # Styles
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    total_font = Font(bold=True, size=11)
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    # Title row
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(BUCKET_LABELS) + 2)
    title_cell = ws.cell(row=1, column=1, value="MSIL DA2.8 — Non-HMI Defect Age Report (< Integrating)")
    title_cell.font = Font(bold=True, size=14)
    ws.cell(row=2, column=1, value=f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')}")

    # Headers at row 4
    headers = ["Domain Dev Teams"] + BUCKET_LABELS + ["Grand Total"]
    for col_idx, h in enumerate(headers, start=1):
        cell = ws.cell(row=4, column=col_idx, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")
        cell.border = thin_border

    # Data rows sorted by grand total desc
    sorted_teams = sorted(pivot.items(), key=lambda x: x[1][-1], reverse=True)
    grand = [0] * (len(AGE_BUCKETS) + 1)

    for row_idx, (team, counts) in enumerate(sorted_teams, start=5):
        ws.cell(row=row_idx, column=1, value=team).border = thin_border
        for col_idx, c in enumerate(counts, start=2):
            cell = ws.cell(row=row_idx, column=col_idx, value=c if c > 0 else None)
            cell.alignment = Alignment(horizontal="center")
            cell.border = thin_border
            grand[col_idx - 2] += c

    # Grand Total row
    total_row = 5 + len(sorted_teams)
    ws.cell(row=total_row, column=1, value="Grand Total").font = total_font
    ws.cell(row=total_row, column=1).border = thin_border
    for col_idx, g in enumerate(grand, start=2):
        cell = ws.cell(row=total_row, column=col_idx, value=g)
        cell.font = total_font
        cell.alignment = Alignment(horizontal="center")
        cell.border = thin_border

    # Auto-width
    for col in ws.columns:
        max_len = max(len(str(cell.value or "")) for cell in col)
        ws.column_dimensions[col[0].column_letter].width = max_len + 4

    wb.save(output_path)
    print(f"\nExcel saved to: {os.path.abspath(output_path)}")


def main():
    import argparse

    parser = argparse.ArgumentParser(description="MSIL DA2.8 Non-HMI Defect Age Report")
    parser.add_argument("--excel", metavar="FILE", help="Also write output to an Excel file")
    parser.add_argument("--json", metavar="FILE", help="Save raw query results to JSON")
    args = parser.parse_args()

    print(f"Querying Elvis DB for MSIL DA2.8 Non-HMI defects (< Integrating)...")

    try:
        rows = fetch_pre_integration_defects()
    except mysql.connector.Error as e:
        print(f"Database error: {e}")
        sys.exit(1)

    print(f"Fetched {len(rows)} defects.")

    if args.json:
        serializable = []
        for r in rows:
            sr = {}
            for k, v in r.items():
                try:
                    json.dumps(v)
                    sr[k] = v
                except (TypeError, ValueError):
                    sr[k] = str(v)
            serializable.append(sr)
        with open(args.json, "w", encoding="utf-8") as f:
            json.dump(serializable, f, indent=2, ensure_ascii=False)
        print(f"Raw JSON saved to: {os.path.abspath(args.json)}")

    pivot, over_max = build_pivot(rows)

    if not pivot:
        print("No defects found matching the criteria.")
        sys.exit(0)

    print_table(pivot)

    if over_max:
        print(f"\nNote: {over_max} defects exceeded the {AGE_BUCKETS[-1][1]}-day bucket (included in Grand Total only).")

    if args.excel:
        write_excel(pivot, args.excel)


if __name__ == "__main__":
    main()
