"""
Weekly Inflow Report — Last 5 weeks inflow grouped by FGroup,
filtered to ExtRef (ReferenceNumber) = empty/0, 1, 2 only.
Shows total inflow and rejected count per FGroup per week.

Usage:
    python scripts/weekly_inflow_report.py
"""
import os
import sys
import io
from datetime import datetime, date, timedelta
from collections import defaultdict

sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8", errors="replace")

from dotenv import load_dotenv
import mysql.connector

# Load .env from workspace root
_script_dir = os.path.dirname(os.path.abspath(__file__))
_env_candidates = [
    os.path.join(_script_dir, ".env"),
    os.path.join(_script_dir, "..", ".env"),
]
for _env in _env_candidates:
    if os.path.exists(_env):
        load_dotenv(_env)
        break

PROJECT_ID = "MSIL_DA2.8"
NUM_WEEKS = 5


def get_connection():
    return mysql.connector.connect(
        host=os.getenv("ELVIS_DB_HOST"),
        user=os.getenv("ELVIS_DB_USER"),
        password=os.getenv("ELVIS_DB_PASSWORD"),
        database=os.getenv("ELVIS_DB_NAME"),
        port=int(os.getenv("ELVIS_DB_PORT", 3306)),
    )


def get_week_ranges(num_weeks):
    """Return list of (label, start_date, end_date) for the last N complete weeks + current partial week."""
    today = date.today()
    # Current week: Monday .. today
    current_monday = today - timedelta(days=today.weekday())
    weeks = []
    for i in range(num_weeks - 1, -1, -1):
        mon = current_monday - timedelta(weeks=i)
        sun = mon + timedelta(days=6)
        if sun > today:
            sun = today
        label = f"{mon.strftime('%d-%b')} to {sun.strftime('%d-%b')}"
        weeks.append((label, mon, sun))
    return weeks


def fetch_inflow(start_date, end_date):
    """Fetch inflow tickets created between start_date and end_date (inclusive),
    with ReferenceNumber IN (0, 1, 2) or NULL, excluding 3, 4, 5."""
    conn = get_connection()
    cursor = conn.cursor(dictionary=True)

    # end_date + 1 day for < comparison to include full end_date
    end_dt = end_date + timedelta(days=1)

    query = (
        "SELECT TicketID, FGroup, ReferenceNumber, Rejected, EnterDateTime "
        "FROM tbl_ElvisSR "
        "WHERE ProjectID = %s "
        "  AND EnterDateTime >= %s "
        "  AND EnterDateTime < %s "
        "  AND (ReferenceNumber IS NULL OR ReferenceNumber IN (0, 1, 2)) "
        "  AND IsDeleted = 'N' "
        "ORDER BY FGroup, EnterDateTime"
    )
    params = [PROJECT_ID, start_date.strftime("%Y-%m-%d"), end_dt.strftime("%Y-%m-%d")]
    cursor.execute(query, params)
    rows = cursor.fetchall()
    cursor.close()
    conn.close()
    return rows


def assign_week(enter_dt, weeks):
    """Return the week index for a given EnterDateTime."""
    if isinstance(enter_dt, datetime):
        d = enter_dt.date()
    else:
        d = enter_dt
    for i, (_, start, end) in enumerate(weeks):
        if start <= d <= end:
            return i
    return None


def build_report(weeks):
    """Build the full report data."""
    # Fetch all data for the entire range
    overall_start = weeks[0][1]
    overall_end = weeks[-1][2]
    rows = fetch_inflow(overall_start, overall_end)

    # Structure: fgroup -> [{inflow, rejected} per week]
    data = defaultdict(lambda: [{"inflow": 0, "rejected": 0} for _ in range(len(weeks))])

    for row in rows:
        fg = (row.get("FGroup") or "Unknown").strip()
        enter_dt = row.get("EnterDateTime")
        rejected = row.get("Rejected", "N")
        if not enter_dt:
            continue

        wi = assign_week(enter_dt, weeks)
        if wi is None:
            continue

        data[fg][wi]["inflow"] += 1
        if rejected == "Y":
            data[fg][wi]["rejected"] += 1

    return dict(data)


def print_report(weeks, data):
    """Pretty-print the weekly inflow report."""
    # Sort FGroups by total inflow descending
    def total_inflow(fg):
        return sum(w["inflow"] for w in data[fg])

    sorted_fgroups = sorted(data.keys(), key=total_inflow, reverse=True)

    # Column widths
    fg_w = max(len("Functional Group"), max((len(fg) for fg in sorted_fgroups), default=15))
    week_labels = [w[0] for w in weeks]
    col_w = max(max((len(l) for l in week_labels), default=12), 12)

    # Header
    print()
    print(f"MSIL DA2.8 — Weekly Inflow Report (Last {NUM_WEEKS} Weeks)")
    print(f"Filter: ExtRef (ReferenceNumber) = empty/0, 1, 2 only (excluding 3, 4, 5)")
    print(f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')}")
    print()

    # Build header row
    header = f"{'Functional Group':<{fg_w}}"
    for label in week_labels:
        header += f"  {label:^{col_w}}"
    header += f"  {'Total':^{col_w}}"
    sep = "-" * len(header)

    # Sub-header (Inflow / Rej)
    sub_header = f"{'':<{fg_w}}"
    for _ in week_labels:
        sub_header += f"  {'In / Rej':^{col_w}}"
    sub_header += f"  {'In / Rej':^{col_w}}"

    print(sep)
    print(header)
    print(sub_header)
    print(sep)

    # Grand totals
    grand_inflow = [0] * len(weeks)
    grand_rejected = [0] * len(weeks)

    for fg in sorted_fgroups:
        row_str = f"{fg:<{fg_w}}"
        total_in = 0
        total_rej = 0
        for i, wdata in enumerate(data[fg]):
            inf = wdata["inflow"]
            rej = wdata["rejected"]
            cell = f"{inf}/{rej}" if rej > 0 else f"{inf}"
            row_str += f"  {cell:^{col_w}}"
            total_in += inf
            total_rej += rej
            grand_inflow[i] += inf
            grand_rejected[i] += rej
        total_cell = f"{total_in}/{total_rej}" if total_rej > 0 else f"{total_in}"
        row_str += f"  {total_cell:^{col_w}}"
        print(row_str)

    print(sep)

    # Grand total row
    total_str = f"{'Grand Total':<{fg_w}}"
    gt_in = 0
    gt_rej = 0
    for i in range(len(weeks)):
        cell = f"{grand_inflow[i]}/{grand_rejected[i]}" if grand_rejected[i] > 0 else f"{grand_inflow[i]}"
        total_str += f"  {cell:^{col_w}}"
        gt_in += grand_inflow[i]
        gt_rej += grand_rejected[i]
    gt_cell = f"{gt_in}/{gt_rej}" if gt_rej > 0 else f"{gt_in}"
    total_str += f"  {gt_cell:^{col_w}}"
    print(total_str)
    print(sep)

    print(f"\nTotal inflow: {gt_in}  |  Total rejected: {gt_rej}  |  Rejection rate: {gt_rej/gt_in*100:.1f}%" if gt_in > 0 else "")
    print(f"Format: Inflow/Rejected (e.g., 25/3 means 25 inflow, 3 rejected)")


def write_excel(weeks, data, output_path):
    """Write the weekly inflow report to an Excel file with formatting."""
    try:
        import openpyxl
        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    except ImportError:
        print("Error: openpyxl not installed. Run: pip install openpyxl")
        return

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Weekly Inflow Report"

    # Styles
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    total_font = Font(bold=True, size=11)
    total_fill = PatternFill(start_color="D9E2F3", end_color="D9E2F3", fill_type="solid")
    red_font = Font(bold=True, color="C00000")
    center = Alignment(horizontal="center", vertical="center")
    left = Alignment(horizontal="left", vertical="center")
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )

    week_labels = [w[0] for w in weeks]

    # Sort FGroups by total inflow descending
    def total_inflow(fg):
        return sum(w["inflow"] for w in data[fg])
    sorted_fgroups = sorted(data.keys(), key=total_inflow, reverse=True)

    # --- Title row ---
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=2 + len(week_labels) * 2 + 2)
    title_cell = ws.cell(row=1, column=1, value="MSIL DA2.8 — Weekly Inflow Report (Last 5 Weeks) | ExtRef: empty/0/1/2 only")
    title_cell.font = Font(bold=True, size=13, color="1F4E79")

    # --- Header row (row 3) ---
    # Columns: # | Functional Group | (Inflow, Rejected) per week | Total Inflow | Total Rejected | Rej %
    headers = ["#", "Functional Group"]
    for label in week_labels:
        headers.append(f"{label} (In)")
        headers.append(f"{label} (Rej)")
    headers += ["Total Inflow", "Total Rejected", "Rej %"]

    for col_idx, h in enumerate(headers, 1):
        cell = ws.cell(row=3, column=col_idx, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center
        cell.border = thin_border

    # --- Data rows ---
    row_num = 4
    grand_inflow = [0] * len(weeks)
    grand_rejected = [0] * len(weeks)

    for idx, fg in enumerate(sorted_fgroups, 1):
        col = 1
        ws.cell(row=row_num, column=col, value=idx).alignment = center
        ws.cell(row=row_num, column=col, value=idx).border = thin_border
        col += 1
        ws.cell(row=row_num, column=col, value=fg).alignment = left
        ws.cell(row=row_num, column=col, value=fg).border = thin_border
        col += 1

        t_in = 0
        t_rej = 0
        for i, wdata in enumerate(data[fg]):
            inf = wdata["inflow"]
            rej = wdata["rejected"]
            # Inflow cell
            c_in = ws.cell(row=row_num, column=col, value=inf if inf > 0 else None)
            c_in.alignment = center
            c_in.border = thin_border
            col += 1
            # Rejected cell
            c_rej = ws.cell(row=row_num, column=col, value=rej if rej > 0 else None)
            c_rej.alignment = center
            c_rej.border = thin_border
            if rej > 0:
                c_rej.font = red_font
            col += 1
            t_in += inf
            t_rej += rej
            grand_inflow[i] += inf
            grand_rejected[i] += rej

        # Total Inflow
        c = ws.cell(row=row_num, column=col, value=t_in)
        c.alignment = center
        c.border = thin_border
        c.font = Font(bold=True)
        col += 1
        # Total Rejected
        c = ws.cell(row=row_num, column=col, value=t_rej if t_rej > 0 else None)
        c.alignment = center
        c.border = thin_border
        if t_rej > 0:
            c.font = red_font
        col += 1
        # Rej %
        pct = f"{t_rej/t_in*100:.1f}%" if t_in > 0 else ""
        c = ws.cell(row=row_num, column=col, value=pct)
        c.alignment = center
        c.border = thin_border

        row_num += 1

    # --- Grand Total row ---
    col = 1
    ws.cell(row=row_num, column=col).border = thin_border
    col += 1
    c = ws.cell(row=row_num, column=col, value="Grand Total")
    c.font = total_font
    c.fill = total_fill
    c.border = thin_border
    col += 1

    gt_in = 0
    gt_rej = 0
    for i in range(len(weeks)):
        c = ws.cell(row=row_num, column=col, value=grand_inflow[i])
        c.font = total_font
        c.fill = total_fill
        c.alignment = center
        c.border = thin_border
        col += 1
        c = ws.cell(row=row_num, column=col, value=grand_rejected[i] if grand_rejected[i] > 0 else None)
        c.font = Font(bold=True, color="C00000") if grand_rejected[i] > 0 else total_font
        c.fill = total_fill
        c.alignment = center
        c.border = thin_border
        col += 1
        gt_in += grand_inflow[i]
        gt_rej += grand_rejected[i]

    c = ws.cell(row=row_num, column=col, value=gt_in)
    c.font = total_font
    c.fill = total_fill
    c.alignment = center
    c.border = thin_border
    col += 1
    c = ws.cell(row=row_num, column=col, value=gt_rej)
    c.font = Font(bold=True, color="C00000")
    c.fill = total_fill
    c.alignment = center
    c.border = thin_border
    col += 1
    c = ws.cell(row=row_num, column=col, value=f"{gt_rej/gt_in*100:.1f}%" if gt_in > 0 else "")
    c.font = total_font
    c.fill = total_fill
    c.alignment = center
    c.border = thin_border

    # --- Column widths ---
    ws.column_dimensions["A"].width = 4
    ws.column_dimensions["B"].width = 28
    for i in range(3, col + 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = 14

    wb.save(output_path)
    print(f"\nExcel saved to: {output_path}")


def main():
    weeks = get_week_ranges(NUM_WEEKS)
    data = build_report(weeks)
    print_report(weeks, data)

    # Always export to Excel in Downloads
    downloads = os.path.join(os.path.expanduser("~"), "Downloads")
    output_path = os.path.join(downloads, "weekly_inflow_report.xlsx")
    write_excel(weeks, data, output_path)


if __name__ == "__main__":
    main()
