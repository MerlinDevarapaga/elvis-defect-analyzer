"""
Weekly Inflow/Outflow Report — Last 5 weeks inflow + outflow by FGroup,
filtered to ExtRef (ReferenceNumber) = empty/0, 1, 2 only.
Also shows how many are P8 / P8_YTB_NA items.

- Inflow  = tickets created  (EnterDateTime) in each week
- Outflow = tickets that reached >= Concluding step (FirstConclDateTime) in each week

Usage:
    python scripts/weekly_inflow_outflow_report.py
"""
import os
import sys
import io
from datetime import datetime, date, timedelta
from collections import defaultdict

sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8", errors="replace")

from dotenv import load_dotenv
import mysql.connector

# Load .env from workspace root
_script_dir = os.path.dirname(os.path.abspath(__file__))
for _env in [os.path.join(_script_dir, "..", ".env"), os.path.join(_script_dir, ".env")]:
    if os.path.exists(_env):
        load_dotenv(_env)
        break

PROJECT_ID = "MSIL_DA2.8"
NUM_WEEKS = 5

P8_PATTERNS = ("p8", "p8_ytb_na", "p8_ytb")


def is_p8(fg_swrev):
    """Check if FG_SWRev starts with P8 or is P8_YTB_NA."""
    if not fg_swrev:
        return False
    val = fg_swrev.strip().lower()
    return val.startswith("p8") or val in P8_PATTERNS


def get_connection():
    return mysql.connector.connect(
        host=os.getenv("ELVIS_DB_HOST"),
        user=os.getenv("ELVIS_DB_USER"),
        password=os.getenv("ELVIS_DB_PASSWORD"),
        database=os.getenv("ELVIS_DB_NAME"),
        port=int(os.getenv("ELVIS_DB_PORT", 3306)),
    )


def get_week_ranges(num_weeks):
    """Return list of (label, start_date, end_date) for the last N weeks."""
    today = date.today()
    current_monday = today - timedelta(days=today.weekday())
    weeks = []
    for i in range(num_weeks - 1, -1, -1):
        mon = current_monday - timedelta(weeks=i)
        sun = mon + timedelta(days=6)
        if sun > today:
            sun = today
        label = f"{mon.strftime('%d-%b')} to {sun.strftime('%d-%b')}"
        weeks.append((label, mon, sun))
    return weeks


def fetch_inflow_outflow(start_date, end_date):
    """Fetch tickets created OR integrated within the date range,
    with ReferenceNumber IN (0, 1, 2) or NULL."""
    conn = get_connection()
    cursor = conn.cursor(dictionary=True)

    end_dt = end_date + timedelta(days=1)

    # Inflow: created in range
    # Outflow: FirstConclDateTime in range (>= Concluding)
    # We need both, so fetch tickets where EITHER applies
    query = (
        "SELECT TicketID, FGroup, FG_SWRev, ReferenceNumber, Rejected, "
        "       EnterDateTime, FirstConclDateTime "
        "FROM tbl_ElvisSR "
        "WHERE ProjectID = %s "
        "  AND (ReferenceNumber IS NULL OR ReferenceNumber IN (0, 1, 2)) "
        "  AND IsDeleted = 'N' "
        "  AND ( "
        "    (EnterDateTime >= %s AND EnterDateTime < %s) "
        "    OR (FirstConclDateTime >= %s AND FirstConclDateTime < %s) "
        "  ) "
        "ORDER BY FGroup, EnterDateTime"
    )
    params = [
        PROJECT_ID,
        start_date.strftime("%Y-%m-%d"), end_dt.strftime("%Y-%m-%d"),
        start_date.strftime("%Y-%m-%d"), end_dt.strftime("%Y-%m-%d"),
    ]
    cursor.execute(query, params)
    rows = cursor.fetchall()
    cursor.close()
    conn.close()
    return rows


def to_date(dt):
    if isinstance(dt, datetime):
        return dt.date()
    if isinstance(dt, date):
        return dt
    return None


def assign_week(d, weeks):
    if d is None:
        return None
    for i, (_, start, end) in enumerate(weeks):
        if start <= d <= end:
            return i
    return None


def build_report(weeks):
    """Build inflow, outflow, p8 counts per FGroup per week."""
    overall_start = weeks[0][1]
    overall_end = weeks[-1][2]
    rows = fetch_inflow_outflow(overall_start, overall_end)

    nw = len(weeks)
    # fgroup -> {inflow: [per week], outflow: [per week], p8_inflow: [per week]}
    data = defaultdict(lambda: {
        "inflow": [0] * nw,
        "outflow": [0] * nw,
        "rejected": [0] * nw,
        "p8_inflow": [0] * nw,
    })

    for row in rows:
        fg = (row.get("FGroup") or "Unknown").strip()
        enter_dt = to_date(row.get("EnterDateTime"))
        concl_dt = to_date(row.get("FirstConclDateTime"))
        rejected = row.get("Rejected", "N")
        fg_swrev = row.get("FG_SWRev", "")

        # Inflow
        wi = assign_week(enter_dt, weeks)
        if wi is not None:
            data[fg]["inflow"][wi] += 1
            if rejected == "Y":
                data[fg]["rejected"][wi] += 1
            if is_p8(fg_swrev):
                data[fg]["p8_inflow"][wi] += 1

        # Outflow (ticket reached >= Concluding)
        wo = assign_week(concl_dt, weeks)
        if wo is not None:
            data[fg]["outflow"][wo] += 1

    return dict(data)


def print_report(weeks, data):
    """Print the weekly inflow/outflow table."""
    nw = len(weeks)

    def total_inflow(fg):
        return sum(data[fg]["inflow"])

    sorted_fgroups = sorted(data.keys(), key=total_inflow, reverse=True)
    week_labels = [w[0] for w in weeks]

    fg_w = max(len("Functional Group"), max((len(fg) for fg in sorted_fgroups), default=15))
    col_w = max(max((len(l) for l in week_labels), default=12), 14)

    print()
    print(f"MSIL DA2.8 — Weekly Inflow / Outflow Report (Last {NUM_WEEKS} Weeks)")
    print(f"Filter: ExtRef (ReferenceNumber) = empty/0, 1, 2 only (excluding 3, 4, 5)")
    print(f"Outflow = tickets that reached >= Concluding step in that week")
    print(f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')}")
    print()

    # Header
    header = f"{'Functional Group':<{fg_w}}"
    for label in week_labels:
        header += f"  {label:^{col_w}}"
    header += f"  {'Total':^{col_w}}"
    sep = "-" * len(header)

    sub = f"{'':<{fg_w}}"
    for _ in week_labels:
        sub += f"  {'In/Out/Rej':^{col_w}}"
    sub += f"  {'In/Out/Rej':^{col_w}}"

    print(sep)
    print(header)
    print(sub)
    print(sep)

    g_in = [0] * nw
    g_out = [0] * nw
    g_rej = [0] * nw
    g_p8 = [0] * nw

    for fg in sorted_fgroups:
        row_str = f"{fg:<{fg_w}}"
        t_in = t_out = t_rej = t_p8 = 0
        for i in range(nw):
            inf = data[fg]["inflow"][i]
            out = data[fg]["outflow"][i]
            rej = data[fg]["rejected"][i]
            cell = f"{inf}/{out}/{rej}"
            row_str += f"  {cell:^{col_w}}"
            t_in += inf
            t_out += out
            t_rej += rej
            t_p8 += data[fg]["p8_inflow"][i]
            g_in[i] += inf
            g_out[i] += out
            g_rej[i] += rej
            g_p8[i] += data[fg]["p8_inflow"][i]
        row_str += f"  {f'{t_in}/{t_out}/{t_rej}':^{col_w}}"
        print(row_str)

    print(sep)

    total_str = f"{'Grand Total':<{fg_w}}"
    gt_in = gt_out = gt_rej = gt_p8 = 0
    for i in range(nw):
        cell = f"{g_in[i]}/{g_out[i]}/{g_rej[i]}"
        total_str += f"  {cell:^{col_w}}"
        gt_in += g_in[i]
        gt_out += g_out[i]
        gt_rej += g_rej[i]
        gt_p8 += g_p8[i]
    total_str += f"  {f'{gt_in}/{gt_out}/{gt_rej}':^{col_w}}"
    print(total_str)
    print(sep)

    print(f"\nTotal inflow: {gt_in}  |  Total outflow: {gt_out}  |  Net: {gt_in - gt_out}  |  Rejected: {gt_rej}")

    # P8 / P8_YTB_NA summary
    print(f"\n{'='*60}")
    print(f"P8 / P8_YTB_NA INFLOW SUMMARY")
    print(f"{'='*60}")

    p8_header = f"{'Functional Group':<{fg_w}}"
    for label in week_labels:
        p8_header += f"  {label:^{col_w}}"
    p8_header += f"  {'Total':^{col_w}}"

    print(p8_header)
    print("-" * len(p8_header))

    any_p8 = False
    gt_p8_total = 0
    for fg in sorted_fgroups:
        t_p8 = sum(data[fg]["p8_inflow"])
        if t_p8 == 0:
            continue
        any_p8 = True
        row_str = f"{fg:<{fg_w}}"
        for i in range(nw):
            v = data[fg]["p8_inflow"][i]
            row_str += f"  {(str(v) if v > 0 else ''):^{col_w}}"
        row_str += f"  {t_p8:^{col_w}}"
        gt_p8_total += t_p8
        print(row_str)

    if not any_p8:
        print("  (No P8/P8_YTB_NA tickets found in this period)")
    else:
        print("-" * len(p8_header))
        total_p8_str = f"{'Total P8/P8_YTB_NA':<{fg_w}}"
        for i in range(nw):
            total_p8_str += f"  {(str(g_p8[i]) if g_p8[i] > 0 else ''):^{col_w}}"
        total_p8_str += f"  {gt_p8_total:^{col_w}}"
        print(total_p8_str)
        print(f"\nP8/P8_YTB_NA tickets are {gt_p8_total} out of {gt_in} total inflow ({gt_p8_total/gt_in*100:.1f}%)" if gt_in > 0 else "")


def write_excel(weeks, data, output_path):
    """Write the inflow/outflow report to Excel."""
    try:
        import openpyxl
        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    except ImportError:
        print("Error: openpyxl not installed. Run: pip install openpyxl")
        return

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Inflow-Outflow Report"

    nw = len(weeks)
    week_labels = [w[0] for w in weeks]

    # Styles
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    in_fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
    out_fill = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")
    rej_fill = PatternFill(start_color="FCE4EC", end_color="FCE4EC", fill_type="solid")
    p8_fill = PatternFill(start_color="FFF3E0", end_color="FFF3E0", fill_type="solid")
    total_font = Font(bold=True, size=11)
    total_fill = PatternFill(start_color="D9E2F3", end_color="D9E2F3", fill_type="solid")
    red_font = Font(bold=True, color="C00000")
    center = Alignment(horizontal="center", vertical="center")
    left_a = Alignment(horizontal="left", vertical="center")
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )

    def total_inflow(fg):
        return sum(data[fg]["inflow"])
    sorted_fgroups = sorted(data.keys(), key=total_inflow, reverse=True)

    # --- Title ---
    ncols = 2 + nw * 3 + 4  # #, FGroup, (In/Out/Rej per week), TotalIn, TotalOut, TotalRej, P8
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=ncols)
    ws.cell(row=1, column=1,
            value="MSIL DA2.8 — Weekly Inflow/Outflow Report (Last 5 Weeks) | ExtRef: empty/0/1/2 only"
    ).font = Font(bold=True, size=13, color="1F4E79")

    # --- Week group headers (row 3) ---
    col = 3  # skip #, FGroup
    for label in week_labels:
        ws.merge_cells(start_row=3, start_column=col, end_row=3, end_column=col + 2)
        c = ws.cell(row=3, column=col, value=label)
        c.font = header_font
        c.fill = header_fill
        c.alignment = center
        c.border = thin_border
        col += 3
    # Total columns
    ws.merge_cells(start_row=3, start_column=col, end_row=3, end_column=col + 2)
    ws.cell(row=3, column=col, value="Total").font = header_font
    ws.cell(row=3, column=col).fill = header_fill
    ws.cell(row=3, column=col).alignment = center
    ws.cell(row=3, column=col).border = thin_border

    # --- Sub-headers (row 4) ---
    headers = ["#", "Functional Group"]
    for _ in week_labels:
        headers += ["In", "Out", "Rej"]
    headers += ["Total In", "Total Out", "Total Rej", "P8/YTB"]
    for ci, h in enumerate(headers, 1):
        c = ws.cell(row=4, column=ci, value=h)
        c.font = Font(bold=True, color="FFFFFF", size=10)
        c.alignment = center
        c.border = thin_border
        if "In" in h and "Total" not in h:
            c.fill = PatternFill(start_color="548235", end_color="548235", fill_type="solid")
        elif "Out" in h:
            c.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        elif "Rej" in h:
            c.fill = PatternFill(start_color="C00000", end_color="C00000", fill_type="solid")
        elif "P8" in h:
            c.fill = PatternFill(start_color="ED7D31", end_color="ED7D31", fill_type="solid")
        else:
            c.fill = header_fill

    # --- Data rows ---
    row_num = 5
    g_in = [0] * nw
    g_out = [0] * nw
    g_rej = [0] * nw
    g_p8 = 0

    for idx, fg in enumerate(sorted_fgroups, 1):
        col = 1
        ws.cell(row=row_num, column=col, value=idx).alignment = center
        ws.cell(row=row_num, column=col).border = thin_border
        col += 1
        ws.cell(row=row_num, column=col, value=fg).alignment = left_a
        ws.cell(row=row_num, column=col).border = thin_border
        col += 1

        t_in = t_out = t_rej = t_p8 = 0
        for i in range(nw):
            inf = data[fg]["inflow"][i]
            out = data[fg]["outflow"][i]
            rej = data[fg]["rejected"][i]

            c = ws.cell(row=row_num, column=col, value=inf if inf else None)
            c.alignment = center; c.border = thin_border; c.fill = in_fill
            col += 1
            c = ws.cell(row=row_num, column=col, value=out if out else None)
            c.alignment = center; c.border = thin_border; c.fill = out_fill
            col += 1
            c = ws.cell(row=row_num, column=col, value=rej if rej else None)
            c.alignment = center; c.border = thin_border; c.fill = rej_fill
            if rej: c.font = red_font
            col += 1

            t_in += inf; t_out += out; t_rej += rej
            g_in[i] += inf; g_out[i] += out; g_rej[i] += rej

        t_p8 = sum(data[fg]["p8_inflow"])
        g_p8 += t_p8

        # Total In
        c = ws.cell(row=row_num, column=col, value=t_in)
        c.alignment = center; c.border = thin_border; c.font = total_font
        col += 1
        # Total Out
        c = ws.cell(row=row_num, column=col, value=t_out)
        c.alignment = center; c.border = thin_border; c.font = total_font
        col += 1
        # Total Rej
        c = ws.cell(row=row_num, column=col, value=t_rej if t_rej else None)
        c.alignment = center; c.border = thin_border
        if t_rej: c.font = red_font
        col += 1
        # P8
        c = ws.cell(row=row_num, column=col, value=t_p8 if t_p8 else None)
        c.alignment = center; c.border = thin_border; c.fill = p8_fill
        col += 1

        row_num += 1

    # --- Grand Total ---
    col = 1
    ws.cell(row=row_num, column=col).border = thin_border
    col += 1
    c = ws.cell(row=row_num, column=col, value="Grand Total")
    c.font = total_font; c.fill = total_fill; c.border = thin_border
    col += 1

    gt_in = gt_out = gt_rej = 0
    for i in range(nw):
        c = ws.cell(row=row_num, column=col, value=g_in[i])
        c.font = total_font; c.fill = total_fill; c.alignment = center; c.border = thin_border
        col += 1
        c = ws.cell(row=row_num, column=col, value=g_out[i])
        c.font = total_font; c.fill = total_fill; c.alignment = center; c.border = thin_border
        col += 1
        c = ws.cell(row=row_num, column=col, value=g_rej[i])
        c.font = total_font; c.fill = total_fill; c.alignment = center; c.border = thin_border
        if g_rej[i]: c.font = Font(bold=True, color="C00000")
        col += 1
        gt_in += g_in[i]; gt_out += g_out[i]; gt_rej += g_rej[i]

    c = ws.cell(row=row_num, column=col, value=gt_in)
    c.font = total_font; c.fill = total_fill; c.alignment = center; c.border = thin_border
    col += 1
    c = ws.cell(row=row_num, column=col, value=gt_out)
    c.font = total_font; c.fill = total_fill; c.alignment = center; c.border = thin_border
    col += 1
    c = ws.cell(row=row_num, column=col, value=gt_rej)
    c.font = Font(bold=True, color="C00000"); c.fill = total_fill; c.alignment = center; c.border = thin_border
    col += 1
    c = ws.cell(row=row_num, column=col, value=g_p8)
    c.font = total_font; c.fill = PatternFill(start_color="FFF3E0", end_color="FFF3E0", fill_type="solid")
    c.alignment = center; c.border = thin_border

    # Column widths
    ws.column_dimensions["A"].width = 4
    ws.column_dimensions["B"].width = 30
    for i in range(3, col + 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = 10

    wb.save(output_path)
    print(f"\nExcel saved to: {output_path}")


def main():
    weeks = get_week_ranges(NUM_WEEKS)
    data = build_report(weeks)
    print_report(weeks, data)

    downloads = os.path.join(os.path.expanduser("~"), "Downloads")
    output_path = os.path.join(downloads, "weekly_inflow_outflow_report.xlsx")
    write_excel(weeks, data, output_path)


if __name__ == "__main__":
    main()
