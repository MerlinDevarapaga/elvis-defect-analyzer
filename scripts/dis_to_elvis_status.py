"""
Query Elvis DB for DIS reference numbers (JIRA) and export Elvis TicketID,
current status, and rejected status to Excel.
Mapping: Elvis TicketID = 3700000 + DIS_number
"""
import os
import sys
import io
from datetime import datetime

sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8", errors="replace")

from dotenv import load_dotenv
import mysql.connector

# Load .env
_script_dir = os.path.dirname(os.path.abspath(__file__))
for _env in [os.path.join(_script_dir, ".env"), os.path.join(_script_dir, "..", ".env")]:
    if os.path.exists(_env):
        load_dotenv(_env)
        break

# DIS reference numbers from the input list
DIS_IDS = [
    "DIS-11137",
    "DIS-10576",
    "DIS-9627",
    "DIS-10674",
    "DIS-10062",
    "DIS-10263",
    "DIS-10033",
    "DIS-10323",
    "DIS-10214",
    "DIS-30080",
    "DIS-10390",
    "DIS-10087",
    "DIS-9915",
    "DIS-10115",
    "DIS-10064",
    "DIS-10004",
    "DIS-9842",
    "DIS-9827",
    "DIS-9764",
    "DIS-9724",
    "DIS-8745",
    "DIS-9622",
    "DIS-9457",
    "DIS-9445",
    "DIS-9424",
    "DIS-9320",
    "DIS-9130",
    "DIS-9067",
    "DIS-8961",
    "DIS-8848",
    "DIS-8727",
    "DIS-8649",
    "DIS-8451",
    "DIS-8043",
    "DIS-7880",
    "DIS-5426",
    "DIS-5392",
    "DIS-3275",
    "DIS-902",
]


def get_connection():
    return mysql.connector.connect(
        host=os.getenv("ELVIS_DB_HOST"),
        user=os.getenv("ELVIS_DB_USER"),
        password=os.getenv("ELVIS_DB_PASSWORD"),
        database=os.getenv("ELVIS_DB_NAME"),
        port=int(os.getenv("ELVIS_DB_PORT", 3306)),
        connection_timeout=15,
    )


def dis_to_elvis_id(dis_ref):
    """Convert DIS-XXXXX to Elvis TicketID: 3700000 + number."""
    num = int(dis_ref.replace("DIS-", ""))
    return 3700000 + num


def query_dis_status(dis_ids):
    """Query Elvis DB for each DIS reference and get status info."""
    conn = get_connection()
    cursor = conn.cursor(dictionary=True)

    results = []
    not_found = []

    for dis_id in dis_ids:
        elvis_id = dis_to_elvis_id(dis_id)
        cursor.execute(
            "SELECT TicketID, TicketStepID, Rejected, RejectReason, Title, FGroup, PriorityID "
            "FROM tbl_ElvisSR "
            "WHERE TicketID = %s",
            (elvis_id,),
        )
        row = cursor.fetchone()

        if row:
            results.append({
                "DIS_Reference": dis_id,
                "Elvis_TicketID": row["TicketID"],
                "Current_Status": row["TicketStepID"],
                "Rejected": row["Rejected"],
                "Reject_Reason": row["RejectReason"] or "",
                "Title": row["Title"],
                "FGroup": row["FGroup"],
                "Priority": row["PriorityID"],
            })
        else:
            not_found.append(dis_id)
            results.append({
                "DIS_Reference": dis_id,
                "Elvis_TicketID": f"{elvis_id} (NOT FOUND)",
                "Current_Status": "",
                "Rejected": "",
                "Reject_Reason": "",
                "Title": "",
                "FGroup": "",
                "Priority": "",
            })

    cursor.close()
    conn.close()
    return results, not_found


def export_to_excel(results, not_found):
    """Export results to Excel."""
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
    except ImportError:
        print("Installing openpyxl...")
        import subprocess
        subprocess.check_call([sys.executable, "-m", "pip", "install", "openpyxl", "-q"])
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "DIS to Elvis Status"

    # Headers
    headers = ["DIS Reference", "Elvis Ticket ID", "Current Status", 
               "Rejected", "Reject Reason", "Title", "FGroup", "Priority"]
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    # Data rows
    for row_idx, r in enumerate(results, 2):
        ws.cell(row=row_idx, column=1, value=r["DIS_Reference"])
        ws.cell(row=row_idx, column=2, value=r["Elvis_TicketID"])
        ws.cell(row=row_idx, column=3, value=r["Current_Status"])
        ws.cell(row=row_idx, column=4, value=r["Rejected"])
        ws.cell(row=row_idx, column=5, value=r["Reject_Reason"])
        ws.cell(row=row_idx, column=6, value=r["Title"])
        ws.cell(row=row_idx, column=7, value=r["FGroup"])
        ws.cell(row=row_idx, column=8, value=r["Priority"])

        # Highlight NOT FOUND rows in red
        if r["Elvis_TicketID"] == "NOT FOUND":
            red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
            for col in range(1, 9):
                ws.cell(row=row_idx, column=col).fill = red_fill

    # Auto-width columns
    for col in ws.columns:
        max_len = 0
        col_letter = col[0].column_letter
        for cell in col:
            if cell.value:
                max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = min(max_len + 2, 50)

    # Save
    downloads = os.path.join(os.path.expanduser("~"), "Downloads")
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    filepath = os.path.join(downloads, f"DIS_Elvis_Status_{timestamp}.xlsx")
    wb.save(filepath)
    return filepath


def main():
    print(f"Querying Elvis DB for {len(DIS_IDS)} DIS references...")
    print(f"Mapping: Elvis TicketID = 3700000 + DIS number")
    results, not_found = query_dis_status(DIS_IDS)

    # Print summary
    found_count = len(DIS_IDS) - len(not_found)
    print(f"\nFound: {found_count}/{len(DIS_IDS)}")
    if not_found:
        print(f"Not found: {len(not_found)}")
        for nf in not_found:
            print(f"  - {nf}")

    # Print table
    print(f"\n{'DIS Ref':<12} {'Elvis ID':<10} {'Status':<15} {'Rejected':<5} {'Reject Reason'}")
    print("-" * 75)
    for r in results:
        print(f"{r['DIS_Reference']:<12} {str(r['Elvis_TicketID']):<10} "
              f"{r['Current_Status']:<15} {r['Rejected']:<5} {r['Reject_Reason']}")

    # Export to Excel
    filepath = export_to_excel(results, not_found)
    print(f"\nExcel exported to: {filepath}")


if __name__ == "__main__":
    main()
