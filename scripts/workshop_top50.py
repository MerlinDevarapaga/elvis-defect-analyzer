"""
Workshop Defect Picker — Select top 50 pre-integrating defects for a cross-team
resolution workshop. Excludes P8/YTB/NA priority items.

Ranking logic:
  1. Priority (A(1) > B(2) > C(3) > ...)
  2. Age (older first)
  3. Spread across all FGroups (teams) so every team has items

Output: Console summary + Excel with workshop agenda.
"""
import os
import sys
import io
import re
from datetime import datetime, date
from collections import defaultdict

sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8", errors="replace")

from dotenv import load_dotenv
import mysql.connector

_script_dir = os.path.dirname(os.path.abspath(__file__))
for _env in [os.path.join(_script_dir, "..", ".env"), os.path.join(_script_dir, ".env")]:
    if os.path.exists(_env):
        load_dotenv(_env)
        break

PROJECT_ID = "MSIL_DA2.8"
PRE_INTEGRATING_STEPS = ("Categorizing", "Processing", "Reproduction")
EXCLUDE_PRIORITIES = ("p8_ytb_NA", "P8_YTB_NA", "p8_ytb_na")
TOP_N = 50

PRIORITY_ORDER = {"A(1)": 1, "top": 2, "B(2)": 3, "C(3)": 4, "D(4)": 5}


def get_connection():
    return mysql.connector.connect(
        host=os.getenv("ELVIS_DB_HOST"),
        user=os.getenv("ELVIS_DB_USER"),
        password=os.getenv("ELVIS_DB_PASSWORD"),
        database=os.getenv("ELVIS_DB_NAME"),
        port=int(os.getenv("ELVIS_DB_PORT", 3306)),
    )


def compute_age(enter_datetime):
    if isinstance(enter_datetime, datetime):
        d = enter_datetime.date()
    elif isinstance(enter_datetime, date):
        d = enter_datetime
    else:
        d = datetime.strptime(str(enter_datetime)[:10], "%Y-%m-%d").date()
    return (date.today() - d).days


def priority_rank(pri):
    if not pri:
        return 99
    return PRIORITY_ORDER.get(pri, 10)


def main():
    conn = get_connection()
    cursor = conn.cursor(dictionary=True)

    placeholders = ", ".join(["%s"] * len(PRE_INTEGRATING_STEPS))
    query = f"""
        SELECT TicketID, Title, EnterDateTime, TicketStepID, FGroup,
               PriorityID, Occurance, ProblemType, `System`, Component
        FROM tbl_ElvisSR
        WHERE ProjectID = %s
          AND TicketStepID IN ({placeholders})
          AND IsDeleted = 'N'
        ORDER BY FGroup, EnterDateTime
    """
    cursor.execute(query, [PROJECT_ID] + list(PRE_INTEGRATING_STEPS))
    all_tickets = cursor.fetchall()
    cursor.close()
    conn.close()

    # Filter out P8/YTB/NA and HMI
    filtered = []
    for t in all_tickets:
        pri = (t.get("PriorityID") or "").strip()
        if pri.lower() in ("p8_ytb_na", "p8_ytb", "na", ""):
            continue
        # Exclude items with 'ytb' or 'na' patterns in priority
        if "ytb" in pri.lower() or pri.lower() == "na":
            continue
        filtered.append(t)

    print("Total pre-integrating tickets: %d" % len(all_tickets))
    print("After excluding P8/YTB/NA: %d" % len(filtered))

    # Compute age and sort score
    for t in filtered:
        t["_age"] = compute_age(t["EnterDateTime"])
        t["_pri_rank"] = priority_rank(t.get("PriorityID"))

    # Group by FGroup
    by_team = defaultdict(list)
    for t in filtered:
        fg = (t.get("FGroup") or "Unknown").strip()
        by_team[fg].append(t)

    # Sort each team's tickets by priority then age (oldest first)
    for fg in by_team:
        by_team[fg].sort(key=lambda x: (x["_pri_rank"], -x["_age"]))

    # Round-robin pick across teams to ensure spread
    teams = sorted(by_team.keys(), key=lambda fg: len(by_team[fg]), reverse=True)
    team_indices = {fg: 0 for fg in teams}
    selected = []
    selected_ids = set()

    rounds = 0
    while len(selected) < TOP_N and rounds < 200:
        picked_any = False
        for fg in teams:
            if len(selected) >= TOP_N:
                break
            idx = team_indices[fg]
            while idx < len(by_team[fg]):
                ticket = by_team[fg][idx]
                idx += 1
                if ticket["TicketID"] not in selected_ids:
                    selected.append(ticket)
                    selected_ids.add(ticket["TicketID"])
                    team_indices[fg] = idx
                    picked_any = True
                    break
            team_indices[fg] = idx
        if not picked_any:
            break
        rounds += 1

    # Final sort: priority first, then age
    selected.sort(key=lambda x: (x["_pri_rank"], -x["_age"]))

    # Print summary
    print("\n" + "=" * 130)
    print("TOP %d DEFECTS FOR WORKSHOP" % len(selected))
    print("=" * 130)

    # Team distribution
    team_counts = defaultdict(int)
    pri_counts = defaultdict(int)
    for t in selected:
        team_counts[(t.get("FGroup") or "Unknown").strip()] += 1
        pri_counts[t.get("PriorityID", "?")] += 1

    print("\nTeam Distribution:")
    for fg, cnt in sorted(team_counts.items(), key=lambda x: -x[1]):
        print("  %-25s : %d tickets" % (fg, cnt))

    print("\nPriority Distribution:")
    for p, cnt in sorted(pri_counts.items(), key=lambda x: priority_rank(x[0])):
        print("  %-10s : %d tickets" % (p, cnt))

    print("\n" + "-" * 130)
    print("%-4s %-10s %-10s %-14s %-22s %-7s %5s  %s" % (
        "#", "TicketID", "Priority", "Step", "FGroup", "Occur", "Age", "Title"))
    print("-" * 130)
    for i, t in enumerate(selected, 1):
        print("%-4d %-10s %-10s %-14s %-22s %-7s %4dd  %s" % (
            i, t["TicketID"], t.get("PriorityID", "?"), t["TicketStepID"],
            (t.get("FGroup") or "?")[:22], (t.get("Occurance") or "?")[:7],
            t["_age"], t["Title"][:70]))

    # Excel export
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

        wb = Workbook()
        ws = wb.active
        ws.title = "Workshop Top 50"

        hf = Font(bold=True, color="FFFFFF", size=10)
        hfill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        a1_fill = PatternFill(start_color="FF6B6B", end_color="FF6B6B", fill_type="solid")
        top_fill = PatternFill(start_color="FF9999", end_color="FF9999", fill_type="solid")
        b2_fill = PatternFill(start_color="FFB347", end_color="FFB347", fill_type="solid")
        c3_fill = PatternFill(start_color="77DD77", end_color="77DD77", fill_type="solid")
        tb = Border(left=Side("thin"), right=Side("thin"), top=Side("thin"), bottom=Side("thin"))
        wrap = Alignment(wrap_text=True, vertical="top")

        headers = ["#", "Ticket ID", "Priority", "Step", "FGroup (Team)", "Occurrence",
                   "Age (days)", "System", "Component", "Title", "Workshop Notes"]
        for c, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=c, value=h)
            cell.font = hf
            cell.fill = hfill
            cell.border = tb
            cell.alignment = Alignment(horizontal="center", wrap_text=True, vertical="center")

        for i, t in enumerate(selected, 1):
            vals = [
                i, t["TicketID"], t.get("PriorityID", "?"), t["TicketStepID"],
                (t.get("FGroup") or "?"), (t.get("Occurance") or "?"),
                t["_age"], (t.get("System") or ""), (t.get("Component") or ""),
                t["Title"], ""
            ]
            for c, v in enumerate(vals, 1):
                cell = ws.cell(row=i + 1, column=c, value=v)
                cell.border = tb
                cell.alignment = wrap

            # Color code by priority
            pri = t.get("PriorityID", "")
            pri_cell = ws.cell(row=i + 1, column=3)
            if pri == "A(1)":
                pri_cell.fill = a1_fill
            elif pri == "top":
                pri_cell.fill = top_fill
            elif pri == "B(2)":
                pri_cell.fill = b2_fill
            elif pri == "C(3)":
                pri_cell.fill = c3_fill

        widths = {"A": 4, "B": 12, "C": 10, "D": 14, "E": 22, "F": 12,
                  "G": 10, "H": 18, "I": 20, "J": 70, "K": 30}
        for col, w in widths.items():
            ws.column_dimensions[col].width = w

        # Team summary sheet
        ws2 = wb.create_sheet("Team Summary")
        ws2.cell(row=1, column=1, value="FGroup (Team)").font = hf
        ws2.cell(row=1, column=1).fill = hfill
        ws2.cell(row=1, column=1).border = tb
        ws2.cell(row=1, column=2, value="Tickets in Workshop").font = hf
        ws2.cell(row=1, column=2).fill = hfill
        ws2.cell(row=1, column=2).border = tb
        for r, (fg, cnt) in enumerate(sorted(team_counts.items(), key=lambda x: -x[1]), 2):
            ws2.cell(row=r, column=1, value=fg).border = tb
            ws2.cell(row=r, column=2, value=cnt).border = tb
        ws2.column_dimensions["A"].width = 25
        ws2.column_dimensions["B"].width = 20

        out = os.path.join(_script_dir, "..", "docs", "output", "workshop_top50_defects.xlsx")
        os.makedirs(os.path.dirname(out), exist_ok=True)
        wb.save(out)
        print("\nExcel saved to:", os.path.abspath(out))
    except ImportError:
        print("\nopenpyxl not installed — skipping Excel export")


if __name__ == "__main__":
    main()
