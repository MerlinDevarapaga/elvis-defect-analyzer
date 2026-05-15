"""
DA2.8 May End Bug Zero filter — replicated from Elvis UI.

Filter logic:
  (Ext. Reference <= 2  AND  FG_SWRev != 'P8_YTB_NA')
  AND (
      PriorityID IN ('A(1)', 'top')
      OR (PriorityID = 'B(2)' AND Occurance != 'Once')
      OR (PriorityID = 'C(3)' AND Occurance != 'Once')
  )
"""
import os, io, sys
from datetime import datetime
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8", errors="replace")
from dotenv import load_dotenv
import mysql.connector

load_dotenv()

conn = mysql.connector.connect(
    host=os.getenv("ELVIS_DB_HOST"),
    user=os.getenv("ELVIS_DB_USER"),
    password=os.getenv("ELVIS_DB_PASSWORD"),
    database=os.getenv("ELVIS_DB_NAME"),
    port=int(os.getenv("ELVIS_DB_PORT", 3306)),
    connection_timeout=15,
)
cursor = conn.cursor(dictionary=True)

QUERY = """
SELECT `TicketID`, `Title`, `TicketStepID`, `PriorityID`, `Occurance`,
       `FGroup`, `FG_SWRev`, `ReferenceNumber`, `Rejected`, `RejectReason`,
       `System`, `Component`, `Owner`, `EnterDateTime`, `LastChangeDateTime`,
       `PlannedFixedDate`, `PlannedFixedVersion`
FROM tbl_ElvisSR
WHERE `ProjectID` = 'MSIL_DA2.8'
  AND `IsDeleted` = 'N'
  AND (`ReferenceNumber` IS NULL OR `ReferenceNumber` <= 2)
  AND (`FG_SWRev` IS NULL OR `FG_SWRev` != 'P8_YTB_NA')
  AND (
      `PriorityID` IN ('A(1)', 'top')
      OR (`PriorityID` = 'B(2)' AND `Occurance` != 'Once')
      OR (`PriorityID` = 'C(3)' AND `Occurance` != 'Once')
  )
ORDER BY `PriorityID`, `TicketStepID`, `TicketID`
"""

print("Querying Elvis DB with Bug Zero filter...")
cursor.execute(QUERY)
rows = cursor.fetchall()
print(f"Total tickets matching filter: {len(rows)}")

# Summary by status
print("\n=== By Status ===")
status_counts = {}
for r in rows:
    s = r["TicketStepID"]
    status_counts[s] = status_counts.get(s, 0) + 1
for s, c in sorted(status_counts.items(), key=lambda x: -x[1]):
    print(f"  {s:<20} {c}")

# Summary by priority
print("\n=== By Priority ===")
pri_counts = {}
for r in rows:
    p = r["PriorityID"]
    pri_counts[p] = pri_counts.get(p, 0) + 1
for p, c in sorted(pri_counts.items()):
    print(f"  {p:<10} {c}")

# Summary by rejected
print("\n=== By Rejected ===")
rej_counts = {}
for r in rows:
    rj = r["Rejected"]
    rej_counts[rj] = rej_counts.get(rj, 0) + 1
for rj, c in sorted(rej_counts.items()):
    print(f"  {rj:<5} {c}")

cursor.close()
conn.close()

# Export to Excel
try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
except ImportError:
    import subprocess
    subprocess.check_call([sys.executable, "-m", "pip", "install", "openpyxl", "-q"])
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment

wb = openpyxl.Workbook()
ws = wb.active
ws.title = "May End Bug Zero"

headers = ["TicketID", "Title", "Status", "Priority", "Occurrence",
           "FGroup", "FG_SWRev", "Ext. Reference", "Rejected", "Reject Reason",
           "System", "Component", "Owner", "Entered", "Last Changed",
           "Planned Fix Date", "Planned Fix Version"]
header_font = Font(bold=True, color="FFFFFF", size=10)
header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")

for col, h in enumerate(headers, 1):
    cell = ws.cell(row=1, column=col, value=h)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = Alignment(horizontal="center")

col_map = ["TicketID", "Title", "TicketStepID", "PriorityID", "Occurance",
           "FGroup", "FG_SWRev", "ReferenceNumber", "Rejected", "RejectReason",
           "System", "Component", "Owner", "EnterDateTime", "LastChangeDateTime",
           "PlannedFixedDate", "PlannedFixedVersion"]

# Color coding for priority
pri_colors = {
    "top": "FF0000",
    "A(1)": "FF6B6B",
    "B(2)": "FFA500",
    "C(3)": "FFD700",
}

for row_idx, r in enumerate(rows, 2):
    for col_idx, key in enumerate(col_map, 1):
        val = r.get(key, "")
        if val and hasattr(val, "strftime"):
            val = val.strftime("%Y-%m-%d")
        ws.cell(row=row_idx, column=col_idx, value=val or "")

    # Color the priority cell
    pri = r.get("PriorityID", "")
    if pri in pri_colors:
        ws.cell(row=row_idx, column=4).fill = PatternFill(
            start_color=pri_colors[pri], end_color=pri_colors[pri], fill_type="solid"
        )
        if pri in ("top", "A(1)"):
            ws.cell(row=row_idx, column=4).font = Font(color="FFFFFF", bold=True)

# Auto-fit columns
for col in ws.columns:
    max_len = 0
    col_letter = col[0].column_letter
    for cell in col:
        if cell.value:
            max_len = max(max_len, min(len(str(cell.value)), 50))
    ws.column_dimensions[col_letter].width = max_len + 2

# Add summary sheet
ws2 = wb.create_sheet("Summary")
ws2.cell(row=1, column=1, value="Bug Zero Filter Summary").font = Font(bold=True, size=14)
ws2.cell(row=2, column=1, value=f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')}")
ws2.cell(row=3, column=1, value=f"Total Tickets: {len(rows)}")

ws2.cell(row=5, column=1, value="Status").font = Font(bold=True)
ws2.cell(row=5, column=2, value="Count").font = Font(bold=True)
for i, (s, c) in enumerate(sorted(status_counts.items(), key=lambda x: -x[1]), 6):
    ws2.cell(row=i, column=1, value=s)
    ws2.cell(row=i, column=2, value=c)

row_start = i + 2
ws2.cell(row=row_start, column=1, value="Priority").font = Font(bold=True)
ws2.cell(row=row_start, column=2, value="Count").font = Font(bold=True)
for j, (p, c) in enumerate(sorted(pri_counts.items()), row_start + 1):
    ws2.cell(row=j, column=1, value=p)
    ws2.cell(row=j, column=2, value=c)

downloads = os.path.join(os.path.expanduser("~"), "Downloads")
timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
filepath = os.path.join(downloads, f"DA28_May_End_Bug_Zero_{timestamp}.xlsx")
wb.save(filepath)
print(f"\nExcel exported to: {filepath}")
