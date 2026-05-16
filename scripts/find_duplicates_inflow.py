"""
Find potential duplicates between last N days inflow and existing pre-integrating tickets.
Uses title similarity (SequenceMatcher) to flag candidates.

Usage:
    python scripts/find_duplicates_inflow.py
    python scripts/find_duplicates_inflow.py --excel   (exports to docs/output/)
"""
import os
import sys
import io
import re
import argparse
from datetime import datetime, timedelta
from difflib import SequenceMatcher
from collections import defaultdict

sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8", errors="replace")

from dotenv import load_dotenv
import mysql.connector

# Load .env
_script_dir = os.path.dirname(os.path.abspath(__file__))
for _env in [os.path.join(_script_dir, "..", ".env"), os.path.join(_script_dir, ".env")]:
    if os.path.exists(_env):
        load_dotenv(_env)
        break

DAYS_BACK = 4
SIMILARITY_THRESHOLD = 0.55
EXCEL_THRESHOLD = 0.70  # Higher threshold for Excel export
PROJECT_ID = "MSIL_DA2.8"
PRE_INTEGRATING_STEPS = ("Categorizing", "Processing", "Reproduction")


def get_connection():
    return mysql.connector.connect(
        host=os.getenv("ELVIS_DB_HOST"),
        user=os.getenv("ELVIS_DB_USER"),
        password=os.getenv("ELVIS_DB_PASSWORD"),
        database=os.getenv("ELVIS_DB_NAME"),
        port=int(os.getenv("ELVIS_DB_PORT", 3306)),
    )


def normalize(title):
    t = title.lower().strip()
    t = re.sub(r"[^a-z0-9 ]", " ", t)
    t = re.sub(r"\s+", " ", t)
    return t


def fetch_data():
    conn = get_connection()
    cursor = conn.cursor(dictionary=True)

    cutoff = (datetime.now() - timedelta(days=DAYS_BACK)).strftime("%Y-%m-%d")

    q_inflow = """
        SELECT TicketID, Title, EnterDateTime, TicketStepID, FGroup, PriorityID, Occurance
        FROM tbl_ElvisSR
        WHERE ProjectID = %s AND EnterDateTime >= %s AND IsDeleted = 'N'
        ORDER BY EnterDateTime DESC
    """
    cursor.execute(q_inflow, [PROJECT_ID, cutoff])
    inflow = cursor.fetchall()

    placeholders = ", ".join(["%s"] * len(PRE_INTEGRATING_STEPS))
    q_existing = f"""
        SELECT TicketID, Title, EnterDateTime, TicketStepID, FGroup, PriorityID, Occurance
        FROM tbl_ElvisSR
        WHERE ProjectID = %s AND TicketStepID IN ({placeholders}) AND IsDeleted = 'N'
        ORDER BY FGroup, EnterDateTime
    """
    cursor.execute(q_existing, [PROJECT_ID] + list(PRE_INTEGRATING_STEPS))
    existing = cursor.fetchall()

    cursor.close()
    conn.close()
    return inflow, existing, cutoff


def find_matches(inflow, existing, threshold):
    matches = []
    for inf in inflow:
        inf_norm = normalize(inf["Title"])
        inf_id = inf["TicketID"]
        for ex in existing:
            if ex["TicketID"] == inf_id:
                continue
            ratio = SequenceMatcher(None, inf_norm, normalize(ex["Title"])).ratio()
            if ratio >= threshold:
                matches.append({
                    "similarity": ratio,
                    "inflow_id": inf_id,
                    "inflow_title": inf["Title"],
                    "inflow_date": str(inf["EnterDateTime"])[:10],
                    "inflow_step": inf["TicketStepID"],
                    "inflow_fgroup": inf["FGroup"],
                    "inflow_priority": inf["PriorityID"],
                    "existing_id": ex["TicketID"],
                    "existing_title": ex["Title"],
                    "existing_date": str(ex["EnterDateTime"])[:10],
                    "existing_step": ex["TicketStepID"],
                    "existing_fgroup": ex["FGroup"],
                    "existing_priority": ex["PriorityID"],
                })

    matches.sort(key=lambda x: x["similarity"], reverse=True)

    seen = set()
    unique = []
    for m in matches:
        key = (m["inflow_id"], m["existing_id"])
        if key not in seen:
            seen.add(key)
            unique.append(m)
    return unique


def print_results(inflow, existing, matches, cutoff):
    print("Looking for inflow since:", cutoff)
    print("Inflow tickets (last %d days): %d" % (DAYS_BACK, len(inflow)))
    print("Existing pre-integrating tickets:", len(existing))
    print()

    for r in inflow:
        print("  %s | %s | %-14s | %-20s | %s | %s" % (
            r["TicketID"], str(r["EnterDateTime"])[:10], r["TicketStepID"],
            r["FGroup"], r["PriorityID"], r["Title"][:90]))

    print()
    print("=" * 120)
    print("POTENTIAL DUPLICATES FOUND:", len(matches))
    print("=" * 120)

    if not matches:
        print("No potential duplicates detected.")
        return

    grouped = defaultdict(list)
    for m in matches:
        grouped[m["inflow_id"]].append(m)

    for inf_id, group in grouped.items():
        first = group[0]
        print()
        print("--- Inflow Ticket: %s (%s) [%s] [%s] ---" % (
            inf_id, first["inflow_date"], first["inflow_step"], first["inflow_fgroup"]))
        print("    Title: %s" % first["inflow_title"][:120])
        print("    Potential matches:")
        for m in group:
            sim_pct = "%.0f%%" % (m["similarity"] * 100)
            print("      -> %s (%s) [%s] [%s] Sim=%s" % (
                m["existing_id"], m["existing_date"], m["existing_step"],
                m["existing_fgroup"], sim_pct))
            print("         Title: %s" % m["existing_title"][:120])


def export_excel(matches, cutoff):
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

    wb = Workbook()
    ws = wb.active
    ws.title = "Potential Duplicates (>=70%)"

    hf = Font(bold=True, color="FFFFFF", size=10)
    hfill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    red_fill = PatternFill(start_color="FF6B6B", end_color="FF6B6B", fill_type="solid")
    orange_fill = PatternFill(start_color="FFB347", end_color="FFB347", fill_type="solid")
    green_fill = PatternFill(start_color="77DD77", end_color="77DD77", fill_type="solid")
    tb = Border(left=Side("thin"), right=Side("thin"), top=Side("thin"), bottom=Side("thin"))
    wrap = Alignment(wrap_text=True, vertical="top")

    headers = [
        "#", "Sim%",
        "Inflow TicketID", "Inflow Date", "Inflow Step", "Inflow FGroup", "Inflow Priority", "Inflow Title",
        "Existing TicketID", "Existing Date", "Existing Step", "Existing FGroup", "Existing Priority", "Existing Title",
    ]
    for c, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=c, value=h)
        cell.font = hf
        cell.fill = hfill
        cell.border = tb
        cell.alignment = Alignment(horizontal="center", wrap_text=True, vertical="center")

    # Filter to >= 70% for Excel
    excel_matches = [m for m in matches if m["similarity"] >= EXCEL_THRESHOLD]

    for i, m in enumerate(excel_matches, 1):
        sim_pct = round(m["similarity"] * 100)
        vals = [
            i, "%d%%" % sim_pct,
            m["inflow_id"], m["inflow_date"], m["inflow_step"], m["inflow_fgroup"], m["inflow_priority"], m["inflow_title"],
            m["existing_id"], m["existing_date"], m["existing_step"], m["existing_fgroup"], m["existing_priority"], m["existing_title"],
        ]
        for c, v in enumerate(vals, 1):
            cell = ws.cell(row=i + 1, column=c, value=v)
            cell.border = tb
            cell.alignment = wrap
        sim_cell = ws.cell(row=i + 1, column=2)
        if sim_pct >= 95:
            sim_cell.fill = red_fill
        elif sim_pct >= 80:
            sim_cell.fill = orange_fill
        else:
            sim_cell.fill = green_fill

    widths = {"A": 4, "B": 6, "C": 14, "D": 12, "E": 14, "F": 18, "G": 10, "H": 60,
              "I": 14, "J": 12, "K": 14, "L": 18, "M": 10, "N": 60}
    for col, w in widths.items():
        ws.column_dimensions[col].width = w

    out = os.path.join(_script_dir, "..", "docs", "output", "inflow_duplicate_analysis.xlsx")
    os.makedirs(os.path.dirname(out), exist_ok=True)
    wb.save(out)
    print()
    print("Excel saved to:", os.path.abspath(out))
    print("High-confidence matches (>=70%%):", len(excel_matches))


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--excel", action="store_true", help="Export to Excel")
    args = parser.parse_args()

    inflow, existing, cutoff = fetch_data()
    matches = find_matches(inflow, existing, SIMILARITY_THRESHOLD)
    print_results(inflow, existing, matches, cutoff)

    if args.excel:
        export_excel(matches, cutoff)


if __name__ == "__main__":
    main()
