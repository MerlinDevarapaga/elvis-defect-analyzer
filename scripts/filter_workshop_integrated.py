"""
Filter Workshop Excel — Updates workshop with new SOA 1-Urgent/2-Very High
tickets from DB, adds them to respective FGroup sheets, and strikes through
any tickets >= Integrating. Keeps 'Reference Sheet' untouched.
"""
import os
import sys
import io
from datetime import datetime, date
from dotenv import load_dotenv
import mysql.connector
import openpyxl
from openpyxl.styles import Font, Alignment
from openpyxl.utils import get_column_letter

sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8", errors="replace")

_script_dir = os.path.dirname(os.path.abspath(__file__))
for _env in [os.path.join(_script_dir, "..", ".env"), os.path.join(_script_dir, ".env")]:
    if os.path.exists(_env):
        load_dotenv(_env)
        break

PROJECT_ID = "MSIL_DA2.8"

INTEGRATED_AND_BEYOND = (
    "Integrating", "Integrated", "Verifying", "Verified",
    "Concluding", "Concluded", "Closed",
)

PRE_INTEGRATING_STEPS = ("Categorizing", "Processing", "Reproduction")

VALID_SOA = ("1-Urgent", "2-Very High")

SKIP_SHEETS = {
    "Reference Sheet", "Repro", "Detail1", "Pivot",
    "Platfrom Dependency", "-1065090410_20260414062004",
}

# FGroup → sheet name mapping
FGROUP_TO_SHEET = {
    "Media": "Media",
    "Bluetooth": "BT",
    "Projection": "Projection",
    "Audio": "Audio",
    "IOC": "IOC",
    "Camera": "Camera",
    "WiFi": "Wifi",
    "Systems - Core": "Systems core",
    "Systems - Infra": "Sys infra",
    "Systems - SWU Software Update": "SWUP",
    "USB": "USB",
    "SVS": "SVS",
}

# 11-col sheet header order
SHORT_HEADERS = [
    "Requested Priority", "Occurrence", "Functional group", "Ticket ID",
    "Title", "Secondary type", "Ticket step", "Reported on",
    "Fix planned in CW", "Platform - Ticket step", "Sequence of Activity",
]

# DB column → short sheet header mapping
DB_TO_SHORT = {
    "PriorityID": "Requested Priority",
    "Occurance": "Occurrence",
    "FGroup": "Functional group",
    "TicketID": "Ticket ID",
    "Title": "Title",
    "ProblemType": "Secondary type",
    "TicketStepID": "Ticket step",
    "EnterDateTime": "Reported on",
    "FixPlannedInCW": "Fix planned in CW",
    "PF_TicketStep": "Platform - Ticket step",
    "SequenceOfActivity": "Sequence of Activity",
}

# 41-col sheet header order (standard, excluding Projection's extra leading None col)
FULL_HEADERS = [
    "Ext. Reference", "Release Gate Issue", "Issue Type", "Stability Category",
    "Free text 7", "Milestone", "Requested Priority", "Occurrence",
    "Functional group", "Ticket ID", "Title", "Secondary type", "Ticket step",
    "Reported on", "Detected by", "Reported by", "System SW-Rev.", "System HW-Rev.",
    "Time since entry", "Reference", "In conclusion since", "Fix planned at",
    "Fix planned in CW", "In integration since", "In verification since",
    "Platform - Ticket step", "Free text 8", "Covering note", "Exchange note",
    "Sent exchange notes", "Problem Description", "In repro since", "Ticket owner",
    "Rejected", "Reject cause", "Note", "Functional group SW-Rev.",
    "Implemented in version", "Fix planned in version", "Free lookup 6",
    "Sequence of Activity",
]

DB_TO_FULL = {
    "ReferenceNumber": "Ext. Reference",
    "StabilityCategory": "Stability Category",
    "FreeField_07": "Free text 7",
    "Milestone": "Milestone",
    "PriorityID": "Requested Priority",
    "Occurance": "Occurrence",
    "FGroup": "Functional group",
    "TicketID": "Ticket ID",
    "Title": "Title",
    "ProblemType": "Secondary type",
    "TicketStepID": "Ticket step",
    "EnterDateTime": "Reported on",
    "DetectedBy": "Detected by",
    "EnteredByUGrp": "Reported by",
    "Sys_SWRev": "System SW-Rev.",
    "Sys_HWRev": "System HW-Rev.",
    "ReferenceNumber": "Reference",
    "FirstConclDateTime": "In conclusion since",
    "PlannedFixedDate": "Fix planned at",
    "FixPlannedInCW": "Fix planned in CW",
    "FirstIntegrDateTime": "In integration since",
    "FirstVeriDateTime": "In verification since",
    "PF_TicketStep": "Platform - Ticket step",
    "FreeField_08": "Free text 8",
    "CoveringNote": "Covering note",
    "ExchangeNote": "Exchange note",
    "ExchangeNoteSent": "Sent exchange notes",
    "ProblemDescription": "Problem Description",
    "FirstReproDateTime": "In repro since",
    "Owner": "Ticket owner",
    "Rejected": "Rejected",
    "RejectReason": "Reject cause",
    "FG_SWRev": "Functional group SW-Rev.",
    "IntegrateVersion": "Implemented in version",
    "PlannedFixedVersion": "Fix planned in version",
    "FreeField_06": "Free lookup 6",
    "SequenceOfActivity": "Sequence of Activity",
}

INPUT_FILE = r"C:\Users\mdevarapaga\Downloads\16-Apr-Defect Workshop (1).xlsx"
OUTPUT_FILE = r"C:\Users\mdevarapaga\Downloads\17-Apr-Defect Workshop.xlsx"

# DB columns needed for new rows
DB_COLUMNS_SHORT = list(DB_TO_SHORT.keys())
DB_COLUMNS_FULL = list(DB_TO_FULL.keys())
ALL_DB_COLUMNS = list(set(DB_COLUMNS_SHORT + DB_COLUMNS_FULL + ["TicketID", "TicketStepID", "FGroup", "SequenceOfActivity"]))


def get_connection():
    return mysql.connector.connect(
        host=os.getenv("ELVIS_DB_HOST"),
        user=os.getenv("ELVIS_DB_USER"),
        password=os.getenv("ELVIS_DB_PASSWORD"),
        database=os.getenv("ELVIS_DB_NAME"),
        port=int(os.getenv("ELVIS_DB_PORT", 3306)),
    )


def fetch_existing_ticket_status(ticket_ids):
    """Query DB for current TicketStepID of existing ticket IDs."""
    conn = get_connection()
    cursor = conn.cursor(dictionary=True)
    result = {}
    ids = list(ticket_ids)
    for i in range(0, len(ids), 200):
        chunk = ids[i : i + 200]
        placeholders = ", ".join(["%s"] * len(chunk))
        cursor.execute(
            f"SELECT TicketID, TicketStepID, FGroup FROM tbl_ElvisSR WHERE TicketID IN ({placeholders})",
            chunk,
        )
        for row in cursor.fetchall():
            result[row["TicketID"]] = row
    cursor.close()
    conn.close()
    return result


def fetch_soa_tickets():
    """Fetch all pre-integrating SOA 1/2 tickets from DB."""
    conn = get_connection()
    cursor = conn.cursor(dictionary=True)
    col_list = ", ".join([f"`{c}`" for c in ALL_DB_COLUMNS])
    placeholders_steps = ", ".join(["%s"] * len(PRE_INTEGRATING_STEPS))
    placeholders_soa = ", ".join(["%s"] * len(VALID_SOA))
    query = f"""
        SELECT {col_list}
        FROM tbl_ElvisSR
        WHERE ProjectID = %s
          AND TicketStepID IN ({placeholders_steps})
          AND SequenceOfActivity IN ({placeholders_soa})
          AND IsDeleted = 'N'
        ORDER BY FGroup, EnterDateTime
    """
    params = [PROJECT_ID] + list(PRE_INTEGRATING_STEPS) + list(VALID_SOA)
    cursor.execute(query, params)
    rows = cursor.fetchall()
    cursor.close()
    conn.close()
    return rows


def fetch_top_a_non_urgent_tickets():
    """Fetch pre-integrating TOP A (PriorityID=A(1)) tickets NOT already SOA 1-Urgent."""
    conn = get_connection()
    cursor = conn.cursor(dictionary=True)
    col_list = ", ".join([f"`{c}`" for c in ALL_DB_COLUMNS])
    placeholders_steps = ", ".join(["%s"] * len(PRE_INTEGRATING_STEPS))
    query = f"""
        SELECT {col_list}
        FROM tbl_ElvisSR
        WHERE ProjectID = %s
          AND TicketStepID IN ({placeholders_steps})
          AND PriorityID = %s
          AND (SequenceOfActivity IS NULL OR SequenceOfActivity NOT IN ('1-Urgent'))
          AND IsDeleted = 'N'
        ORDER BY FGroup, EnterDateTime
    """
    params = [PROJECT_ID] + list(PRE_INTEGRATING_STEPS) + ["A(1)"]
    cursor.execute(query, params)
    rows = cursor.fetchall()
    cursor.close()
    conn.close()
    return rows


def apply_strikethrough(ws, row_num):
    for col_num in range(1, ws.max_column + 1):
        cell = ws.cell(row=row_num, column=col_num)
        orig = cell.font
        cell.font = Font(
            name=orig.name, size=orig.size, bold=orig.bold,
            italic=orig.italic, color=orig.color, strikethrough=True,
        )


def add_row_to_sheet(ws, headers, db_row, db_map):
    """Append a new row to the sheet using DB data mapped to headers."""
    new_row = []
    for h in headers:
        # Find the DB column for this header
        db_col = None
        for dc, hdr in db_map.items():
            if hdr == h:
                db_col = dc
                break
        if db_col and db_col in db_row:
            new_row.append(db_row[db_col])
        else:
            new_row.append(None)
    ws.append(new_row)
    return ws.max_row


def main():
    print(f"Loading {INPUT_FILE} ...")
    wb = openpyxl.load_workbook(INPUT_FILE)
    print(f"Sheets: {wb.sheetnames}")

    # --- Step 1: Collect existing ticket IDs per sheet ---
    all_ticket_ids = set()
    sheet_ticket_col = {}  # sname → 0-based col idx for Ticket ID
    sheet_headers = {}     # sname → header list
    for sname in wb.sheetnames:
        if sname in SKIP_SHEETS:
            continue
        ws = wb[sname]
        headers = [cell.value for cell in ws[1]]
        sheet_headers[sname] = headers
        if "Ticket ID" in headers:
            col_idx = headers.index("Ticket ID")
            sheet_ticket_col[sname] = col_idx
            for row in ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=True):
                tid = row[col_idx]
                if tid and isinstance(tid, (int, float)):
                    all_ticket_ids.add(int(tid))

    print(f"Existing ticket IDs in sheets: {len(all_ticket_ids)}")

    # --- Step 2: Query DB for current status of existing tickets ---
    print("Querying DB for existing ticket status ...")
    db_status = fetch_existing_ticket_status(all_ticket_ids)

    integrated_ids = set()
    for tid, info in db_status.items():
        if info.get("TicketStepID", "") in INTEGRATED_AND_BEYOND:
            integrated_ids.add(tid)

    print(f"Tickets now >= Integrated: {len(integrated_ids)}")
    for tid in sorted(integrated_ids):
        info = db_status[tid]
        print(f"  {tid}: {info['TicketStepID']} ({info.get('FGroup', '?')})")

    # --- Step 3: Fetch new SOA 1/2 pre-integrating tickets from DB ---
    print("\nQuerying DB for SOA 1-Urgent / 2-Very High pre-integrating tickets ...")
    soa_tickets = fetch_soa_tickets()
    print(f"Total SOA 1/2 pre-integrating tickets in DB: {len(soa_tickets)}")

    # --- Step 3b: Fetch TOP A (PriorityID=A(1)) tickets not already SOA 1-Urgent ---
    print("\nQuerying DB for TOP A non-urgent pre-integrating tickets ...")
    top_a_tickets = fetch_top_a_non_urgent_tickets()
    print(f"Total TOP A non-urgent pre-integrating tickets in DB: {len(top_a_tickets)}")

    # Merge: SOA tickets + TOP A non-urgent (deduplicate by TicketID)
    seen_ids = {t["TicketID"] for t in soa_tickets}
    merged_tickets = list(soa_tickets)
    for t in top_a_tickets:
        if t["TicketID"] not in seen_ids:
            merged_tickets.append(t)
            seen_ids.add(t["TicketID"])
    print(f"Merged total (SOA + TOP A non-urgent): {len(merged_tickets)}")

    # Filter to only new tickets not already in any sheet
    new_tickets = [t for t in merged_tickets if t["TicketID"] not in all_ticket_ids]
    print(f"New tickets to add: {len(new_tickets)}")

    # --- Step 4: Add new tickets to respective sheets ---
    added_counts = {}
    for ticket in new_tickets:
        fg = ticket.get("FGroup", "")
        sname = FGROUP_TO_SHEET.get(fg)
        if not sname or sname not in sheet_ticket_col:
            print(f"  Skipping {ticket['TicketID']} - no sheet for FGroup '{fg}'")
            continue

        ws = wb[sname]
        headers = sheet_headers[sname]
        is_short = len(headers) <= 15
        db_map = DB_TO_SHORT if is_short else DB_TO_FULL

        row_num = add_row_to_sheet(ws, headers, ticket, db_map)
        added_counts.setdefault(sname, []).append(ticket["TicketID"])

    print("\n--- New tickets added per sheet ---")
    for sname, tids in added_counts.items():
        print(f"  {sname}: {len(tids)} added — {tids}")
    if not added_counts:
        print("  (none)")

    # --- Step 5: Strikethrough >= Integrated rows (all data sheets) ---
    # Re-collect all ticket IDs (including newly added)
    strike_counts = {}
    for sname in wb.sheetnames:
        if sname in SKIP_SHEETS or sname not in sheet_ticket_col:
            continue
        ws = wb[sname]
        col_idx = sheet_ticket_col[sname]

        struck = 0
        for row_num in range(2, ws.max_row + 1):
            tid = ws.cell(row=row_num, column=col_idx + 1).value
            if not tid or not isinstance(tid, (int, float)):
                continue
            if int(tid) in integrated_ids:
                struck += 1
                apply_strikethrough(ws, row_num)
        strike_counts[sname] = struck

    print("\n--- Rows struck through per sheet ---")
    for sname, count in strike_counts.items():
        if count > 0:
            print(f"  {sname}: {count} rows struck through")
        else:
            print(f"  {sname}: no changes")

    # --- Step 6: Wrap text for all sheets (compact line spacing) ---
    for sname in wb.sheetnames:
        ws = wb[sname]
        for row in ws.iter_rows(min_row=1, max_row=ws.max_row, max_col=ws.max_column):
            for cell in row:
                cell.alignment = Alignment(wrap_text=True, vertical="center", shrink_to_fit=False)
        # Set compact row heights for data rows
        for row_num in range(2, ws.max_row + 1):
            ws.row_dimensions[row_num].height = 15

    # Save
    wb.save(OUTPUT_FILE)
    print(f"\nSaved to: {OUTPUT_FILE}")


if __name__ == "__main__":
    main()
