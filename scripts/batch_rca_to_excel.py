"""
Batch fetch Elvis defects and populate RCA/Countermeasure columns in the Excel template.
Reads from 'Till 17th March' sheet structure, writes results to 'RCA Output' sheet.
"""
import os
import sys
import json
import datetime

# Add the skill scripts folder to path so we can reuse fetch_defect
_script_dir = os.path.dirname(os.path.abspath(__file__))
_skill_scripts = os.path.join(_script_dir, "..", ".github", "skills", "elvis-defect-analyzer", "scripts")
sys.path.insert(0, _skill_scripts)

from fetch_defect import fetch_defect, COLUMN_GROUPS, _is_populated

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side


TICKET_IDS = [
    3727322, 3725174, 3725645, 3723553, 3719292,
    3719329, 3726560, 3726402, 3729547, 3714420,
    3727656, 3724991, 3710131, 3703960, 3726579,
]

EXCEL_PATH = r"C:\Users\mdevarapaga\Downloads\MSIL_DA2.8_Defect_List_6-Apr-2026.xlsx"
OUTPUT_SHEET = "RCA Output"


def extract_rca(defect):
    """Extract root cause from defect fields."""
    if not defect:
        return "Ticket not found in Elvis DB"

    parts = []

    # Primary: Cause field
    cause = (defect.get("Cause") or "").strip()
    if cause and cause.lower() not in ("none", "0", ""):
        parts.append(cause)

    # BugTaxonomy
    taxonomy = (defect.get("BugTaxonomy") or "").strip()
    if taxonomy and taxonomy.lower() not in ("none", "0", ""):
        parts.append(f"[Bug Taxonomy: {taxonomy}]")

    # CauseID
    cause_id = (defect.get("CauseID") or "").strip()
    if cause_id and cause_id.lower() not in ("none", "0", ""):
        parts.append(f"[CauseID: {cause_id}]")

    # RespNote as fallback analysis
    resp_note = (defect.get("RespNote") or "").strip()
    if resp_note and resp_note.lower() not in ("none", "0", "") and not parts:
        # Use first 300 chars of RespNote as RCA hint
        parts.append(resp_note[:300])

    # InternalStatement as fallback
    internal = (defect.get("InternalStatement") or "").strip()
    if internal and internal.lower() not in ("none", "0", "") and not parts:
        parts.append(internal[:300])

    # Result field (verification notes)
    result = (defect.get("Result") or "").strip()
    if result and result.lower() not in ("none", "0", "") and not parts:
        parts.append(f"Analysis: {result[:300]}")

    # ProblemDescription for context if still empty
    if not parts:
        desc = (defect.get("ProblemDescription") or "").strip()
        if desc:
            parts.append(f"RCA pending. Issue: {desc[:200]}")
        else:
            parts.append("RCA not yet available in Elvis")

    return " | ".join(parts)


def extract_countermeasure(defect):
    """Extract countermeasure/fix info from defect fields."""
    if not defect:
        return "N/A"

    parts = []

    # Primary: Measures field
    measures = (defect.get("Measures") or "").strip()
    if measures and measures.lower() not in ("none", "0", ""):
        parts.append(measures)

    # Avoidance
    avoidance = (defect.get("Avoidance") or "").strip()
    if avoidance and avoidance.lower() not in ("none", "0", ""):
        parts.append(f"[Avoidance: {avoidance}]")

    # OfficialStatement
    official = (defect.get("OfficialStatement") or "").strip()
    if official and official.lower() not in ("none", "0", ""):
        if not parts:
            parts.append(official[:300])

    # FixedInVersion
    fixed_ver = (defect.get("FixedInVersion") or "").strip()
    if fixed_ver and fixed_ver.lower() not in ("none", "0", ""):
        parts.append(f"[Fixed in: {fixed_ver}]")

    # PlannedFixedVersion / Date
    planned_ver = str(defect.get("PlannedFixedVersion") or "").strip()
    planned_date = str(defect.get("PlannedFixedDate") or "").strip()
    if planned_ver and planned_ver.lower() not in ("none", "0", "0000-00-00"):
        parts.append(f"[Planned fix: {planned_ver}]")
    if planned_date and planned_date not in ("0000-00-00", "0000-00-00 00:00:00", "", "None"):
        parts.append(f"[Target: {planned_date}]")

    if not parts:
        # Check step/state for status hint
        step = str(defect.get("TicketStepID") or "").strip()
        state = str(defect.get("StateID") or "").strip()
        owner = (defect.get("Owner") or "").strip()
        parts.append(f"Countermeasure pending [Step: {step}, State: {state}, Owner: {owner}]")

    return " | ".join(parts)


def determine_rca_identified(defect):
    """Determine if root cause has been identified (Yes/No)."""
    if not defect:
        return "No"
    cause = (defect.get("Cause") or "").strip()
    taxonomy = (defect.get("BugTaxonomy") or "").strip()
    measures = (defect.get("Measures") or "").strip()
    if (cause and cause.lower() not in ("none", "0", "")) or \
       (taxonomy and taxonomy.lower() not in ("none", "0", "")):
        return "Yes"
    if measures and measures.lower() not in ("none", "0", ""):
        return "Yes"
    return "No"


def determine_fix_in_progress(defect):
    """Determine if fix is in progress based on workflow state."""
    if not defect:
        return "No"
    step = str(defect.get("TicketStepID") or "").strip()
    fixed_ver = (defect.get("FixedInVersion") or "").strip()
    measures = (defect.get("Measures") or "").strip()

    # If already fixed
    if fixed_ver and fixed_ver.lower() not in ("none", "0", ""):
        return "Yes (Fixed)"
    # If measures exist, work is happening
    if measures and measures.lower() not in ("none", "0", ""):
        return "Yes"
    # Processing/Integration steps typically mean active work
    if step in ("3", "4", "5", "6", "7"):
        return "Yes"
    return "No"


def estimate_closure(defect):
    """Estimate when ticket can be closed."""
    if not defect:
        return "N/A"

    # Already closed?
    state = str(defect.get("StateID") or "").strip()
    last_close = str(defect.get("LastCloseDateTime") or "").strip()
    if state == "6" or (last_close and last_close not in ("0000-00-00 00:00:00", "")):
        return f"Closed ({last_close})" if last_close and last_close not in ("0000-00-00 00:00:00", "") else "Closed"

    # Planned fix date
    planned_date = str(defect.get("PlannedFixedDate") or "").strip()
    if planned_date and planned_date not in ("0000-00-00", "0000-00-00 00:00:00", ""):
        return planned_date

    # Fixed version exists but not closed
    fixed_ver = (defect.get("FixedInVersion") or "").strip()
    if fixed_ver and fixed_ver.lower() not in ("none", "0", ""):
        return "Awaiting verification"

    return "TBD"


def main():
    print(f"Fetching {len(TICKET_IDS)} defects from Elvis DB...\n")

    results = []
    for tid in TICKET_IDS:
        print(f"  Fetching {tid}...", end=" ", flush=True)
        try:
            defect = fetch_defect(tid)
            if defect:
                print("OK")
            else:
                print("NOT FOUND")
        except Exception as e:
            print(f"ERROR: {e}")
            defect = None

        # Extract fields matching template
        row = {
            "Requested Priority": (defect.get("PriorityID") or "") if defect else "",
            "Occurrence": (defect.get("Occurance") or "") if defect else "",
            "Platform/Project": (defect.get("Product") or defect.get("SubProject") or "") if defect else "",
            "Functional group": (defect.get("FGroup") or "") if defect else "",
            "Ticket ID": tid,
            "Title": (defect.get("Title") or "") if defect else "",
            "Reported on": str(defect.get("EnterDateTime") or "") if defect else "",
            "Root cause identified\n[Yes/No]": determine_rca_identified(defect),
            " Fix in Progress [Yes/No]": determine_fix_in_progress(defect),
            "What is the root cause?": extract_rca(defect),
            "What is the counter measure?": extract_countermeasure(defect),
            "When this ticket can be closed?": estimate_closure(defect),
        }
        results.append(row)

    print(f"\nFetched all tickets. Writing to Excel...")

    # Read existing Excel (pandas for reading, openpyxl for writing)
    df_out = pd.DataFrame(results)

    # Write to a new sheet in the same workbook using openpyxl
    # We need to handle the pivot table issue by writing directly
    try:
        wb = load_workbook(EXCEL_PATH)
    except Exception:
        # If openpyxl fails on pivot tables, create a new workbook
        print("Note: Could not open existing workbook (pivot table issue). Creating standalone output file.")
        from openpyxl import Workbook
        wb = Workbook()
        ws = wb.active
        ws.title = OUTPUT_SHEET
        output_path = EXCEL_PATH.replace(".xlsx", "_RCA_Output.xlsx")
        _write_sheet(ws, df_out)
        wb.save(output_path)
        print(f"Saved to: {output_path}")
        return

    # Remove existing output sheet if present
    if OUTPUT_SHEET in wb.sheetnames:
        del wb[OUTPUT_SHEET]

    ws = wb.create_sheet(OUTPUT_SHEET)
    _write_sheet(ws, df_out)
    wb.save(EXCEL_PATH)
    print(f"Saved '{OUTPUT_SHEET}' sheet in: {EXCEL_PATH}")


def _write_sheet(ws, df):
    """Write dataframe to worksheet with formatting."""
    headers = list(df.columns)

    # Styles
    header_font = Font(bold=True, size=11, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell_align = Alignment(vertical="top", wrap_text=True)
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    # Write headers
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border

    # Write data
    for row_idx, (_, row) in enumerate(df.iterrows(), 2):
        for col_idx, header in enumerate(headers, 1):
            val = row[header]
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.alignment = cell_align
            cell.border = thin_border

    # Column widths
    col_widths = {
        "Requested Priority": 14,
        "Occurrence": 12,
        "Platform/Project": 16,
        "Functional group": 16,
        "Ticket ID": 12,
        "Title": 50,
        "Reported on": 14,
        "Root cause identified\n[Yes/No]": 14,
        " Fix in Progress [Yes/No]": 14,
        "What is the root cause?": 60,
        "What is the counter measure?": 60,
        "When this ticket can be closed?": 20,
    }
    for col_idx, header in enumerate(headers, 1):
        ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = col_widths.get(header, 15)

    # Freeze top row
    ws.freeze_panes = "A2"

    # Auto-filter
    ws.auto_filter.ref = ws.dimensions

    print(f"  Written {len(df)} rows with headers.")


if __name__ == "__main__":
    main()
