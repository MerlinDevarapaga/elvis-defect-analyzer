"""
Generate AI-inferred RCA and Countermeasures for Elvis defects.
Reads all_defects_dump.json, applies domain knowledge to produce crisp RCA/CMs.
Writes output to Excel matching 'Till 17th March' template.
"""
import json
import sys
import io
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8', errors='replace')

EXCEL_OUT = r"C:\Users\mdevarapaga\Downloads\MSIL_DA2.8_Defect_List_6-Apr-2026_RCA_Output.xlsx"

# ── AI-generated RCA and Countermeasures per ticket ──────────────────────────

RCA_DATA = {
    "3727322": {
        "rca_identified": "No",
        "fix_in_progress": "Yes",
        "root_cause": (
            "Probable RCA: BT discovery asymmetry — HU's Bluetooth stack (likely BlueZ/Fluoride) "
            "is failing to respond to inquiry scan from certain phone models during device discovery, "
            "while the HU itself is broadcasting and visible to phones. "
            "Likely cause: Incompatible BT inquiry mode (standard vs interlaced) or BT controller "
            "firmware not properly handling page scan parameters for specific BT chipsets. "
            "The issue occurring 'Always' with certain devices points to a BT profile compatibility gap."
        ),
        "counter_measure": (
            "1. Collect BT HCI snoop logs from HU during failed discovery.\n"
            "2. Verify inquiry scan type (standard vs interlaced) and page scan interval configuration.\n"
            "3. Check BT firmware version and compare against known compatibility matrix.\n"
            "4. Test with BT controller parameter tuning (inquiry scan window/interval).\n"
            "5. If firmware gap — update BT controller FW; if stack issue — patch Fluoride inquiry handler."
        ),
        "closure": "Post BT log analysis — target R9 (2026-04-17)",
    },
    "3725174": {
        "rca_identified": "No",
        "fix_in_progress": "No",
        "root_cause": (
            "Probable RCA: BT pairing state machine race condition — When a second phone initiates "
            "pairing while HFP/A2DP profiles are active on the first phone, the DCSM (Device Connection "
            "State Manager) fails to register the second device in the paired device database. "
            "The pairing key exchange succeeds at the BT stack level, but the device entry is not "
            "propagated to the HMI paired-device list. Likely a missing callback or event-drop in "
            "DCSM's multi-device connection handler."
        ),
        "counter_measure": (
            "1. Review DCSM bonding callback flow for multi-device pairing scenario.\n"
            "2. Add logging at DCSM layer to trace device-add events after bond completion.\n"
            "3. Verify BT stack bond_state_changed callback reaches DCSM for 2nd device.\n"
            "4. Fix: Ensure DCSM registers paired device in DB regardless of active profile count.\n"
            "5. Validate with 3+ phone pairing scenario."
        ),
        "closure": "Pending DCSM analysis — TBD",
    },
    "3725645": {
        "rca_identified": "Yes",
        "fix_in_progress": "Yes",
        "root_cause": (
            "Confirmed RCA: EVSHAL (Embedded Vision System HAL) memory leak in Qualcomm camera HAL. "
            "Design flaw — mDirectRendering flag never becomes true since importExternalBuffers is "
            "not called in the MSIL use case, causing buffer handles to be allocated but never released. "
            "Continuous memory leak leads to SOC memory exhaustion, causing HU boot failure."
        ),
        "counter_measure": (
            "1. Fix applied: Release buffer handles when mDirectRendering is false (Gerrit 611568, 610972).\n"
            "2. Patch in packages/services/Car and vendor/qcom-proprietary/ais.\n"
            "3. Validated on OND — no memory leak observed.\n"
            "4. Awaiting code review and pre-integration promotion to R9.\n"
            "5. Long-term: Add memory monitoring watchdog for EVSHAL buffer pool."
        ),
        "closure": "Fix validated, awaiting R9 integration (target 2026-04-02)",
    },
    "3723553": {
        "rca_identified": "No",
        "fix_in_progress": "Yes",
        "root_cause": (
            "Probable RCA: SWU (Software Update) installation progress persistence failure. "
            "When HU is rebooted mid-installation, the OTA agent does not checkpoint the installation "
            "progress to persistent storage. On reboot, it reads 0% from the progress file and "
            "restarts the installation from scratch. Likely missing fsync/commit of installation "
            "progress state, or the progress tracker uses volatile memory only."
        ),
        "counter_measure": (
            "1. Review OTA agent's installation progress persistence mechanism.\n"
            "2. Add periodic checkpoint writes (every 5% or per-partition) to persistent storage.\n"
            "3. Implement resume-from-checkpoint logic in OTA agent's reboot recovery path.\n"
            "4. Validate with forced reboot at 30%, 50%, 80% installation progress.\n"
            "5. Ensure AB partition scheme supports partial installation resume."
        ),
        "closure": "Pending SWU team analysis — target 2026-03-27",
    },
    "3719292": {
        "rca_identified": "Yes",
        "fix_in_progress": "Yes",
        "root_cause": (
            "Confirmed RCA: Incorrect/Missing Design — Old inventory in tree.xml. "
            "The SWU activation flow uses tree.xml for component inventory validation. "
            "Stale default inventory entries caused Redbend initialization blocker, "
            "resulting in HU stuck at 'Activation in Progress' screen indefinitely."
        ),
        "counter_measure": (
            "1. Fix: Updated tree.xml with new defaults including MSIL-specific configurations.\n"
            "2. Resolved Redbend initialization blocker.\n"
            "3. Gerrit: androidhub/harman/packages/SoftwareUpdate/+/596428.\n"
            "4. Preventive: Add tree.xml inventory validation in SWU pre-check step.\n"
            "5. Add integration test for USB offline update activation flow."
        ),
        "closure": "Fix submitted — pending verification on R8.1+ build",
    },
    "3719329": {
        "rca_identified": "Yes",
        "fix_in_progress": "Yes",
        "root_cause": (
            "Confirmed RCA: Same root cause as 3719292 — Incorrect/Missing Design. "
            "Old inventory in tree.xml causes Redbend initialization to misbehave, "
            "triggering premature activation popup at 0% installation. "
            "The SWU state machine incorrectly transitions to ACTIVATION_READY state "
            "before installation begins due to stale configuration."
        ),
        "counter_measure": (
            "1. Fix: Same as 3719292 — updated tree.xml with MSIL configurations + Redbend init fix.\n"
            "2. Gerrit: androidhub/harman/packages/SoftwareUpdate/+/596428.\n"
            "3. Preventive: Add state machine guard — block ACTIVATION_READY if install progress < 100%.\n"
            "4. Add assertion/log for unexpected state transitions in SWU flow."
        ),
        "closure": "Fix submitted — pending verification on R8.1+ build",
    },
    "3726560": {
        "rca_identified": "No",
        "fix_in_progress": "No (In Reproduction)",
        "root_cause": (
            "Probable RCA: BT stack crash/restart loop. Failure to pair AND unpair + BT toggling On/Off "
            "with black screen suggests the Bluetooth system service is crashing repeatedly. "
            "Likely cause: Null pointer or resource exhaustion in BT bonding manager when handling "
            "simultaneous pair/unpair requests, causing BT service to restart in a loop. "
            "The black screen indicates the HMI layer loses its BT service connection."
        ),
        "counter_measure": (
            "1. Collect logcat/tombstone during reproduction — check for BT service crashes.\n"
            "2. Review BT bonding manager for concurrent pair/unpair race conditions.\n"
            "3. Add crash recovery handling — prevent BT toggle loop on service restart.\n"
            "4. HMI: Add fallback UI instead of black screen when BT service is unavailable.\n"
            "5. Test with specific devices: Vivo T3X and Moto G85 (Android 15)."
        ),
        "closure": "In reproduction phase — TBD after root cause confirmation",
    },
    "3726402": {
        "rca_identified": "No",
        "fix_in_progress": "Yes",
        "root_cause": (
            "Probable RCA: Camera subsystem initialization causing boot loop. "
            "FGroup is Camera — after battery cycle, the camera HAL or EVS (Extended View System) "
            "initialization may fail, triggering a watchdog restart. Multiple restarts with "
            "welcome logo/music indicate the system is reaching early boot but crashing before "
            "fully initializing. Likely a race condition between camera service startup and "
            "SOC power sequencing after battery disconnect."
        ),
        "counter_measure": (
            "1. Capture boot logs across multiple restart cycles — identify crashing service.\n"
            "2. Check camera HAL init timeout and watchdog configuration.\n"
            "3. Add graceful degradation — skip camera init on repeated failures instead of restarting.\n"
            "4. Review SOC power sequencing timing after battery cycle.\n"
            "5. Planned fix in R9 (target 2026-04-10)."
        ),
        "closure": "Planned fix in R9 — target 2026-04-10",
    },
    "3729547": {
        "rca_identified": "No",
        "fix_in_progress": "Yes (Analysis pending)",
        "root_cause": (
            "Probable RCA: SVS (Surround View System) calibration data or configuration is wiped "
            "during factory reset but not re-initialized on first boot. "
            "After factory reset, the SVS camera feed shows black screen with view icons because "
            "the camera pipeline starts but has no valid calibration/configuration data to render. "
            "The view icons render (HMI layer works) but the camera rendering layer has no valid input."
        ),
        "counter_measure": (
            "1. Verify which SVS data/configs are cleared during factory reset.\n"
            "2. Ensure SVS default calibration is restored as part of factory reset recovery.\n"
            "3. Add SVS self-check: if calibration data missing, re-initialize with defaults.\n"
            "4. Clone to Platform for investigation (per RespNote).\n"
            "5. Test: Factory reset → reverse gear → verify SVS renders correctly."
        ),
        "closure": "Analysis pending (cloned to Platform) — target 2026-03-31",
    },
    "3714420": {
        "rca_identified": "No",
        "fix_in_progress": "No",
        "root_cause": (
            "Probable RCA: HMI keyboard language selection UI does not persist/highlight the "
            "active language state correctly. The keyboard language preference is stored and "
            "functionally applied (keyboard works in Hindi) but the UI selection indicator "
            "does not read the current language setting on the language picker screen. "
            "Likely a missing binding between the language preference value and the "
            "highlight state of the language list item in the HMI framework."
        ),
        "counter_measure": (
            "1. Review keyboard language selection HMI code — check if selected language ID is "
            "bound to the list item highlight state.\n"
            "2. Fix: Read active keyboard language from settings and set highlight on matching list item.\n"
            "3. Low priority (C3) — awaiting MSIL OPL confirmation.\n"
            "4. Target: R9.1."
        ),
        "closure": "C3 priority — deferred to R9.1 (2026-04-17+), awaiting MSIL OPL",
    },
    "3727656": {
        "rca_identified": "Yes",
        "fix_in_progress": "Yes (In Integration)",
        "root_cause": (
            "Confirmed RCA: Incorrect Construction (Bug Taxonomy: 4). "
            "Android Auto projection displays black screen when AA media source is selected "
            "while another projection (CP) is connected but not active. "
            "The AA app fails to handle the media source switch when a competing projection "
            "session exists. Issue was not reproducible consistently — moved to integration "
            "with fix Gerrit 604786 (AA projection app patch)."
        ),
        "counter_measure": (
            "1. Fix in Gerrit: androidhub/harman/projection/apps/androidauto/prj/msil_da2.8/+/604786.\n"
            "2. Handles media source switch correctly when multiple projections are connected.\n"
            "3. Planned for R9 integration (target 2026-04-07).\n"
            "4. Verify with: AA + CP connected → switch to AA media source."
        ),
        "closure": "In integration — target R9 (2026-04-07)",
    },
    "3724991": {
        "rca_identified": "No",
        "fix_in_progress": "No (Validate on latest SW)",
        "root_cause": (
            "Probable RCA: Camera switch disconnect event not propagated to DID 0x1000 diagnostic data. "
            "When SVS camera switch is disconnected from battery negative terminal, the IOC does not "
            "update Position 28 Bit 0 to '0'. Likely cause: The GPIO interrupt or ADC threshold for "
            "camera switch disconnect is not configured correctly, or the DID update handler only "
            "writes on connect events (edge-triggered on rising only, not falling)."
        ),
        "counter_measure": (
            "1. Review IOC camera switch GPIO/ADC detection — verify both connect and disconnect edges trigger DID update.\n"
            "2. Check DID 0x1000 write handler for Position 28 Bit 0 — ensure it handles both states.\n"
            "3. Validate on latest SW (per RespNote).\n"
            "4. If confirmed: Fix IOC diagnostic handler to update DID on disconnect event.\n"
            "5. Target: 2026-04-24."
        ),
        "closure": "Pending validation on latest SW — target 2026-04-24",
    },
    "3710131": {
        "rca_identified": "Yes",
        "fix_in_progress": "Yes",
        "root_cause": (
            "Confirmed RCA: Additional speed-related handling missing in SVS 360° view. "
            "When 360° view opens at vehicle speed >10 km/h, the previous camera frame buffer "
            "is momentarily displayed (splash) before the speed-based view restriction kicks in. "
            "The SVS rendering pipeline renders the stale frame before the speed check logic "
            "blanks/restricts the view. Architecture review confirmed the rendering order needs correction."
        ),
        "counter_measure": (
            "1. Reorder SVS rendering pipeline: perform speed check BEFORE frame render.\n"
            "2. Clear/blank the frame buffer before opening 360° view at any speed.\n"
            "3. Architecture review in progress (per RespNote).\n"
            "4. Planned fix: R8.1 (target 2026-04-24).\n"
            "5. Validate: Open 360° at >10 km/h — no stale frame flash visible."
        ),
        "closure": "Architecture review ongoing — target R8.1 (2026-04-24)",
    },
    "3703960": {
        "rca_identified": "No",
        "fix_in_progress": "Yes (Analysis pending)",
        "root_cause": (
            "Probable RCA: 3D rendering artifact in SVS 360° view — grey patch at the bottom of "
            "the 3D vehicle model. Likely cause: Incorrect UV mapping or texture atlas gap in the "
            "3D car model's underside mesh, or a Z-buffer clipping issue where the ground plane "
            "intersects the vehicle model's lower geometry. Could also be a missing texture tile "
            "in the rendering pipeline's bird's-eye stitching algorithm."
        ),
        "counter_measure": (
            "1. Check 3D vehicle model mesh — inspect bottom vertices for UV mapping gaps.\n"
            "2. Review ground plane Z-offset relative to vehicle model in 360° render config.\n"
            "3. If texture issue — fix model asset; if rendering — adjust Z-clipping plane.\n"
            "4. Planned fix: R8.1 (target 2026-04-24).\n"
            "5. Validate: 360° view from all angles — no grey patches visible."
        ),
        "closure": "Analysis pending — target R8.1 (2026-04-24)",
    },
    "3726579": {
        "rca_identified": "No",
        "fix_in_progress": "No (In Reproduction)",
        "root_cause": (
            "Probable RCA: SOC boot failure after battery cycle — system hangs at black screen. "
            "Similar pattern to 3726402 (multiple restarts after battery cycle). "
            "Core system service crash during early boot prevents HMI from launching. "
            "Intermittent nature ('Sometimes') suggests a timing-dependent race condition in "
            "power management or system services initialization sequence after abrupt power loss. "
            "Could be filesystem corruption or incomplete shutdown state persisted to storage."
        ),
        "counter_measure": (
            "1. Capture serial/UART logs during black screen occurrence to identify stuck service.\n"
            "2. Check for filesystem corruption markers (fsck results) after battery cycle.\n"
            "3. Review power management shutdown sequence — ensure clean state on abrupt power loss.\n"
            "4. Add watchdog-based recovery to force full reboot if black screen exceeds timeout.\n"
            "5. Test: 20 battery cycles — verify no persistent black screen."
        ),
        "closure": "In reproduction phase — TBD after root cause confirmation",
    },
}


def main():
    with open('all_defects_dump.json', 'r', encoding='utf-8') as f:
        defects = json.load(f)

    rows = []
    for tid_str, rca_info in RCA_DATA.items():
        d = defects.get(tid_str, {})
        rows.append({
            "Requested Priority": d.get("PriorityID", ""),
            "Occurrence": d.get("Occurance", ""),
            "Platform/Project": d.get("SubProject", "") or "Intelligent Cockpit Platform",
            "Functional group": d.get("FGroup", ""),
            "Ticket ID": int(tid_str),
            "Title": d.get("Title", ""),
            "Reported on": d.get("EnterDateTime", ""),
            "Root cause identified\n[Yes/No]": rca_info["rca_identified"],
            " Fix in Progress [Yes/No]": rca_info["fix_in_progress"],
            "What is the root cause?": rca_info["root_cause"],
            "What is the counter measure?": rca_info["counter_measure"],
            "When this ticket can be closed?": rca_info["closure"],
        })

    df = pd.DataFrame(rows)

    # Write to Excel with formatting
    wb = Workbook()
    ws = wb.active
    ws.title = "RCA Output"

    headers = list(df.columns)
    header_font = Font(bold=True, size=11, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell_align = Alignment(vertical="top", wrap_text=True)
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )

    # Conditional fill for RCA status
    yes_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    no_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")

    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border

    for row_idx, (_, row) in enumerate(df.iterrows(), 2):
        for col_idx, header in enumerate(headers, 1):
            val = row[header]
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.alignment = cell_align
            cell.border = thin_border
            # Color-code RCA identified column
            if header == "Root cause identified\n[Yes/No]":
                cell.fill = yes_fill if val == "Yes" else no_fill
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    col_widths = {
        "Requested Priority": 14, "Occurrence": 12, "Platform/Project": 22,
        "Functional group": 18, "Ticket ID": 12, "Title": 55,
        "Reported on": 14, "Root cause identified\n[Yes/No]": 14,
        " Fix in Progress [Yes/No]": 16, "What is the root cause?": 70,
        "What is the counter measure?": 70, "When this ticket can be closed?": 28,
    }
    for col_idx, header in enumerate(headers, 1):
        col_letter = ws.cell(row=1, column=col_idx).column_letter
        ws.column_dimensions[col_letter].width = col_widths.get(header, 15)

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    wb.save(EXCEL_OUT)
    print(f"Done! Saved {len(rows)} tickets to: {EXCEL_OUT}")
    print(f"Sheet: 'RCA Output'")
    print()
    for r in rows:
        tid = r["Ticket ID"]
        rca = r["Root cause identified\n[Yes/No]"]
        print(f"  {tid} | RCA: {rca} | {r['Title'][:60]}")


if __name__ == "__main__":
    main()
